# Databricks notebook source
# =============================================================================
# Medical Necessity Analysis v3
#
# Built from the 7/21 meeting action items:
#   1. EXAMPLE TRIPS per clinical bucket for Friday's review with Dave
#   2. Map clinical questions to CMS criteria (not just internal criteria)
#   3. Risk tiers / payment-likelihood framework by bucket
#   4. Fix the parser error that inflated the questionnaire results
#   5. Scope filter - exclude rideshare, air, and emergent from the denominator
#   6. Draft minimum data-entry standards for Rev Cycle
#
# READ-ONLY. Temp views only, no tables created.
#
# Source: `prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster
#         (from prod.silver_transbroker.triprequest + tripleg)
# =============================================================================


# COMMAND ----------

# Run once, then comment out:
# %pip install openpyxl
# dbutils.library.restartPython()


# COMMAND ----------

# -----------------------------------------------------------------------------
# 0. CONFIG
# -----------------------------------------------------------------------------

TRIPMASTER  = "`prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster"
OUTPUT_DIR  = "/Workspace/Users/josh.smitherman@gmr.net/med_nec/data"
OUTPUT_XLSX = f"{OUTPUT_DIR}/med_nec_analysis_v3.xlsx"

# How many example trips to pull per bucket for the Dave review.
EXAMPLES_PER_BUCKET = 25

import os, json, re
from pyspark.sql import functions as F, types as T

spark.sql("USE CATALOG prod")
os.makedirs(OUTPUT_DIR, exist_ok=True)
print("output ->", OUTPUT_XLSX)


# COMMAND ----------

# =============================================================================
# 1. SCOPE FILTER
# The raw table mixes in transport types that need no medical-necessity
# documentation. Including them overstates the problem (rideshare alone was
# 9,292 of the 24,303 "nothing typed" orders).
# =============================================================================

# Verify the actual code values before trusting the filter.
print("Level-of-service values present:")
display(spark.sql(f"""
    SELECT LevelOfService, LevelOfServiceDescription, ServiceType, count(*) AS orders
    FROM {TRIPMASTER}
    GROUP BY 1,2,3
    ORDER BY orders DESC
"""))


# COMMAND ----------

# Out-of-scope patterns. Adjust after reviewing the cell above.
OUT_OF_SCOPE = """
    (
        upper(coalesce(ServiceType,''))               IN ('TAXI','ROTOR','EMG')
     OR upper(coalesce(LevelOfService,''))            IN ('TAXI','ROTOR','ALS-EMG','EMG',
                                                          'CCT-EMG','A1','A2','B1','B2')
     OR upper(coalesce(LevelOfServiceDescription,'')) LIKE '%EMERGEN%'
     OR upper(coalesce(LevelOfServiceDescription,'')) IN ('RIDESHARE','LYFT','ROTOR',
                                                          'FIXED WING QUOTE','TEAM')
    )
"""

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW scoped AS
SELECT *,
       CASE WHEN {OUT_OF_SCOPE} THEN 'out_of_scope' ELSE 'in_scope' END AS scope_flag
FROM {TRIPMASTER}
""")

print("Scope split:")
display(spark.sql("""
    SELECT scope_flag, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM scoped GROUP BY scope_flag ORDER BY orders DESC
"""))

print("What got excluded (verify this list is right):")
display(spark.sql("""
    SELECT LevelOfServiceDescription, ServiceType, count(*) AS orders
    FROM scoped WHERE scope_flag='out_of_scope'
    GROUP BY 1,2 ORDER BY orders DESC
"""))


# COMMAND ----------

# =============================================================================
# 2. PARSER FIX
#
# The v2 error: the parser took the FIRST key that looked like an answer. In
# this payload that is a form property (rendering/visibility flag), which is
# almost always the literal "True" - it produced "True" 4,062,545 times out of
# 4,117,796 answers.
#
# Fix: collect ALL candidate answers in an object and rank them.
#   1st choice - a real string value (not true/false/null)
#   2nd choice - a number
#   last resort - a boolean
# Also return WHICH key supplied the answer, so we can verify.
# =============================================================================

QUESTION_HINTS = ("question", "prompt", "label", "text", "title", "name", "caption")
ANSWER_HINTS   = ("answer", "value", "response", "selected", "checked", "result",
                  "chosen", "input", "reply", "entry")

BOOLEANISH = {"true", "false", "1", "0", "yes", "no", "y", "n"}
NEGATIVE   = {"", "false", "0", "no", "n", "none", "n/a", "na", "null", "unanswered",
              "not applicable", "not selected", "unchecked"}

QA_SCHEMA = T.ArrayType(T.StructType([
    T.StructField("question",   T.StringType()),
    T.StructField("answer",     T.StringType()),
    T.StructField("answer_key", T.StringType()),
    T.StructField("answered",   T.BooleanType()),
]))


def extract_qa(payload):
    if not payload:
        return []
    try:
        obj = json.loads(payload)
    except Exception:
        return []

    out, seen = [], set()

    def qtext(d):
        for k, v in d.items():
            if any(h in k.lower() for h in QUESTION_HINTS):
                if isinstance(v, str) and 3 < len(v) < 400:
                    return v.strip()
        return None

    def best_answer(d):
        """Rank candidate answers: real string > number > boolean."""
        best, best_key, best_rank = None, None, 99
        for k, v in d.items():
            if not any(h in k.lower() for h in ANSWER_HINTS):
                continue
            if v is None or isinstance(v, (dict, list)):
                continue
            s = str(v).strip()
            if not s:
                continue
            if isinstance(v, bool) or s.lower() in BOOLEANISH:
                rank = 3
            elif isinstance(v, (int, float)):
                rank = 2
            else:
                rank = 1
            if rank < best_rank:
                best, best_key, best_rank = s, k, rank
        return best, best_key

    def rec(o, depth=0):
        if depth > 12:
            return
        if isinstance(o, dict):
            q = qtext(o)
            a, akey = best_answer(o)
            if q and a is not None:
                key = (q[:200], a[:200], akey)
                if key not in seen:
                    seen.add(key)
                    out.append((q[:400], a[:200], akey,
                                a.strip().lower() not in NEGATIVE))
            for v in o.values():
                rec(v, depth + 1)
        elif isinstance(o, list):
            for v in o:
                rec(v, depth + 1)

    rec(obj)
    return out


extract_qa_udf = F.udf(extract_qa, QA_SCHEMA)

spark.sql("""
CREATE OR REPLACE TEMP VIEW scoped_in AS
SELECT * FROM scoped WHERE scope_flag='in_scope'
""")

qa_exploded = (spark.table("scoped_in")
    .select("TripRequestId", "ContractName", "LevelOfServiceDescription",
            "ClinicalData", "LosQuestions")
    .withColumn("qa", extract_qa_udf(F.col("LosQuestions")))
    .withColumn("p", F.explode_outer("qa"))
    .select("TripRequestId", "ContractName", "LevelOfServiceDescription",
            F.col("p.question").alias("question"),
            F.col("p.answer").alias("answer"),
            F.col("p.answer_key").alias("answer_key"),
            F.col("p.answered").alias("answered")))

qa_exploded.createOrReplaceTempView("qa_pairs")

print("PARSER DIAGNOSTIC - which key supplied the answer, and how varied are its values?")
print("A key with 1-2 distinct values across millions of rows is a form property, not an answer.")
display(spark.sql("""
    SELECT answer_key,
           count(*)                   AS uses,
           count(DISTINCT answer)      AS distinct_values,
           min(answer)                 AS example_min,
           max(answer)                 AS example_max
    FROM qa_pairs
    WHERE question IS NOT NULL
    GROUP BY answer_key
    ORDER BY uses DESC
"""))


# COMMAND ----------

# Health check. Compare against v2: answered_pairs should no longer be ~98% "True".
display(spark.sql("""
    SELECT count(*)                                  AS total_pairs,
           count(DISTINCT TripRequestId)             AS orders,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS answered_pairs,
           count(DISTINCT question)                  AS distinct_questions,
           count(DISTINCT answer)                    AS distinct_answers
    FROM qa_pairs WHERE question IS NOT NULL
"""))

print("Top answer values - should now look like real clinical responses:")
display(spark.sql("""
    SELECT answer, count(*) AS uses
    FROM qa_pairs WHERE answered AND answer IS NOT NULL
    GROUP BY answer ORDER BY uses DESC LIMIT 50
"""))


# COMMAND ----------

# =============================================================================
# 3. CMS CRITERIA MAPPING
#
# Maps the discovered questions to published CMS ambulance medical-necessity
# criteria. Per the 7/21 meeting, pull the authoritative text via Copilot and
# replace/extend these entries - the mapping below is a starting scaffold, not
# a citation.
#
# CMS core test: non-emergency ambulance is covered when the patient's
# condition contraindicates transport by any other means.
# =============================================================================

CMS_CRITERIA = {
    "BED_CONFINED": {
        "criterion": "Bed-confined - ALL THREE must be met: unable to get up from bed without "
                     "assistance, unable to ambulate, unable to sit in a chair or wheelchair.",
        "match": r"bed confined|bed[\s-]*bound|get up from bed|unable to ambulate|"
                 r"unable to sit in a chair|sit in a chair or wheelchair",
        "strength": "primary",
    },
    "OTHER_TRANSPORT_CONTRA": {
        "criterion": "Transport by other means is contraindicated by the patient's condition.",
        "match": r"contraindicat|other means|unable to (be )?transport|special positioning|"
                 r"handling due to illness|cannot be transported by",
        "strength": "primary",
    },
    "MONITORING_EN_ROUTE": {
        "criterion": "Requires monitoring or treatment of a medical condition during transport.",
        "match": r"monitoring of a medical condition|requires monitoring|monitor.*during transport|"
                 r"cardiac|telemetry|vital",
        "strength": "primary",
    },
    "AIRWAY_RESPIRATORY": {
        "criterion": "Requires airway management or respiratory support during transport.",
        "match": r"ventilator|intubat|trach|suction|cpap|bipap|oxygen|airway",
        "strength": "supporting",
    },
    "MEDICATION_EN_ROUTE": {
        "criterion": "Requires administration or monitoring of medication during transport.",
        "match": r"administration of medication|medication during transport|iv |infusion|drip",
        "strength": "supporting",
    },
    "RESTRAINT_BEHAVIORAL": {
        "criterion": "Requires restraints or monitoring for a psychiatric/behavioral condition.",
        "match": r"restrain|psychiatric|behavioral|altered mental status|sedation|psychiatric hold",
        "strength": "supporting",
    },
    "WOUND_DRAINAGE": {
        "criterion": "Open, draining, or pressure wound requiring positioning or care.",
        "match": r"draining|pressure wound|wound is stage|ostomy|drain",
        "strength": "supporting",
    },
    "ISOLATION": {
        "criterion": "Isolation or infection-control precautions required during transport.",
        "match": r"isolation|precaution|contagio|infect",
        "strength": "supporting",
    },
}

cms_case = "\n        ".join(
    f"WHEN lower(question) RLIKE '{v['match']}' THEN '{k}'" for k, v in CMS_CRITERIA.items())

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW qa_cms AS
SELECT *,
    CASE
        {cms_case}
        ELSE 'UNMAPPED'
    END AS cms_criterion
FROM qa_pairs
WHERE question IS NOT NULL
""")

print("Questions mapped to CMS criteria:")
display(spark.sql("""
    SELECT cms_criterion,
           count(DISTINCT question)                  AS distinct_questions,
           count(*)                                  AS times_asked,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS times_answered
    FROM qa_cms GROUP BY cms_criterion ORDER BY times_answered DESC
"""))


# COMMAND ----------

# The mapping itself - review this with Jen Jones / Michelle's team.
print("Question -> CMS criterion (review list):")
display(spark.sql("""
    SELECT cms_criterion, question,
           count(*)                                  AS times_asked,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS times_answered
    FROM qa_cms
    GROUP BY cms_criterion, question
    ORDER BY cms_criterion, times_answered DESC
"""))


# COMMAND ----------

# Unmapped questions - these need a CMS criterion assigned or explicit exclusion.
display(spark.sql("""
    SELECT question, count(*) AS times_asked
    FROM qa_cms WHERE cms_criterion='UNMAPPED'
    GROUP BY question ORDER BY times_asked DESC LIMIT 100
"""))


# COMMAND ----------

# =============================================================================
# 4. DOCUMENTATION BUCKETS  (free text only - the corrected method)
# =============================================================================

TEXT_CONCEPTS = {
    "mobility_deficit":  r"hemipar|hemipleg|paraly|non[\s-]*ambulat|unable to (bear weight|ambulate|walk|stand)|"
                         r"fracture|amputat|contracture|bear weight",
    "cannot_sit":        r"cannot sit|unable to sit|special positioning|supine|must lie|stretcher|"
                         r"cannot support trunk|unable to maintain",
    "bed_confined":      r"bed[\s-]*(bound|confined)|unable to get up|cannot get out of bed",
    "oxygen":            r"oxygen|\bo2\b|\blpm\b|nasal cannula|\bbipap\b|\bcpap\b|\bnc\b",
    "cardiac":           r"cardiac|telemetry|\bekg\b|\becg\b|\bnstemi\b|\bstemi\b|arrhythm|\bafib\b|chest pain",
    "ventilator":        r"ventilat|\bvent\b|trach|intubat",
    "suctioning":        r"suction",
    "iv_medication":     r"\biv\b|infusion|\bdrip\b|heparin|antibiotic|\btpn\b|\bppn\b",
    "wound_ostomy":      r"wound|ostomy|ulcer|decubitus|drain",
    "isolation":         r"isolation|\bmrsa\b|c\.? ?diff|precaution",
    "behavioral":        r"dementia|alzheimer|combative|agitat|altered mental|flight risk|elope|psych",
    "bariatric":         r"bariatric|morbid",
    # quality problems
    "weakness_only":     r"general(iz)?e?d? weakness|generally weak|^\s*weak",
    "fall_risk_only":    r"fall risk|unsteady|deconditio",
    "nonclinical":       r"per protocol|convenience|no other transport|unable to arrange|family request",
}

WEAK   = ["weakness_only", "fall_risk_only", "nonclinical"]
STRONG = [c for c in TEXT_CONCEPTS if c not in WEAK]

cols = ",\n        ".join(
    f"CASE WHEN lower(ClinicalData) RLIKE '{rx}' THEN 1 ELSE 0 END AS t_{n}"
    for n, rx in TEXT_CONCEPTS.items())

s_sum = " + ".join(f"t_{c}" for c in STRONG)
w_sum = " + ".join(f"t_{c}" for c in WEAK)

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW buckets AS
SELECT TripRequestId, ContractName, RequesterFacility,
       LevelOfService, LevelOfServiceDescription, ServiceType,
       ClinicalData, length(ClinicalData) AS clinical_len,
       {cols}
FROM scoped_in
""")

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW bucketed AS
SELECT *,
    ({s_sum}) AS strong_ct,
    ({w_sum}) AS weak_ct,
    CASE
        WHEN ClinicalData IS NULL OR length(trim(ClinicalData))=0 THEN 'no_documentation'
        WHEN ({s_sum})=0 AND ({w_sum})>0                          THEN 'weak_only'
        WHEN ({s_sum})=0                                          THEN 'unclassified'
        WHEN ({w_sum})>0                                          THEN 'mixed'
        ELSE 'specific'
    END AS bucket
FROM buckets
""")

print("Documentation buckets, IN-SCOPE ONLY (the corrected baseline):")
display(spark.sql("""
    SELECT bucket, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM bucketed GROUP BY bucket ORDER BY orders DESC
"""))


# COMMAND ----------

# =============================================================================
# 5. RISK TIERS / PAYMENT LIKELIHOOD
#
# The 7/21 meeting proposed scoring buckets by expected payment - roughly 5%
# for at-risk, up to 60% for valid clinical reasons.
#
# *** THESE RATES ARE PLACEHOLDERS, NOT MEASURED. ***
# They are the meeting's working assumption. Replace them with actual paid /
# denied rates once the Integra denial data lands and can be joined by trip.
# Until then this table shows RELATIVE risk and volume, not dollars.
# =============================================================================

ASSUMED_PAY_RATE = {          # placeholder - replace with measured rates
    "specific":         0.60,
    "mixed":            0.45,
    "unclassified":     0.30,
    "weak_only":        0.05,
    "no_documentation": 0.05,
}

RISK_TIER = {
    "specific":         "1 - low",
    "mixed":            "2 - medium",
    "unclassified":     "3 - unknown",
    "weak_only":        "4 - high",
    "no_documentation": "4 - high",
}

pay_case  = "\n        ".join(f"WHEN bucket='{k}' THEN {v}" for k, v in ASSUMED_PAY_RATE.items())
tier_case = "\n        ".join(f"WHEN bucket='{k}' THEN '{v}'" for k, v in RISK_TIER.items())

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW risk_scored AS
SELECT *,
    CASE {pay_case}  ELSE 0.0  END AS assumed_pay_rate,
    CASE {tier_case} ELSE '3 - unknown' END AS risk_tier
FROM bucketed
""")

print("Risk tiers (pay rates are ASSUMED - validate with denial data):")
display(spark.sql("""
    SELECT risk_tier, bucket, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct_of_inscope,
           max(assumed_pay_rate) AS assumed_pay_rate
    FROM risk_scored GROUP BY risk_tier, bucket ORDER BY risk_tier, orders DESC
"""))


# COMMAND ----------

# Risk by level of service - where the exposure concentrates.
display(spark.sql("""
    SELECT LevelOfServiceDescription,
           count(*)                                                        AS orders,
           sum(CASE WHEN risk_tier LIKE '4%' THEN 1 ELSE 0 END)            AS high_risk,
           round(100.0*sum(CASE WHEN risk_tier LIKE '4%' THEN 1 ELSE 0 END)
                 /count(*),1)                                              AS pct_high_risk
    FROM risk_scored
    GROUP BY LevelOfServiceDescription
    HAVING count(*) >= 100
    ORDER BY high_risk DESC
"""))


# COMMAND ----------

# =============================================================================
# 6. EXAMPLE TRIPS PER BUCKET   *** FRIDAY DELIVERABLE FOR DAVE ***
# Real orders from each bucket so the categories are concrete, not abstract.
# =============================================================================

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW examples AS
SELECT * FROM (
    SELECT bucket, risk_tier, ContractName, RequesterFacility,
           LevelOfServiceDescription, clinical_len,
           strong_ct, weak_ct,
           ClinicalData,
           row_number() OVER (PARTITION BY bucket ORDER BY clinical_len DESC) AS rn
    FROM risk_scored
) WHERE rn <= {EXAMPLES_PER_BUCKET}
""")

for b in ["no_documentation", "weak_only", "unclassified", "mixed", "specific"]:
    print(f"\n{'='*70}\nBUCKET: {b}\n{'='*70}")
    display(spark.sql(f"""
        SELECT ContractName, LevelOfServiceDescription, clinical_len, ClinicalData
        FROM examples WHERE bucket='{b}' ORDER BY rn
    """))


# COMMAND ----------

# Side-by-side contrast for the deck: same level of service, opposite quality.
print("CONTRAST PAIRS - BLS orders, weak vs specific:")
display(spark.sql("""
    SELECT bucket, ContractName, ClinicalData
    FROM risk_scored
    WHERE LevelOfServiceDescription = 'Basic life support'
      AND bucket IN ('weak_only','specific')
      AND ClinicalData IS NOT NULL
      AND length(ClinicalData) BETWEEN 20 AND 300
    ORDER BY bucket, length(ClinicalData)
    LIMIT 40
"""))


# COMMAND ----------

# =============================================================================
# 7. DRAFT MINIMUM DATA-ENTRY STANDARDS
# From the meeting: set a documentation floor so quality is consistent
# regardless of payer variability. Measured against current performance.
# =============================================================================

spark.sql("""
CREATE OR REPLACE TEMP VIEW standards AS
SELECT *,
    CASE WHEN ClinicalData IS NOT NULL AND length(trim(ClinicalData)) > 0
         THEN 1 ELSE 0 END                                 AS std_1_not_empty,
    CASE WHEN strong_ct > 0 THEN 1 ELSE 0 END              AS std_2_has_clinical_reason,
    CASE WHEN weak_ct = 0   THEN 1 ELSE 0 END              AS std_3_no_vague_only,
    CASE WHEN length(coalesce(ClinicalData,'')) >= 25
         THEN 1 ELSE 0 END                                 AS std_4_min_length
FROM risk_scored
""")

print("Current pass rate against each proposed minimum standard:")
display(spark.sql("""
    SELECT
        round(100.0*avg(std_1_not_empty),1)          AS pct_pass_not_empty,
        round(100.0*avg(std_2_has_clinical_reason),1) AS pct_pass_clinical_reason,
        round(100.0*avg(std_3_no_vague_only),1)       AS pct_pass_no_vague,
        round(100.0*avg(std_4_min_length),1)          AS pct_pass_min_length,
        round(100.0*avg(CASE WHEN std_1_not_empty=1 AND std_2_has_clinical_reason=1
                             THEN 1 ELSE 0 END),1)    AS pct_pass_all_core
    FROM standards
"""))

print("By contract - who needs the intervention most:")
display(spark.sql("""
    SELECT ContractName, count(*) AS orders,
           round(100.0*avg(std_1_not_empty),1)           AS pct_not_empty,
           round(100.0*avg(std_2_has_clinical_reason),1) AS pct_clinical_reason
    FROM standards GROUP BY ContractName ORDER BY orders DESC
"""))


# COMMAND ----------

# Worst-performing facilities - the targeted training list.
display(spark.sql("""
    SELECT RequesterFacility, ContractName, count(*) AS orders,
           round(100.0*avg(std_2_has_clinical_reason),1) AS pct_clinical_reason,
           sum(CASE WHEN risk_tier LIKE '4%' THEN 1 ELSE 0 END) AS high_risk_orders
    FROM standards
    GROUP BY RequesterFacility, ContractName
    HAVING count(*) >= 200
    ORDER BY pct_clinical_reason ASC
    LIMIT 30
"""))


# COMMAND ----------

# =============================================================================
# 8. EXCEL EXPORT
# =============================================================================

import pandas as pd
def q(sql): return spark.sql(sql).toPandas()

scope_summary = q("""
    SELECT scope_flag, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM scoped GROUP BY scope_flag ORDER BY orders DESC
""")
scope_detail = q("""
    SELECT LevelOfServiceDescription, ServiceType, count(*) AS orders
    FROM scoped WHERE scope_flag='out_of_scope' GROUP BY 1,2 ORDER BY orders DESC
""")
parser_diag = q("""
    SELECT answer_key, count(*) AS uses, count(DISTINCT answer) AS distinct_values
    FROM qa_pairs WHERE question IS NOT NULL GROUP BY answer_key ORDER BY uses DESC
""")
qa_health = q("""
    SELECT count(*) AS total_pairs, count(DISTINCT TripRequestId) AS orders,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS answered_pairs,
           count(DISTINCT question) AS distinct_questions,
           count(DISTINCT answer) AS distinct_answers
    FROM qa_pairs WHERE question IS NOT NULL
""").T.reset_index(); qa_health.columns = ["Metric", "Value"]

cms_summary = q("""
    SELECT cms_criterion, count(DISTINCT question) AS distinct_questions,
           count(*) AS times_asked, sum(CASE WHEN answered THEN 1 ELSE 0 END) AS times_answered
    FROM qa_cms GROUP BY cms_criterion ORDER BY times_answered DESC
""")
cms_mapping = q("""
    SELECT cms_criterion, question, count(*) AS times_asked,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS times_answered
    FROM qa_cms GROUP BY cms_criterion, question ORDER BY cms_criterion, times_answered DESC
""")
cms_unmapped = q("""
    SELECT question, count(*) AS times_asked FROM qa_cms
    WHERE cms_criterion='UNMAPPED' GROUP BY question ORDER BY times_asked DESC LIMIT 200
""")
answer_vals = q("""
    SELECT answer, count(*) AS uses FROM qa_pairs
    WHERE answered AND answer IS NOT NULL GROUP BY answer ORDER BY uses DESC LIMIT 300
""")
bucket_summary = q("""
    SELECT bucket, count(*) AS orders, round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM bucketed GROUP BY bucket ORDER BY orders DESC
""")
risk_summary = q("""
    SELECT risk_tier, bucket, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct_of_inscope,
           max(assumed_pay_rate) AS assumed_pay_rate
    FROM risk_scored GROUP BY risk_tier, bucket ORDER BY risk_tier, orders DESC
""")
risk_by_los = q("""
    SELECT LevelOfServiceDescription, count(*) AS orders,
           sum(CASE WHEN risk_tier LIKE '4%' THEN 1 ELSE 0 END) AS high_risk,
           round(100.0*sum(CASE WHEN risk_tier LIKE '4%' THEN 1 ELSE 0 END)/count(*),1) AS pct_high_risk
    FROM risk_scored GROUP BY LevelOfServiceDescription
    HAVING count(*) >= 50 ORDER BY high_risk DESC
""")
bucket_by_contract = q("""
    SELECT ContractName, bucket, count(*) AS orders
    FROM bucketed GROUP BY ContractName, bucket ORDER BY ContractName, orders DESC
""")
examples_df = q("""
    SELECT bucket, risk_tier, ContractName, LevelOfServiceDescription,
           clinical_len, strong_ct, weak_ct, ClinicalData
    FROM examples ORDER BY bucket, rn
""")
contrast_df = q("""
    SELECT bucket, ContractName, ClinicalData
    FROM risk_scored
    WHERE LevelOfServiceDescription='Basic life support'
      AND bucket IN ('weak_only','specific')
      AND ClinicalData IS NOT NULL AND length(ClinicalData) BETWEEN 20 AND 300
    ORDER BY bucket, length(ClinicalData) LIMIT 60
""")
standards_overall = q("""
    SELECT round(100.0*avg(std_1_not_empty),1) AS pct_not_empty,
           round(100.0*avg(std_2_has_clinical_reason),1) AS pct_clinical_reason,
           round(100.0*avg(std_3_no_vague_only),1) AS pct_no_vague,
           round(100.0*avg(std_4_min_length),1) AS pct_min_length,
           round(100.0*avg(CASE WHEN std_1_not_empty=1 AND std_2_has_clinical_reason=1
                                THEN 1 ELSE 0 END),1) AS pct_pass_core
    FROM standards
""").T.reset_index(); standards_overall.columns = ["Standard", "Current pass rate %"]

standards_contract = q("""
    SELECT ContractName, count(*) AS orders,
           round(100.0*avg(std_1_not_empty),1) AS pct_not_empty,
           round(100.0*avg(std_2_has_clinical_reason),1) AS pct_clinical_reason
    FROM standards GROUP BY ContractName ORDER BY orders DESC
""")
worst_facilities = q("""
    SELECT RequesterFacility, ContractName, count(*) AS orders,
           round(100.0*avg(std_2_has_clinical_reason),1) AS pct_clinical_reason,
           sum(CASE WHEN risk_tier LIKE '4%' THEN 1 ELSE 0 END) AS high_risk_orders
    FROM standards GROUP BY RequesterFacility, ContractName
    HAVING count(*) >= 200 ORDER BY pct_clinical_reason ASC LIMIT 50
""")
text_concepts = pd.concat([
    spark.sql(f"""SELECT '{n}' AS concept, sum(t_{n}) AS orders,
                  round(100.0*sum(t_{n})/count(*),2) AS pct FROM bucketed""").toPandas()
    for n in TEXT_CONCEPTS], ignore_index=True).sort_values("orders", ascending=False)
text_concepts["type"] = text_concepts["concept"].apply(
    lambda c: "quality problem" if c in WEAK else "supports necessity")

print("tables built")


# COMMAND ----------

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR = PatternFill("solid", fgColor="1F4E79")
HF  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BF  = Font(name="Arial", size=10)
TF  = Font(name="Arial", bold=True, size=14, color="1F4E79")
NF  = Font(name="Arial", size=9, italic=True, color="666666")


def sheet(wb, name, df, title=None, note=None):
    ws = wb.create_sheet(name[:31]); r = 1
    if title: ws.cell(row=1, column=1, value=title).font = TF; r = 2
    if note:  ws.cell(row=r, column=1, value=note).font = NF; r += 1
    r += 1; h = r
    for j, c in enumerate(df.columns, 1):
        x = ws.cell(row=h, column=j, value=str(c)); x.fill, x.font = HDR, HF
        x.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(df.itertuples(index=False), h + 1):
        for j, v in enumerate(row, 1):
            if isinstance(v, str) and len(v) > 900: v = v[:900] + "..."
            ws.cell(row=i, column=j, value=v).font = BF
    for j, c in enumerate(df.columns, 1):
        try: w = max(len(str(c)), int(df[c].astype(str).str.len().quantile(0.9)))
        except Exception: w = len(str(c))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 3, 12), 80)
    ws.freeze_panes = ws.cell(row=h + 1, column=1)
    return ws, h, h + len(df)


def bar(ws, t, cc, vc, h, l, a):
    ch = BarChart(); ch.type, ch.style, ch.title = "col", 10, t
    ch.add_data(Reference(ws, min_col=vc, min_row=h, max_row=l), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=cc, min_row=h + 1, max_row=l))
    ch.height, ch.width, ch.legend = 8, 18, None; ws.add_chart(ch, a)


def pie(ws, t, cc, vc, h, l, a):
    ch = PieChart(); ch.title = t
    ch.add_data(Reference(ws, min_col=vc, min_row=h, max_row=l), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=cc, min_row=h + 1, max_row=l))
    ch.height, ch.width = 9, 13; ws.add_chart(ch, a)


wb = Workbook(); wb.remove(wb.active)

ws = wb.create_sheet("README"); ws.column_dimensions["A"].width = 115
for i, (t, f) in enumerate([
    ("Medical Necessity Analysis v3", TF), ("", BF),
    ("Source: `prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster (read-only).", BF),
    ("Scope: non-emergent GROUND only. Rideshare, air, and emergent are excluded.", BF),
    ("", BF),
    ("WHAT CHANGED FROM v2", Font(name="Arial", bold=True, size=11)),
    ("1. Scope filter applied - v2 counted rideshare and air in the denominator.", BF),
    ("2. Parser fixed - v2 grabbed a form property ('True' 98.7% of answers) instead of", BF),
    ("   the real answer. Now ranks candidates: string > number > boolean.", BF),
    ("3. Questions mapped to CMS criteria, not just internal concepts.", BF),
    ("4. Risk tiers added with assumed pay rates.", BF),
    ("5. Example trips per bucket for the Dave review.", BF),
    ("6. Draft minimum data-entry standards with current pass rates.", BF),
    ("", BF),
    ("BUCKET DEFINITIONS", Font(name="Arial", bold=True, size=11)),
    ("  specific         - clinical reason present, no vague filler", BF),
    ("  mixed            - clinical reason AND vague filler", BF),
    ("  weak_only        - vague filler only", BF),
    ("  unclassified     - text present, nothing recognized (vocabulary gap, not a verdict)", BF),
    ("  no_documentation - free-text box empty", BF),
    ("", BF),
    ("CAVEATS", Font(name="Arial", bold=True, size=11)),
    ("- Pay rates in Risk Tiers are ASSUMED from the 7/21 meeting, not measured.", BF),
    ("  Replace with actual paid/denied rates once Integra denial data is joined.", BF),
    ("- CMS mapping is a scaffold. Pull authoritative text via Copilot and revise.", BF),
    ("- Concept term lists not yet validated by Jen Jones / Michelle's team.", BF),
], 1):
    ws.cell(row=i, column=1, value=t).font = f

sheet(wb, "Scope", scope_summary, "In-scope vs out-of-scope",
      "Rideshare, air, and emergent need no medical-necessity documentation.")
sheet(wb, "Scope Detail", scope_detail, "What was excluded", "Verify this list is correct.")
sheet(wb, "Parser Diagnostic", parser_diag, "Which key supplied each answer",
      "A key with 1-2 distinct values across millions of rows is a form property, not an answer.")
sheet(wb, "QA Health", qa_health, "Question/answer extraction health")

ws, h, l = sheet(wb, "CMS Summary", cms_summary, "Questions mapped to CMS criteria")
bar(ws, "Answered by CMS criterion", 1, 4, h, l, "G4")

sheet(wb, "CMS Mapping", cms_mapping, "Every question and its CMS criterion",
      "Review with Jen Jones / Michelle's team.")
sheet(wb, "CMS Unmapped", cms_unmapped, "Questions with no CMS criterion assigned",
      "Assign a criterion or mark explicitly out of scope.")
sheet(wb, "Answer Values", answer_vals, "What nurses actually answer")

ws, h, l = sheet(wb, "Buckets", bucket_summary, "Documentation buckets (in-scope only)")
pie(ws, "Documentation buckets", 1, 2, h, l, "F4")

sheet(wb, "Buckets x Contract", bucket_by_contract, "Buckets by customer")

ws, h, l = sheet(wb, "Risk Tiers", risk_summary, "Risk tiers and assumed pay rates",
                 "PAY RATES ARE ASSUMED, NOT MEASURED. Validate with denial data.")
bar(ws, "Orders by risk tier", 2, 3, h, l, "G4")

ws, h, l = sheet(wb, "Risk by LOS", risk_by_los, "High-risk orders by level of service")
bar(ws, "High-risk orders by LOS", 1, 3, h, l, "G4")

ws, h, l = sheet(wb, "Text Concepts", text_concepts, "Clinical concepts in the free text")
bar(ws, "Concepts in free text", 1, 2, h, l, "F4")

sheet(wb, "EXAMPLES", examples_df, "Example trips per bucket  *** FOR DAVE REVIEW ***",
      "Real orders. Review for patient information before sharing outside the team.")
sheet(wb, "Contrast BLS", contrast_df, "BLS: weak vs specific, side by side",
      "Same service level, opposite documentation quality. Good for the deck.")

sheet(wb, "Standards", standards_overall, "Draft minimum data-entry standards",
      "Current pass rate against each proposed floor.")
sheet(wb, "Standards x Contract", standards_contract, "Standards pass rate by customer")
sheet(wb, "Worst Facilities", worst_facilities, "Facilities needing intervention first",
      "Min 200 orders. Sorted by lowest clinical-reason rate.")

wb.save(OUTPUT_XLSX)
print("WROTE:", OUTPUT_XLSX)
print("sheets:", wb.sheetnames)


# COMMAND ----------

import os
for f in sorted(os.listdir(OUTPUT_DIR)):
    print(f"{f}  ({os.path.getsize(os.path.join(OUTPUT_DIR,f)):,} bytes)")

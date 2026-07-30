# Medical Necessity Scoring - revised schema

# Reframes the non-emergent ground analysis around **medical necessity and transport
# appropriateness** rather than payment or denial. Per the 24 July review:

# - Two scoring axes, not one label: a **mobility** axis (why other transport is
#   contraindicated) and a **monitoring** axis (why this level of service).
# - A **two-bucket** determination - clearly necessary / clearly not necessary - with an
#   explicit **indeterminate** middle to be refined over time.
# - **Weighted concepts** and a **confidence** score, both driven from one config block.
# - No payment-focused terminology in any output column.

# Read-only throughout. No table writes.

# CMS source documents (every field, concept, and term below cites one of these):
#   [BPM10]  Medicare Benefit Policy Manual, Pub. 100-02, Ch. 10 - Ambulance Services
#            https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/bp102c10.pdf
#            - 10.2.1 general medical necessity test; 10.2.3 bed-confined three-prong test.
#   [410.40] 42 CFR 410.40 - Coverage of ambulance services (levels of service, physician cert.)
#            https://www.ecfr.gov/current/title-42/chapter-IV/subchapter-B/part-410/subpart-B/section-410.40
#   [414.605] 42 CFR 414.605 - Definitions of BLS, ALS1, ALS2, SCT, ALS assessment/intervention
#            https://www.ecfr.gov/current/title-42/chapter-IV/subchapter-B/part-414/subpart-H/section-414.605
#   [CPM15]  Medicare Claims Processing Manual, Pub. 100-04, Ch. 15 - Ambulance (adjudication)
#            https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c15ambulance.pdf
#   [RSN]    CMS Ambulance Transport Reason Codes and Statements (denial codes, e.g. AM600)
#            https://www.cms.gov/files/document/ambulance-transport-reason-codes-statements.pdf
#   [MLN]    CMS MLN Ambulance Services - provider compliance tips (documentation guidance)
#            https://www.cms.gov/training-education/medicare-learning-networkr-mln/compliance/medicare-provider-compliance-tips/ambulance-services

# 0. Environment

# Written for Databricks (`spark` session assumed present). The `display()` helper below
# lets the same notebook run in a plain Jupyter/VS Code kernel, where it falls back to a
# pandas-style preview. Delete this cell if running only in Databricks.

try:
    display  # provided by Databricks
except NameError:
    def display(sdf):
        try:
            print(sdf.limit(50).toPandas().to_string(index=False))
        except AttributeError:
            print(sdf)

# 1. Source fields

# The source table is a read-only snapshot. Each field below is cited to the CMS document
# that makes it relevant to the medical necessity determination.

# The catalog name contains a hyphen, so each identifier part must be back-quoted for
# Spark SQL. SOURCE_TABLE keeps the plain name; SOURCE_TABLE_SQL is the quoted form used
# when reading. (Unquoted `prod-sandbox` raises INVALID_IDENTIFIER / SQLSTATE 42602.)
SOURCE_TABLE     = "prod-sandbox.vivekkumar_patel.temp_tnet_tripmaster"
SOURCE_TABLE_SQL = ".".join(f"`{p}`" for p in SOURCE_TABLE.split("."))

# Logical fields mapped to candidate column names. The first candidate present in the table
# wins. This avoids INVALID_IDENTIFIER / UNRESOLVED_COLUMN when the real schema differs.
# Edit the candidate lists if the actual column names are not covered.
FIELD_CANDIDATES = {
    "free_text": ["ClinicalData", "clinical_data", "ClinicalNotes", "Reason", "ReasonForTransport"],
    "los":       ["LevelOfService", "level_of_service", "LOS", "ServiceLevel"],
    "customer":  ["Customer", "CustomerName", "HealthSystem", "Facility", "Account"],
    "order_id":  ["OrderId", "order_id", "OrderID", "TripId", "TripID", "TripLegId"],
}

# 42 CFR / BPM10 citations for the logical fields (see resolution below):
#   free_text -> [BPM10] 10.2.1, 10.2.4 ; [RSN] AM600   (reason must be recorded)
#   los       -> [414.605] ; [410.40](c)                (level billed must be necessary)
#   customer / order_id -> operational only, no CMS basis

# 2. Concept dictionary - the single source of truth

# Every concept carries its axis, weight, group, CMS basis, source tag, source URL, and the
# term pattern matched against ClinicalData. The url column makes the citation explicit next to
# the terms it governs. Weights are 0-3, reflecting how directly the concept establishes the
# criterion (not clinical severity).

#   axis: "mobility"   = transport necessity   [BPM10] 10.2.1 / 10.2.3  (why not a wheelchair van/car)
#         "monitoring" = level of service      [414.605]               (why ALS / CCT rather than BLS)
#         "filler"     = no necessity alone     [MLN] / [BPM10] 10.2.1
#   basis: "named"    = concept is explicit in the cited CMS text
#          "inferred" = derived from the [BPM10] 10.2.1 general test, not separately named

BPM10   = "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/bp102c10.pdf"
CFR414  = "https://www.ecfr.gov/current/title-42/chapter-IV/subchapter-B/part-414/subpart-H/section-414.605"
MLN     = "https://www.cms.gov/training-education/medicare-learning-networkr-mln/compliance/medicare-provider-compliance-tips/ambulance-services"

# fields: (concept, axis, weight, group, basis, cms_ref, source_url, term_pattern)
CONCEPTS = [
    # --- mobility axis: transport necessity, BPM10 10.2.3 bed-confined test (prongs) + 10.2.1 ---
    # bed_confined: [BPM10] 10.2.3 - "unable to get up from bed without assistance". One prong.
    ("bed_confined",     "mobility",     3, "A", "named",    "BPM10 10.2.3 (prong: get up)",       BPM10,
        r"bed ?(bound|confined)|unable to get up|cannot get out of bed"),
    # mobility_deficit: [BPM10] 10.2.3 - "unable to ambulate" (prong 2); also 10.2.1 contraindication.
    ("mobility_deficit", "mobility",     3, "A", "named",    "BPM10 10.2.3 (prong: ambulate)",     BPM10,
        r"hemipar|hemipleg|paraly|non ?ambulat|unable to (bear weight|ambulate|walk|stand)|fracture|amputat|contracture|bear weight"),
    # cannot_sit: [BPM10] 10.2.3 - "unable to sit in a chair or wheelchair" (prong 3).
    ("cannot_sit",       "mobility",     3, "A", "named",    "BPM10 10.2.3 (prong: sit)",          BPM10,
        r"cannot sit|unable to sit|special positioning|supine|must lie|stretcher|cannot support trunk"),
    # bariatric: not named by CMS; inferred handling requirement under [BPM10] 10.2.1 general test.
    ("bariatric",        "mobility",     2, "A", "inferred", "BPM10 10.2.1 (inferred: handling)",  BPM10,
        r"bariatric|morbid"),
    # wound_ostomy: not named; inferred positioning need under [BPM10] 10.2.1.
    ("wound_ostomy",     "mobility",     1, "A", "inferred", "BPM10 10.2.1 (inferred: position)",  BPM10,
        r"wound|ostomy|ulcer|decubitus|drain"),
    # behavioral: not named; inferred safety need under [BPM10] 10.2.1. Weakest - CMS wants a
    #   physical limitation, so this is a candidate to move to Group B on SME review.
    ("behavioral",       "mobility",     1, "A", "inferred", "BPM10 10.2.1 (inferred: safety)",    BPM10,
        r"dementia|alzheimer|combative|agitat|altered mental|flight risk|elope"),

    # --- monitoring axis: level of service, 42 CFR 414.605 definitions ---
    # ventilator: [414.605] ALS2 (airway management) / SCT (respiratory care) territory.
    ("ventilator",       "monitoring",   3, "A", "named",    "414.605 (ALS2 airway / SCT)",        CFR414,
        r"ventilat|vent|trach|intubat"),
    # suctioning: [414.605] SCT - care beyond the EMT-Paramedic scope of practice.
    ("suctioning",       "monitoring",   3, "A", "named",    "414.605 (SCT)",                      CFR414,
        r"suction"),
    # iv_medication: [414.605] ALS2 - administration of 3+ IV medications / central line.
    ("iv_medication",    "monitoring",   3, "A", "named",    "414.605 (ALS2 IV meds)",             CFR414,
        r"\biv\b|infusion|drip|heparin|antibiotic|tpn"),
    # cardiac: [414.605] ALS assessment / ALS2 cardiac procedures (monitoring, defib, pacing).
    ("cardiac",          "monitoring",   2, "A", "named",    "414.605 (ALS assessment)",           CFR414,
        r"cardiac|telemetry|ekg|ecg|nstemi|stemi|arrhythm|afib"),
    # oxygen: not a named necessity concept; [BPM10] 10.2.1 - chronic O2 alone does not qualify.
    ("oxygen",           "monitoring",   1, "A", "inferred", "BPM10 10.2.1 (O2 alone insufficient)", BPM10,
        r"oxygen|\bo2\b|lpm|nasal cannula|bipap|cpap"),
    # isolation: not named; inferred under [BPM10] 10.2.1 - infection control, not a patient contraindication.
    ("isolation",        "monitoring",   1, "A", "inferred", "BPM10 10.2.1 (infection control)",   BPM10,
        r"isolation|mrsa|c\.? ?diff|precaution"),

    # --- filler: appears in text but establishes no necessity on its own ---
    # weakness_only: [MLN] / MAC guidance - "severe generalized weakness" counts only with a
    #   qualifier and functional consequence; the bare term is vague and of little value on review.
    ("weakness_only",    "filler",       0, "B", "named",    "MLN / MAC (vague weakness)",         MLN,
        r"general(iz)?e?d? weakness|generally weak|^\s*weak"),
    # fall_risk_only: [MLN] / MAC guidance - a risk is not a contraindication to other transport.
    ("fall_risk_only",   "filler",       0, "B", "named",    "MLN / MAC (risk != contra)",         MLN,
        r"fall risk|unsteady|deconditio"),
    # nonclinical: [BPM10] 10.2.1 / [410.40](d) - a physician order alone does not prove necessity.
    ("nonclinical",      "filler",       0, "B", "named",    "BPM10 10.2.1 (order != necessity)",  BPM10,
        r"per protocol|convenience|no other transport|unable to arrange|family request"),
]

# Thresholds - tune here. A single named, weight-3 concept (score 3) is a clear reason.
MOBILITY_CLEAR    = 3   # mobility score at/above this = clearly necessary on the mobility axis
MONITORING_CLEAR  = 3   # monitoring score at/above this = clearly necessary on the monitoring axis
INDETERMINATE_MIN = 1   # any Group A signal below the clear thresholds = indeterminate

# The three prongs of the bed-confined test. [BPM10] 10.2.3 requires ALL THREE to be met, and
# states bed confinement is neither sufficient nor necessary by itself - it is one factor.
BED_CONFINED_PRONGS = ["bed_confined", "mobility_deficit", "cannot_sit"]  # cite: [BPM10] 10.2.3

# 3. Scope - non-emergent ground only

# Rideshare, air, and emergent trips are excluded; they do not carry the [BPM10] 10.2.1
# non-emergency documentation requirement. Adjust the filter to the real column values.

from pyspark.sql import functions as F

df = spark.table(SOURCE_TABLE_SQL)

# Resolve each logical field to a real column. First candidate present wins; unresolved fields
# come back as None and any output depending on them is skipped rather than erroring.
_cols_lower = {c.lower(): c for c in df.columns}
def resolve(field):
    for cand in FIELD_CANDIDATES[field]:
        if cand.lower() in _cols_lower:
            return _cols_lower[cand.lower()]
    return None

FREE_TEXT_COL = resolve("free_text")
LOS_COL       = resolve("los")
CUSTOMER_COL  = resolve("customer")
ORDER_ID_COL  = resolve("order_id")

print("Resolved columns:")
print(f"  free_text -> {FREE_TEXT_COL}")
print(f"  los       -> {LOS_COL}")
print(f"  customer  -> {CUSTOMER_COL}")
print(f"  order_id  -> {ORDER_ID_COL}")
if FREE_TEXT_COL is None:
    raise ValueError(
        "No free-text column found. Add the real column name to "
        "FIELD_CANDIDATES['free_text']. Available columns: " + ", ".join(df.columns)
    )

# Scope: non-emergent ground only. Rideshare, air, and emergent trips are excluded; they do not
# carry the [BPM10] 10.2.1 non-emergency documentation requirement.

# IMPORTANT: scope is defined by EXCLUSION, not by an exact-match whitelist. A whitelist silently
# drops every row whose level-of-service string is spelled or coded differently than expected
# (e.g. "BLS" vs "Basic life support"), which can collapse the dataset to a handful of rows.
# Air / rideshare are removed by pattern; emergent is removed via a trip-type flag if present.

raw_count = df.count()

# Show what level-of-service values actually exist, so the scope can be verified, not assumed.
if LOS_COL is not None:
    print("Distinct " + LOS_COL + " values (top 40 by count):")
    df.groupBy(LOS_COL).count().orderBy(F.desc("count")).show(40, truncate=False)

# Remove non-ground levels of service by pattern (keeps BLS/ALS/CCT/Team/wheelchair/ambulatory
# regardless of exact spelling). Extend the pattern if the real data uses other labels.
NON_GROUND_PATTERN = r"air|fixed ?wing|rotor|helicopter|flight|rideshare|uber|lyft|livery|taxi"
if LOS_COL is not None:
    df = df.filter(~F.lower(F.coalesce(F.col(LOS_COL), F.lit(""))).rlike(NON_GROUND_PATTERN))
else:
    print("WARNING: no level-of-service column resolved - non-ground exclusion NOT applied.")

# Remove emergent trips via a trip-type / priority flag if one is present.
EMERGENT_PATTERN = r"emergen|911|lights|code ?3|stat"
for trip_type_col in ("TripType", "trip_type", "TransportType", "Priority", "CallType"):
    if trip_type_col in df.columns:
        df = df.filter(~F.lower(F.coalesce(F.col(trip_type_col), F.lit(""))).rlike(EMERGENT_PATTERN))
        print(f"Applied emergent exclusion on column: {trip_type_col}")
        break
else:
    print("NOTE: no trip-type column found for emergent exclusion - verify emergent trips are "
          "not in this table, or add the column to the loop above.")

df = df.withColumn("_text", F.lower(F.coalesce(F.col(FREE_TEXT_COL), F.lit(""))))
df = df.withColumn("_has_text", F.length(F.trim(F.col("_text"))) > 0)

scope_count = df.count()
has_text_count = df.filter(F.col("_has_text")).count()
print(f"\nRaw rows:            {raw_count:,}")
print(f"In scope:            {scope_count:,}  ({scope_count/raw_count*100:.1f}% of raw)")
print(f"  with free text:    {has_text_count:,}  ({has_text_count/max(scope_count,1)*100:.1f}% of scope)")

# Guardrails - catch the failure mode that produced the TEAM-only, all-zero result.
if scope_count < 0.2 * raw_count:
    print("\nWARNING: scope kept under 20% of rows. Check the distinct level-of-service values "
          "above - the exclusion pattern may be dropping valid ground rows, or the wrong column "
          "resolved. Do not use these results until the scope looks right.")
if has_text_count < 0.5 * max(scope_count, 1):
    print("\nWARNING: over half of in-scope orders have no free text. Confirm FREE_TEXT_COL "
          "resolved to the real reason-for-transport field (see the 'Resolved columns' output).")
    print("Sample of the resolved free-text field:")
    df.select(FREE_TEXT_COL).filter(F.col("_has_text")).show(5, truncate=80)

# 4. Concept tagging

# One boolean column per concept, the term pattern matched against ClinicalData with `rlike`.
# The CMS citation for each pattern is in the CONCEPTS table above.

for name, axis, weight, group, basis, ref, url, pattern in CONCEPTS:
    df = df.withColumn(f"c_{name}", F.col("_text").rlike(pattern) & F.col("_has_text"))

# 5. Scores, classification, confidence, and level-of-service recommendation

# All derived from the config - no hand-coded per-concept logic below this point.

def axis_score(axis):
    terms = [F.when(F.col(f"c_{n}"), F.lit(w)).otherwise(F.lit(0))
             for n, a, w, *_ in CONCEPTS if a == axis]
    expr = terms[0]
    for t in terms[1:]:
        expr = expr + t
    return expr

# mobility_score: sum of mobility-axis concept weights. Basis: [BPM10] 10.2.1 / 10.2.3.
df = df.withColumn("mobility_score",   axis_score("mobility"))
# monitoring_score: sum of monitoring-axis concept weights. Basis: [414.605].
df = df.withColumn("monitoring_score", axis_score("monitoring"))

# named vs inferred / filler counts, for confidence and the clear threshold.
# named = concept explicit in CMS text; only named concepts can reach clearly_necessary alone.
named_a = [n for n, a, w, g, b, *_ in CONCEPTS if g == "A" and b == "named"]
df = df.withColumn("named_a_hits", sum([F.col(f"c_{n}").cast("int") for n in named_a]))
df = df.withColumn("has_named_a", F.col("named_a_hits") > 0)
df = df.withColumn("any_a", (F.col("mobility_score") + F.col("monitoring_score")) > 0)

# bed_confined_full: all three [BPM10] 10.2.3 prongs present = the full objective standard.
prong_expr = F.col(f"c_{BED_CONFINED_PRONGS[0]}")
for p in BED_CONFINED_PRONGS[1:]:
    prong_expr = prong_expr & F.col(f"c_{p}")
df = df.withColumn("bed_confined_full", prong_expr)  # cite: [BPM10] 10.2.3

filler_cols = [f"c_{n}" for n, a, *_ in CONCEPTS if a == "filler"]
df = df.withColumn("any_filler", sum([F.col(c).cast("int") for c in filler_cols]) > 0)

# necessity_class: clearly_necessary / clearly_not_necessary / indeterminate.
# clearly_necessary requires the full [BPM10] 10.2.3 test, OR a named CMS concept at the clear
#   threshold - inferred concepts alone cannot reach it.
# clearly_not_necessary maps to [RSN] AM600 territory: no documented reason other transport
#   was contraindicated (empty, or filler only).
df = df.withColumn(
    "necessity_class",
    F.when(
        F.col("bed_confined_full") |
        (F.col("has_named_a") &
         ((F.col("mobility_score") >= MOBILITY_CLEAR) | (F.col("monitoring_score") >= MONITORING_CLEAR))),
        F.lit("clearly_necessary"),          # cite: [BPM10] 10.2.3 / 10.2.1 ; [414.605]
    ).when(
        ~F.col("any_a") & (~F.col("_has_text") | F.col("any_filler")),
        F.lit("clearly_not_necessary"),      # cite: [RSN] AM600 ; [MLN]
    ).otherwise(F.lit("indeterminate")),
)

# indeterminate_reason: text_unmatched (text present, no term matched - the LLM target) vs
#   weak_or_inferred_only (some signal, below the clear threshold).
df = df.withColumn(
    "indeterminate_reason",
    F.when(F.col("necessity_class") != "indeterminate", F.lit(None))
     .when(F.col("_has_text") & ~F.col("any_a") & ~F.col("any_filler"), F.lit("text_unmatched"))
     .otherwise(F.lit("weak_or_inferred_only")),
)

# confidence in the classification itself (not a CMS construct - internal quality flag).
df = df.withColumn(
    "confidence",
    F.when(F.col("bed_confined_full"), F.lit("high"))
     .when((F.col("named_a_hits") >= 2), F.lit("high"))
     .when(~F.col("_has_text"), F.lit("high"))
     .when(F.col("named_a_hits") == 1, F.lit("medium"))
     .when(F.col("any_filler") & ~F.col("any_a"), F.lit("medium"))
     .otherwise(F.lit("low")),
)

# recommended_los: level of service implied by the monitoring axis, per [414.605] definitions.
#   ventilator/suctioning -> CCT/SCT ; multiple monitoring or IV meds -> ALS ; else BLS.
#   Descriptive only until CMS BLS/ALS eligibility criteria are confirmed - not a billing call.
df = df.withColumn(
    "recommended_los",
    F.when(F.col("c_ventilator") | F.col("c_suctioning"), F.lit("CCT/SCT"))          # cite: [414.605] SCT/ALS2
     .when((F.col("monitoring_score") >= 2) | F.col("c_iv_medication"), F.lit("ALS"))  # cite: [414.605] ALS1/ALS2
     .when(F.col("mobility_score") >= INDETERMINATE_MIN, F.lit("BLS"))               # cite: [414.605] BLS
     .otherwise(F.lit("none/indeterminate")),
)

# total_necessity_score: the aggregate the 24 July review asked for. It combines both axes and
# adds a bonus when the full [BPM10] 10.2.3 three-prong test is met, so a complete objective
# standard outscores an equal sum of unrelated concepts.
#   total = mobility_score + monitoring_score + (BED_CONFINED_BONUS if all three prongs present)
# Interpretation is provisional - the multiplier/bonus is a tunable modelling choice, not a CMS
# figure, and should be validated by the SMEs and against denial outcomes.
BED_CONFINED_BONUS = 3   # cite: [BPM10] 10.2.3 - full three-prong test is the strongest signal
df = df.withColumn(
    "total_necessity_score",
    F.col("mobility_score") + F.col("monitoring_score")
    + F.when(F.col("bed_confined_full"), F.lit(BED_CONFINED_BONUS)).otherwise(F.lit(0)),
)

# 6. Summary outputs (display only - nothing is written)

# Necessity distribution
display(
    df.groupBy("necessity_class")
      .agg(F.count("*").alias("orders"))
      .orderBy(F.desc("orders"))
)

# Indeterminate breakdown - text_unmatched is the LLM opportunity ([BPM10] 10.2.1 reason present
# but not captured by term matching)
display(
    df.filter(F.col("necessity_class") == "indeterminate")
      .groupBy("indeterminate_reason")
      .agg(F.count("*").alias("orders"))
)

# Necessity by customer (skipped if no customer column resolved)
if CUSTOMER_COL is not None:
    display(
        df.groupBy(CUSTOMER_COL, "necessity_class")
          .agg(F.count("*").alias("orders"))
          .orderBy(CUSTOMER_COL, "necessity_class")
    )
else:
    print("Skipped necessity-by-customer: no customer column resolved.")

# Transport appropriateness: recommended vs requested level of service ([414.605])
if LOS_COL is not None:
    display(
        df.groupBy(LOS_COL, "recommended_los")
          .agg(F.count("*").alias("orders"))
          .orderBy(F.desc("orders"))
    )
else:
    display(
        df.groupBy("recommended_los")
          .agg(F.count("*").alias("orders"))
          .orderBy(F.desc("orders"))
    )

# 7. Example order texts per category

# For the 24 July action item: sample real free-text per class so stakeholders can see how each
# category is defined. Sampled, de-identified review only - not written back.

_example_cols = [c for c in [ORDER_ID_COL, LOS_COL] if c is not None] + [
    "mobility_score", "monitoring_score", "confidence", "indeterminate_reason", FREE_TEXT_COL
]
for cls in ["clearly_necessary", "indeterminate", "clearly_not_necessary"]:
    print("=" * 70)
    print(cls.upper())
    display(
        df.filter(F.col("necessity_class") == cls)
          .select(*_example_cols)
          .limit(15)
    )

# 8. Export summary workbook - med_nec_buckets_v2.xlsx

# Writes the aggregates that back the slides to a five-tab workbook. This is an OUTPUT FILE, not
# a table write - the source table is never modified. Mirrors the original med_nec_buckets.xlsx,
# updated to the medical-necessity schema (necessity_class, mobility/monitoring, appropriateness).

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT_XLSX = "med_nec_buckets_v2.xlsx"
# Destination directory. Workspace paths (/Workspace/...) are a real mounted filesystem, so the
# workbook is written there directly with pandas/openpyxl - no dbutils.fs.cp needed.
# For a Volume or DBFS destination instead, use e.g.:
#   "/Volumes/prod-sandbox/vivekkumar_patel/exports"   (Unity Catalog volume)
#   "dbfs:/FileStore/med_nec"                            (download via /files/ URL)
OUTPUT_DIR = "/Workspace/Users/josh.smitherman@gmr.net/med_nec/data"

total = df.count()

def with_pct(sdf, count_col="orders"):
    pdf = sdf.toPandas()
    pdf["pct_of_scope"] = (pdf[count_col] / total * 100).round(1)
    return pdf

# --- Summary tab: necessity, confidence, indeterminate reasons ---
nec_pdf = with_pct(
    df.groupBy("necessity_class").agg(F.count("*").alias("orders")).orderBy(F.desc("orders"))
)
conf_pdf = with_pct(
    df.groupBy("confidence").agg(F.count("*").alias("orders")).orderBy(F.desc("orders"))
)
indet_pdf = with_pct(
    df.filter(F.col("necessity_class") == "indeterminate")
      .groupBy("indeterminate_reason").agg(F.count("*").alias("orders")).orderBy(F.desc("orders"))
)

# --- Concepts tab: dictionary with counts and CMS citations (backs the concept slide) ---
concept_sums = df.agg(
    *[F.sum(F.col(f"c_{n}").cast("int")).alias(n) for n, *_ in CONCEPTS]
).collect()[0].asDict()
concept_pdf = pd.DataFrame([
    {"concept": n, "group": g, "axis": a, "weight": w, "basis": b,
     "cms_ref": ref, "source_url": url,
     "tags": int(concept_sums[n] or 0),
     "pct_of_orders": round((concept_sums[n] or 0) / total * 100, 1)}
    for (n, a, w, g, b, ref, url, pat) in CONCEPTS
]).sort_values("tags", ascending=False).reset_index(drop=True)

# --- Where tab: by customer, by level of service, transport appropriateness ---
where_blocks = []
if CUSTOMER_COL is not None:
    where_blocks.append((
        "Necessity by customer",
        df.groupBy(CUSTOMER_COL, "necessity_class").agg(F.count("*").alias("orders"))
          .orderBy(CUSTOMER_COL, "necessity_class").toPandas()
    ))
if LOS_COL is not None:
    where_blocks.append((
        "Necessity by level of service",
        df.groupBy(LOS_COL, "necessity_class").agg(F.count("*").alias("orders"))
          .orderBy(LOS_COL, "necessity_class").toPandas()
    ))
    where_blocks.append((
        "Transport appropriateness (requested vs recommended)",
        df.groupBy(LOS_COL, "recommended_los").agg(F.count("*").alias("orders"))
          .orderBy(F.desc("orders")).toPandas()
    ))

# --- Examples tab: sampled free-text per class ---
ex_cols = [c for c in [ORDER_ID_COL, LOS_COL] if c is not None] + [
    "necessity_class", "total_necessity_score", "mobility_score", "monitoring_score",
    "confidence", "indeterminate_reason", FREE_TEXT_COL,
]
examples_pdf = pd.concat([
    df.filter(F.col("necessity_class") == cls).select(*ex_cols).limit(20).toPandas()
    for cls in ["clearly_necessary", "indeterminate", "clearly_not_necessary"]
], ignore_index=True)

# --- Definitions tab: embedded so the workbook is self-explaining ---
definitions_pdf = pd.DataFrame([
    ["necessity_class = clearly_necessary", "Full bed-confined test met, or a named CMS concept at the clear weight threshold.", "BPM10 10.2.3 / 10.2.1; 414.605"],
    ["necessity_class = clearly_not_necessary", "No Group A concept present - field empty or vague filler only.", "RSN AM600; MLN"],
    ["necessity_class = indeterminate", "Some signal but below the clear threshold, or text present that matched no term.", "BPM10 10.2.1"],
    ["indeterminate_reason = text_unmatched", "Text entered but no term matched - the target for language-model classification.", "BPM10 10.2.1"],
    ["indeterminate_reason = weak_or_inferred_only", "Only inferred or low-weight concepts present.", "BPM10 10.2.1"],
    ["mobility_score", "Weighted sum of mobility-axis concepts - why other transport is contraindicated.", "BPM10 10.2.1 / 10.2.3"],
    ["monitoring_score", "Weighted sum of monitoring-axis concepts - why this level of service.", "414.605"],
    ["total_necessity_score", "mobility_score + monitoring_score + bed-confined bonus. See the Scoring tab.", "composite (provisional)"],
    ["confidence", "Internal quality flag on the classification (high / medium / low). Not a CMS construct.", "n/a"],
    ["recommended_los", "Level of service implied by the monitoring axis. Descriptive, not a billing call.", "414.605; 410.40(c)"],
    ["Group A vs Group B", "A = a clinical reason (12 concepts). B = vague filler that confers no necessity (3).", "BPM10 10.2.1; MLN"],
    ["named vs inferred", "named = explicit in CMS text. inferred = derived from the 10.2.1 general test.", "BPM10 10.2.1"],
    ["Scope", f"Non-emergent ground only. {total:,} orders in scope, 2024 to present.", "BPM10 10.2.1"],
], columns=["term", "definition", "cms_ref"])

# --- Scoring tab: weight table, aggregation formula, and score distribution ---
weight_pdf = pd.DataFrame([
    {"concept": n, "axis": a, "group": g, "basis": b, "weight": w, "cms_ref": ref}
    for (n, a, w, g, b, ref, url, pat) in CONCEPTS
]).sort_values(["axis", "weight"], ascending=[True, False]).reset_index(drop=True)

formula_pdf = pd.DataFrame([
    ["mobility_score", "sum of weights of matched mobility-axis concepts", "BPM10 10.2.1 / 10.2.3"],
    ["monitoring_score", "sum of weights of matched monitoring-axis concepts", "414.605"],
    ["bed_confined bonus", f"+{BED_CONFINED_BONUS} when all three 10.2.3 prongs are present", "BPM10 10.2.3"],
    ["total_necessity_score", "mobility_score + monitoring_score + bed_confined bonus", "composite (provisional)"],
    ["clear threshold", f"mobility >= {MOBILITY_CLEAR} or monitoring >= {MONITORING_CLEAR}, with a named concept", "BPM10 / 414.605"],
    ["note", "weights and bonus are tunable modelling choices, not CMS figures - validate with SMEs", "n/a"],
], columns=["element", "definition", "cms_ref"])

total_dist_pdf = with_pct(
    df.groupBy("total_necessity_score").agg(F.count("*").alias("orders"))
      .orderBy("total_necessity_score")
)

score_by_class_pdf = (
    df.groupBy("necessity_class")
      .agg(F.round(F.avg("total_necessity_score"), 2).alias("avg_total_score"),
           F.min("total_necessity_score").alias("min_score"),
           F.max("total_necessity_score").alias("max_score"),
           F.count("*").alias("orders"))
      .orderBy(F.desc("avg_total_score")).toPandas()
)

sheets = {
    "Definitions": [("Definitions and CMS basis", definitions_pdf)],
    "Summary":     [("Necessity", nec_pdf), ("Confidence", conf_pdf),
                    ("Indeterminate - why", indet_pdf)],
    "Scoring":     [("Concept weights", weight_pdf),
                    ("Aggregation formula", formula_pdf),
                    ("Total score distribution", total_dist_pdf),
                    ("Total score by necessity class", score_by_class_pdf)],
    "Concepts":    [("Concept dictionary with CMS citations", concept_pdf)],
    "Where":       where_blocks or [("No customer/LOS column resolved", pd.DataFrame({"note": ["set FIELD_CANDIDATES"]}))],
    "Examples":    [("Sampled order text per class", examples_pdf)],
}

def _coerce(v):
    if pd.isna(v):
        return None
    return v.item() if hasattr(v, "item") else v

def build_workbook(path, sheets):
    wb = Workbook(); wb.remove(wb.active)
    for name, blocks in sheets.items():
        ws = wb.create_sheet(name[:31]); r = 1
        for title, pdf in blocks:
            if title:
                ws.cell(r, 1, title).font = Font(bold=True); r += 1
            for j, col in enumerate(pdf.columns, 1):
                ws.cell(r, j, str(col)).font = Font(bold=True)
            r += 1
            for _, row in pdf.iterrows():
                for j, val in enumerate(row, 1):
                    ws.cell(r, j, _coerce(val))
                r += 1
            r += 1
        for col_cells in ws.columns:
            w = min(70, max((len(str(c.value)) for c in col_cells if c.value is not None), default=8) + 2)
            ws.column_dimensions[col_cells[0].column_letter].width = w
    wb.save(path)

dest = f"{OUTPUT_DIR.rstrip('/')}/{OUTPUT_XLSX}"
if OUTPUT_DIR.startswith("/Workspace/") or OUTPUT_DIR.startswith("/dbfs/") or OUTPUT_DIR.startswith("/Volumes/"):
    # These are real mounted filesystem paths - write directly, no local hop or cp needed.
    import os
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    build_workbook(dest, sheets)
    print("Saved workbook to:", dest)
else:
    # dbfs:/ or other fs-scheme paths - write locally, then copy with dbutils.
    local_path = f"/tmp/{OUTPUT_XLSX}"
    build_workbook(local_path, sheets)
    try:
        dbutils.fs.mkdirs(OUTPUT_DIR)
        dbutils.fs.cp(f"file:{local_path}", dest)
        print("Saved workbook to:", dest)
        if OUTPUT_DIR.startswith("dbfs:/FileStore"):
            print("Download via: <workspace-url>/files/" + dest.split("/FileStore/", 1)[1])
    except Exception as e:
        print("Wrote local copy:", local_path)
        print("Copy to", dest, "failed - set OUTPUT_DIR to a writable path.")
        print("Reason:", e)

# Notes / open items

# - Weights and thresholds are provisional. They encode the CMS hierarchy (named [BPM10] 10.2.3
#   concepts strongest, inferred [BPM10] 10.2.1 concepts weakest) but should be validated by the
#   SMEs and, once available, against actual denial outcomes in [RSN].
# - text_unmatched is the LLM target - text present that no term captured. That count is the
#   concrete case for language-model classification over term matching.
# - recommended_los is descriptive until CMS BLS/ALS eligibility criteria in [414.605] and
#   [410.40](c) are confirmed; it surfaces requested-vs-recommended mismatches, not billing calls.
# - behavioral, oxygen, isolation, wound_ostomy, bariatric are inferred from [BPM10] 10.2.1 and
#   are the first candidates for SME challenge; behavioral is the weakest.
# - Second free-text field: the review raised expanding beyond ClinicalData. Add fields to the
#   match step only after confirming they are point-of-order per [BPM10] 10.2.4, not post-hoc.

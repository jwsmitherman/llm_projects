# Medical Necessity Scoring - revised schema

# Reframes the non-emergent ground analysis around **medical necessity and transport
# appropriateness** rather than payment or denial. Per the 24 July review:

# - Two scoring axes, not one label: a **mobility** axis (why other transport is
#   contraindicated) and a **monitoring** axis (why this level of service).
# - A single combined score drives the buckets by cutoff: 0 = not_necessary; >= 3 with a
#   named concept = necessary; in between = indeterminate.
# - **Weighted concepts** driven from one config block.
# - No payment-focused terminology in any output column.

# The source table is read only and is never modified. The one write is the derived Genie table
# in section 10, to a schema you control (WRITE_GENIE_TABLE).

# CMS source documents are defined in a separate module, cms_references.py, keyed by the CMS
# source they come from (BPM10, BPM10_10_2_3, CFR_414_605, RSN, MLN, ...). The notebook imports it
# so citations live in one place. In Databricks, keep both files in the same Workspace folder with
# Files enabled; if import is unavailable, run `%run ./cms_references` in a cell first.
from cms_references import ref_url

# 0. Environment

# Written for Databricks (`spark` session assumed present). The `display()` helper below
# lets the same notebook run in a plain Jupyter/VS Code kernel, where it falls back to a
# pandas-style preview. Delete this cell if running only in Databricks.

try:
    display  # provided by Databricks
except NameError:
    def display(sdf):
        try:
            print(sdf.limit(50).toPandas().to_string(index=False))
        except AttributeError:
            print(sdf)

# 1. Source fields

# The source table is a read-only snapshot. Each field below is cited to the CMS document
# that makes it relevant to the medical necessity determination.

# The catalog name contains a hyphen, so each identifier part must be back-quoted for
# Spark SQL. SOURCE_TABLE keeps the plain name; SOURCE_TABLE_SQL is the quoted form used
# when reading. (Unquoted `prod-sandbox` raises INVALID_IDENTIFIER / SQLSTATE 42602.)
SOURCE_TABLE     = "prod-sandbox.vivekkumar_patel.temp_tnet_tripmaster"
SOURCE_TABLE_SQL = ".".join(f"`{p}`" for p in SOURCE_TABLE.split("."))

# Logical fields mapped to candidate column names. The first candidate present in the table
# wins. This avoids INVALID_IDENTIFIER / UNRESOLVED_COLUMN when the real schema differs.
# Edit the candidate lists if the actual column names are not covered.
FIELD_CANDIDATES = {
    "free_text": ["ClinicalData", "clinical_data", "ClinicalNotes", "Reason", "ReasonForTransport"],
    "los":       ["LevelOfService", "level_of_service", "LOS", "ServiceLevel"],
    "customer":  ["Customer", "CustomerName", "HealthSystem", "Facility", "Account"],
    "order_id":  ["OrderId", "order_id", "OrderID", "TripId", "TripID", "TripLegId"],
}

# 42 CFR / BPM10 citations for the logical fields (see resolution below):
#   free_text -> [BPM10] 10.2.1, 10.2.4 ; [RSN] AM600   (reason must be recorded)
#   los       -> [414.605] ; [410.40](c)                (level billed must be necessary)
#   customer / order_id -> operational only, no CMS basis

# 2. Concept dictionary - the single source of truth

# Every concept carries its axis, weight, group, CMS basis, source tag, source URL, and the
# term pattern matched against ClinicalData. The url column makes the citation explicit next to
# the terms it governs. Weights are 0-3, reflecting how directly the concept establishes the
# criterion (not clinical severity).

#   axis: "mobility"   = transport necessity   [BPM10] 10.2.1 / 10.2.3  (why not a wheelchair van/car)
#         "monitoring" = level of service      [414.605]               (why ALS / CCT rather than BLS)
#         "filler"     = no necessity alone     [MLN] / [BPM10] 10.2.1
#   basis: "named"    = concept is explicit in the cited CMS text
#          "inferred" = derived from the [BPM10] 10.2.1 general test, not separately named

# Concept source URLs, resolved from the reference module (keyed by CMS source).
BPM10   = ref_url("BPM10")
CFR414  = ref_url("CFR_414_605")
MLN     = ref_url("MLN")

# fields: (concept, axis, weight, group, basis, cms_ref, source_url, term_pattern)
CONCEPTS = [
    # --- mobility axis: transport necessity, BPM10 10.2.3 bed-confined test (prongs) + 10.2.1 ---
    # bed_confined: [BPM10] 10.2.3 - "unable to get up from bed without assistance". One prong.
    ("bed_confined",     "mobility",     3, "A", "named",    "BPM10 10.2.3 (prong: get up)",       BPM10,
        r"bed ?(bound|confined)|unable to get up|cannot get out of bed"),
    # mobility_deficit: [BPM10] 10.2.3 - "unable to ambulate" (prong 2); also 10.2.1 contraindication.
    ("mobility_deficit", "mobility",     3, "A", "named",    "BPM10 10.2.3 (prong: ambulate)",     BPM10,
        r"hemipar|hemipleg|paraly|non ?ambulat|unable to (bear weight|ambulate|walk|stand)|fracture|amputat|contracture|bear weight"),
    # cannot_sit: [BPM10] 10.2.3 - "unable to sit in a chair or wheelchair" (prong 3).
    ("cannot_sit",       "mobility",     3, "A", "named",    "BPM10 10.2.3 (prong: sit)",          BPM10,
        r"cannot sit|unable to sit|special positioning|supine|must lie|stretcher|cannot support trunk"),
    # bariatric: not named by CMS; inferred handling requirement under [BPM10] 10.2.1 general test.
    ("bariatric",        "mobility",     2, "A", "inferred", "BPM10 10.2.1 (inferred: handling)",  BPM10,
        r"bariatric|morbid"),
    # wound_ostomy: not named; inferred positioning need under [BPM10] 10.2.1.
    ("wound_ostomy",     "mobility",     1, "A", "inferred", "BPM10 10.2.1 (inferred: position)",  BPM10,
        r"wound|ostomy|ulcer|decubitus|drain"),
    # behavioral: not named; inferred safety need under [BPM10] 10.2.1. Weakest - CMS wants a
    #   physical limitation, so this is a candidate to move to Group B on SME review.
    ("behavioral",       "mobility",     1, "A", "inferred", "BPM10 10.2.1 (inferred: safety)",    BPM10,
        r"dementia|alzheimer|combative|agitat|altered mental|flight risk|elope"),

    # --- monitoring axis: level of service, 42 CFR 414.605 definitions ---
    # ventilator: [414.605] ALS2 (airway management) / SCT (respiratory care) territory.
    ("ventilator",       "monitoring",   3, "A", "named",    "414.605 (ALS2 airway / SCT)",        CFR414,
        r"ventilat|vent|trach|intubat"),
    # suctioning: [414.605] SCT - care beyond the EMT-Paramedic scope of practice.
    ("suctioning",       "monitoring",   3, "A", "named",    "414.605 (SCT)",                      CFR414,
        r"suction"),
    # iv_medication: [414.605] ALS2 - administration of 3+ IV medications / central line.
    ("iv_medication",    "monitoring",   3, "A", "named",    "414.605 (ALS2 IV meds)",             CFR414,
        r"\biv\b|infusion|drip|heparin|antibiotic|tpn"),
    # cardiac: [414.605] ALS assessment / ALS2 cardiac procedures (monitoring, defib, pacing).
    ("cardiac",          "monitoring",   2, "A", "named",    "414.605 (ALS assessment)",           CFR414,
        r"cardiac|telemetry|ekg|ecg|nstemi|stemi|arrhythm|afib"),
    # oxygen: not a named necessity concept; [BPM10] 10.2.1 - chronic O2 alone does not qualify.
    ("oxygen",           "monitoring",   1, "A", "inferred", "BPM10 10.2.1 (O2 alone insufficient)", BPM10,
        r"oxygen|\bo2\b|lpm|nasal cannula|bipap|cpap"),
    # isolation: not named; inferred under [BPM10] 10.2.1 - infection control, not a patient contraindication.
    ("isolation",        "monitoring",   1, "A", "inferred", "BPM10 10.2.1 (infection control)",   BPM10,
        r"isolation|mrsa|c\.? ?diff|precaution"),

    # --- filler: appears in text but establishes no necessity on its own ---
    # weakness_only: [MLN] / MAC guidance - "severe generalized weakness" counts only with a
    #   qualifier and functional consequence; the bare term is vague and of little value on review.
    ("weakness_only",    "filler",       0, "B", "named",    "MLN / MAC (vague weakness)",         MLN,
        r"general(iz)?e?d? weakness|generally weak|^\s*weak"),
    # fall_risk_only: [MLN] / MAC guidance - a risk is not a contraindication to other transport.
    ("fall_risk_only",   "filler",       0, "B", "named",    "MLN / MAC (risk != contra)",         MLN,
        r"fall risk|unsteady|deconditio"),
    # nonclinical: [BPM10] 10.2.1 / [410.40](d) - a physician order alone does not prove necessity.
    ("nonclinical",      "filler",       0, "B", "named",    "BPM10 10.2.1 (order != necessity)",  BPM10,
        r"per protocol|convenience|no other transport|unable to arrange|family request"),
]

# Thresholds - tune here. A single named, weight-3 concept (score 3) is a clear reason.
NECESSARY_CUTOFF  = 3   # total_score at/above this, with a named concept, = necessary

# 3. Scope - non-emergent ground only

# Rideshare, air, and emergent trips are excluded; they do not carry the [BPM10] 10.2.1
# non-emergency documentation requirement. Adjust the filter to the real column values.

from pyspark.sql import functions as F

df = spark.table(SOURCE_TABLE_SQL)

# Resolve each logical field to a real column. First candidate present wins; unresolved fields
# come back as None and any output depending on them is skipped rather than erroring.
_cols_lower = {c.lower(): c for c in df.columns}
def resolve(field):
    for cand in FIELD_CANDIDATES[field]:
        if cand.lower() in _cols_lower:
            return _cols_lower[cand.lower()]
    return None

FREE_TEXT_COL = resolve("free_text")
LOS_COL       = resolve("los")
CUSTOMER_COL  = resolve("customer")
ORDER_ID_COL  = resolve("order_id")

print("Resolved columns:")
print(f"  free_text -> {FREE_TEXT_COL}")
print(f"  los       -> {LOS_COL}")
print(f"  customer  -> {CUSTOMER_COL}")
print(f"  order_id  -> {ORDER_ID_COL}")
if FREE_TEXT_COL is None:
    raise ValueError(
        "No free-text column found. Add the real column name to "
        "FIELD_CANDIDATES['free_text']. Available columns: " + ", ".join(df.columns)
    )

# Scope: non-emergent GROUND AMBULANCE only. Exclusions are explicit and configurable below,
# because the real level-of-service field uses short codes (BLS, ALS, WC, EMG, FWQUOTE, ...) that
# a generic pattern does not reliably catch. The distinct-value print confirms them against data.

raw_count = df.count()

# Show what level-of-service values actually exist, so scope is verified rather than assumed.
if LOS_COL is not None:
    print("Distinct " + LOS_COL + " values (top 40 by count):")
    df.groupBy(LOS_COL).count().orderBy(F.desc("count")).show(40, truncate=False)

# --- Exclusion 1: air / non-ground. Not subject to the ground ambulance necessity test. ---
NON_GROUND_LOS = ["FWQUOTE", "ORGAN"]          # fixed-wing quote, organ transport
NON_GROUND_PATTERN = r"air|fixed ?wing|rotor|helicopter|flight|rideshare|uber|lyft|livery|taxi"

# --- Exclusion 2: emergent. Outside the NON-emergent scope ([BPM10] 10.2.1). ---
# Catches standalone EMG and the -EMG suffix (ALS-EMG, CCT-EMG), plus spelled-out emergent.
EMERGENT_LOS_PATTERN = r"(^|[-_ ])emg($|[-_ ])|emergen"

# --- Exclusion 3: non-ambulance ground. A wheelchair van or ambulatory transport is the cheaper
# alternative, so ambulance necessity concepts do not apply and produce all-indeterminate noise.
# Excluded by default. Set INCLUDE_NON_AMBULANCE = True to keep them in scope. ---
NON_AMBULANCE_LOS = ["WC", "AMBLTY"]           # wheelchair van, ambulatory
INCLUDE_NON_AMBULANCE = False

# --- Unidentified codes: kept in scope and flagged. Identify with the business before treating
# the headline as final - e.g. ST (large, all-indeterminate) and CC_CF_C5. ---
REVIEW_LOS = ["ST", "CC_CF_C5"]

if LOS_COL is not None:
    los_l = F.lower(F.coalesce(F.col(LOS_COL), F.lit("")))
    exclude_exact = [c.lower() for c in NON_GROUND_LOS]
    if not INCLUDE_NON_AMBULANCE:
        exclude_exact += [c.lower() for c in NON_AMBULANCE_LOS]
    df = df.filter(~los_l.isin(exclude_exact))
    df = df.filter(~los_l.rlike(NON_GROUND_PATTERN))
    df = df.filter(~los_l.rlike(EMERGENT_LOS_PATTERN))
    # Flag any review codes still present, so they are not silently trusted.
    present_review = [c for c in REVIEW_LOS
                      if df.filter(los_l == c.lower()).limit(1).count() > 0]
    if present_review:
        print("REVIEW: unidentified level-of-service codes still in scope - "
              "identify before finalising: " + ", ".join(present_review))
else:
    print("WARNING: no level-of-service column resolved - scope exclusions NOT applied.")

# Also remove emergent trips via a trip-type / priority flag if one is present.
EMERGENT_PATTERN = r"emergen|911|lights|code ?3"
for trip_type_col in ("TripType", "trip_type", "TransportType", "Priority", "CallType"):
    if trip_type_col in df.columns:
        df = df.filter(~F.lower(F.coalesce(F.col(trip_type_col), F.lit(""))).rlike(EMERGENT_PATTERN))
        print(f"Applied emergent exclusion on column: {trip_type_col}")
        break
else:
    print("NOTE: no trip-type column found - emergent exclusion relies on the LOS codes above.")

df = df.withColumn("_text", F.lower(F.coalesce(F.col(FREE_TEXT_COL), F.lit(""))))
df = df.withColumn("_has_text", F.length(F.trim(F.col("_text"))) > 0)

scope_count = df.count()
has_text_count = df.filter(F.col("_has_text")).count()
print(f"\nRaw rows:            {raw_count:,}")
print(f"In scope:            {scope_count:,}  ({scope_count/raw_count*100:.1f}% of raw)")
print(f"  with free text:    {has_text_count:,}  ({has_text_count/max(scope_count,1)*100:.1f}% of scope)")
print("Excluded: air/non-ground (" + ", ".join(NON_GROUND_LOS) + " + patterns), emergent (EMG), "
      + ("non-ambulance (" + ", ".join(NON_AMBULANCE_LOS) + ")" if not INCLUDE_NON_AMBULANCE else "non-ambulance KEPT"))

# Guardrail - catch a scope that collapses to too few rows.
if scope_count < 0.2 * raw_count:
    print("\nWARNING: scope kept under 20% of rows. Check the distinct level-of-service values "
          "above - an exclusion may be dropping valid ground rows, or the wrong column resolved.")
if has_text_count < 0.5 * max(scope_count, 1):
    print("\nWARNING: over half of in-scope orders have no free text. Confirm FREE_TEXT_COL "
          "resolved to the real reason-for-transport field (see the 'Resolved columns' output).")
    print("Sample of the resolved free-text field:")
    df.select(FREE_TEXT_COL).filter(F.col("_has_text")).show(5, truncate=80)

# 4. Concept tagging

# One boolean column per concept, the term pattern matched against ClinicalData with `rlike`.
# The CMS citation for each pattern is in the CONCEPTS table above.

for name, axis, weight, group, basis, ref, url, pattern in CONCEPTS:
    df = df.withColumn(f"c_{name}", F.col("_text").rlike(pattern) & F.col("_has_text"))

# 5. Scores, classification, and level-of-service recommendation

# All derived from the config. The bucket is driven by one combined score plus one named flag.

def axis_score(axis):
    terms = [F.when(F.col(f"c_{n}"), F.lit(w)).otherwise(F.lit(0))
             for n, a, w, *_ in CONCEPTS if a == axis]
    expr = terms[0]
    for t in terms[1:]:
        expr = expr + t
    return expr

# Component scores (shown so the total is fully traceable). mobility + monitoring = total_score
# exactly - there is no bonus or adjustment.
df = df.withColumn("mobility_score",   axis_score("mobility"))    # [BPM10] 10.2.1 / 10.2.3
df = df.withColumn("monitoring_score", axis_score("monitoring"))  # [414.605]
df = df.withColumn("total_score", F.col("mobility_score") + F.col("monitoring_score"))

# Per-concept weighted contribution: w_<concept> = the concept's weight when it matched, else 0.
# The 1/0 concept flags alone cannot be added up to reach the scores, because each concept carries
# a different weight that otherwise exists only in the CONCEPTS block above. These columns make the
# arithmetic reconcile from the data itself:
#     sum(w_* over mobility-axis concepts)   = mobility_score
#     sum(w_* over monitoring-axis concepts) = monitoring_score
#     mobility_score + monitoring_score      = total_score
# Filler concepts carry weight 0 by design and contribute nothing, so they get no w_ column - their
# 1/0 flag is kept because the presence of filler wording is itself informative.
SCORING_CONCEPTS = [(n, a, w) for n, a, w, *_ in CONCEPTS if a in ("mobility", "monitoring")]
for _n, _a, _w in SCORING_CONCEPTS:
    df = df.withColumn(f"w_{_n}",
                       F.when(F.col(f"c_{_n}"), F.lit(_w)).otherwise(F.lit(0)).cast("int"))

# named_score: points from NAMED CMS concepts only (those explicit in CMS text). Shown because a
# named concept is required to reach the necessary band - this column is what drives that.
# IMPORTANT: named_score is a SUBSET of total_score, not a third component. The named concepts are
# already counted inside mobility_score and monitoring_score. Adding named_score to those will
# double-count and is the most likely reason a hand-checked total appears not to reconcile.
named_a = [n for n, a, w, g, b, *_ in CONCEPTS if g == "A" and b == "named"]
named_terms = [F.when(F.col(f"c_{n}"), F.lit(w)).otherwise(F.lit(0))
               for n, a, w, g, b, *_ in CONCEPTS if g == "A" and b == "named"]
named_expr = named_terms[0]
for t in named_terms[1:]:
    named_expr = named_expr + t
df = df.withColumn("named_score", named_expr)
df = df.withColumn("named_a_hits", sum([F.col(f"c_{n}").cast("int") for n in named_a]))
df = df.withColumn("has_named_concept", (F.col("named_a_hits") > 0).cast("int"))

# Reconciliation check: confirm the visible w_ columns actually sum to the scores, and that the
# two axis scores sum to the total. Prints a count of any row that fails, so a scoring change that
# breaks the arithmetic is caught here rather than by someone hand-adding columns later.
_mob_w = [f"w_{n}" for n, a, w in SCORING_CONCEPTS if a == "mobility"]
_mon_w = [f"w_{n}" for n, a, w in SCORING_CONCEPTS if a == "monitoring"]
_recon = (df
    .withColumn("_mob_sum", sum([F.col(c) for c in _mob_w]))
    .withColumn("_mon_sum", sum([F.col(c) for c in _mon_w]))
    .withColumn("_bad",
        (F.col("_mob_sum") != F.col("mobility_score"))
        | (F.col("_mon_sum") != F.col("monitoring_score"))
        | ((F.col("mobility_score") + F.col("monitoring_score")) != F.col("total_score")))
)
_bad_rows = _recon.filter(F.col("_bad")).count()
print(f"Score reconciliation: {_bad_rows:,} rows fail (expected 0).")
print("  sum(w_* mobility) = mobility_score; sum(w_* monitoring) = monitoring_score; "
      "mobility_score + monitoring_score = total_score")
print("  named_score is a SUBSET of total_score, not a third component - do not add it in.")
if _bad_rows:
    display(_recon.filter(F.col("_bad"))
                  .select(*_mob_w, "_mob_sum", "mobility_score",
                          *_mon_w, "_mon_sum", "monitoring_score", "total_score")
                  .limit(20))

# unmatched_text: 1 when text was entered but nothing scored (words no rule recognized). Purely
# informational - it does NOT change the bucket - and flags the language-model target set.
filler_cols = [f"c_{n}" for n, a, *_ in CONCEPTS if a == "filler"]
df = df.withColumn("any_filler", sum([F.col(c).cast("int") for c in filler_cols]) > 0)
df = df.withColumn(
    "unmatched_text",
    (F.col("_has_text") & (F.col("total_score") == 0) & ~F.col("any_filler")).cast("int"),
)

# necessity_class: driven by total_score and the named flag. Simple cutoffs:
#   total_score >= NECESSARY_CUTOFF and a named concept present -> necessary
#   total_score == 0                                            -> not_necessary
#   otherwise                                                   -> indeterminate
df = df.withColumn(
    "necessity_class",
    F.when((F.col("total_score") >= NECESSARY_CUTOFF) & (F.col("has_named_concept") == 1),
           F.lit("necessary"))
     .when(F.col("total_score") == 0, F.lit("not_necessary"))
     .otherwise(F.lit("indeterminate")),
)

# recommended_los: level of service implied by the monitoring axis, per [414.605] definitions.
#   Descriptive only until CMS BLS/ALS eligibility criteria are confirmed - not a billing call.
df = df.withColumn(
    "recommended_los",
    F.when(F.col("c_ventilator") | F.col("c_suctioning"), F.lit("CCT/SCT"))            # [414.605] SCT/ALS2
     .when((F.col("monitoring_score") >= 2) | F.col("c_iv_medication"), F.lit("ALS"))  # [414.605] ALS1/ALS2
     .when(F.col("mobility_score") >= 1, F.lit("BLS"))                                 # [414.605] BLS
     .otherwise(F.lit("none/indeterminate")),
)

# gy_disposition: relates the class to the GY-modifier process (4 Aug review-process meeting).
# not_necessary = no documented reason (a GY candidate); necessary = documented reason present;
# indeterminate = partial / review. Describes the order text at order time, not a billing call.
df = df.withColumn(
    "gy_disposition",
    F.when(F.col("necessity_class") == "not_necessary", F.lit("no documented reason - GY candidate"))
     .when(F.col("necessity_class") == "necessary", F.lit("documented reason present"))
     .otherwise(F.lit("partial - review")),
)

# classification_method: how the labels were assigned. Keyword/term matching today; the
# unmatched_text orders are the set a language model would categorise in a later pass.
df = df.withColumn("classification_method", F.lit("keyword_match"))

# why_labeled: for each order, the concepts that matched and the exact text that triggered each -
# e.g. "mobility_deficit[unable to ambulate]; oxygen[o2]". This is the reviewer's view of HOW an
# order was labeled: the actual words behind each concept. Empty when nothing matched (the
# text_unmatched case), which is itself the signal that the rules found nothing to go on.
_match_exprs = []
for _n, _a, _w, _g, _b, _ref, _url, _pat in CONCEPTS:
    _m = F.regexp_extract(F.col("_text"), _pat, 0)
    _match_exprs.append(
        F.when(F.col(f"c_{_n}"), F.concat(F.lit(_n + "["), _m, F.lit("]"))).otherwise(F.lit(None))
    )
df = df.withColumn("why_labeled", F.concat_ws("; ", F.array(*_match_exprs)))

# 6. Summary outputs (display only - nothing is written)

# Necessity distribution
display(
    df.groupBy("necessity_class")
      .agg(F.count("*").alias("orders"))
      .orderBy(F.desc("orders"))
)

# Unmatched-text breakdown - words present that no rule scored (the language-model opportunity).
display(
    df.groupBy("unmatched_text").agg(F.count("*").alias("orders")).orderBy("unmatched_text")
)

# Necessity by customer (skipped if no customer column resolved)
if CUSTOMER_COL is not None:
    display(
        df.groupBy(CUSTOMER_COL, "necessity_class")
          .agg(F.count("*").alias("orders"))
          .orderBy(CUSTOMER_COL, "necessity_class")
    )
else:
    print("Skipped necessity-by-customer: no customer column resolved.")

# Transport appropriateness: recommended vs requested level of service ([414.605])
if LOS_COL is not None:
    display(
        df.groupBy(LOS_COL, "recommended_los")
          .agg(F.count("*").alias("orders"))
          .orderBy(F.desc("orders"))
    )
else:
    display(
        df.groupBy("recommended_los")
          .agg(F.count("*").alias("orders"))
          .orderBy(F.desc("orders"))
    )

# 7. Example order texts per category

# For the 24 July action item: sample real free-text per class so stakeholders can see how each
# category is defined. Sampled, de-identified review only - not written back.

_example_cols = [c for c in [ORDER_ID_COL, LOS_COL] if c is not None] + [
    "total_score", "mobility_score", "monitoring_score", "named_score", "has_named_concept", "unmatched_text", FREE_TEXT_COL
]
for cls in ["necessary", "indeterminate", "not_necessary"]:
    print("=" * 70)
    print(cls.upper())
    display(
        df.filter(F.col("necessity_class") == cls)
          .select(*_example_cols)
          .limit(15)
    )
# 8. Output helpers - shared by the LLM section and every export below
# (numbering: 8 helpers, 9 LLM, 10 buckets workbook, 11 review workbook, 12 Genie table)

# Excel-safe string cleaning, the run timestamp, and the pandas import. Defined once here because
# the LLM section and all three exports use them.

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import re as _re

# openpyxl rejects ASCII control characters (except tab/newline/carriage return) with
# IllegalCharacterError. Real order text can contain them (stray glyphs from copy/paste, form
# artifacts like the character before "L1 kyphoplasty"), so strip them from every string cell
# before any Excel write. Applied to workbook builders and to the review/CSV frames.
_ILLEGAL_XLSX = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def _clean(v):
    if isinstance(v, str):
        return _ILLEGAL_XLSX.sub("", v)
    return v

def sanitize_df(pdf):
    import pandas.api.types as _pt
    pdf = pdf.copy()
    for c in pdf.columns:
        # Clean any column that is not purely numeric/bool (dtype may be 'object' or 'string').
        if not (_pt.is_numeric_dtype(pdf[c]) or _pt.is_bool_dtype(pdf[c])):
            pdf[c] = pdf[c].map(_clean)
    return pdf

# Run timestamp (YYYYMMDD_HHMMSS) appended to every output file so runs stay grouped and do not
# overwrite each other. Set once here and reused for the workbook, detail CSV, and review workbook.
RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")
# 9. LLM determination - the second, independent read of the same order text

# There are two determinations per order, and they are kept side by side rather than merged:
#   - the SCORE determination (necessity_class), from the weighted concept rules above
#   - the LLM determination (necessity_class_llm), from this section
# Both use the SAME cutoffs, so any difference comes from what was read out of the text, not from
# different thresholds. The LLM path additionally records WHY: a plain-English rationale plus a
# verbatim evidence quote for every fact it asserts.
#
# Pattern is the NurseNav one (OpenAI SDK -> Databricks serving endpoint, extract-then-judge): the
# model extracts facts only, and the same deterministic rules engine assigns the label. The model
# never picks the label itself, which is what keeps the two paths comparable and auditable.
#
# OFF by default - it calls a paid endpoint. Sample first, check the valid-extraction rate, then
# widen. This section runs BEFORE the outputs so both determinations land in the workbook and table.

RUN_LLM = False

# Which orders the LLM reads:
#   "unmatched"  - only orders where text was entered but no rule scored it (the rules' blind spot)
#   "with_text"  - every order that has free text (full rule-vs-LLM comparison)
#   "all"        - every in-scope order, including empty documentation
# "unmatched" is the cheapest useful run. "with_text" is the real comparison but is a large job:
# at ~323k orders this is a batch workload, not a notebook loop - sample first and size it.
LLM_TARGET   = "unmatched"
LLM_SAMPLE_N = 200        # cap while validating the prompt; None = the whole target set
LLM_WORKERS  = 8          # parallel requests; raise carefully, the endpoint is shared
LLM_MODEL    = "databricks-gpt-oss-120b"
LLM_BASE_URL = "https://adb-2790612761746757.17.azuredatabricks.net/serving-endpoints"

# Columns the LLM path contributes. Defined outside the RUN_LLM block so the downstream outputs
# have a stable schema whether or not the LLM ran - unrun orders carry nulls and llm_status="not_run".
LLM_OUTPUT_COLS = ["necessity_class_llm", "mobility_score_llm", "monitoring_score_llm",
                   "total_score_llm", "has_named_concept_llm", "llm_rationale",
                   "llm_evidence", "llm_summary", "llm_status", "determination_agreement"]

llm_pdf = None

if RUN_LLM:
    import json as _json
    import os
    from concurrent.futures import ThreadPoolExecutor
    from openai import OpenAI

    _token = (
        dbutils.notebook.entry_point.getDbutils().notebook().getContext().apiToken().get()
        if "dbutils" in dir() else os.environ.get("DATABRICKS_TOKEN", "")
    )
    _client = OpenAI(api_key=_token, base_url=LLM_BASE_URL)

    def llm_call(system_prompt, user_prompt, max_tokens=1500):
        resp = _client.chat.completions.create(
            model=LLM_MODEL,
            messages=[{"role": "system", "content": system_prompt},
                      {"role": "user",   "content": user_prompt}],
            temperature=0.1, max_tokens=max_tokens,
        )
        content = resp.choices[0].message.content
        if isinstance(content, list):
            for item in content:
                if isinstance(item, dict) and item.get("type") == "text":
                    return item.get("text", "")
            return _json.dumps(content)
        return content

    # --- Extraction prompt: the model pulls facts and explains them; the rules assign the label ---
    MED_NEC_SYSTEM_PROMPT = """
You are a clinical documentation information extraction assistant for non-emergent ground ambulance orders.
Your ONLY task is to extract structured facts from the free-text transport reason recorded at order time,
and to explain in plain English what the text does and does not establish.

CRITICAL CONSTRAINTS
- Do NOT provide medical advice, diagnose, or decide whether the transport was appropriate.
- Do NOT decide the necessity label. A separate rules engine does that from your extracted facts.
- Do NOT guess. Only extract what is explicitly supported by the text.
- Handle negations correctly (for example "no oxygen" is false).
- For every boolean set to true you MUST provide a verbatim evidence quote from the text.
- If no clinical reason is documented, set documentation.reason_documented = false.

ABBREVIATIONS: BLS/ALS=basic/advanced life support, O2=oxygen, IV=intravenous, w/c=wheelchair,
amb=ambulate/ambulatory, ETT=endotracheal tube, trach=tracheostomy.

OUTPUT: valid JSON only, all keys present, booleans never null.

SCHEMA:
{
  "order_id": string,
  "text_summary": string,
  "rationale": string,
  "mobility": {
    "bed_confined": boolean, "cannot_ambulate": boolean, "cannot_sit": boolean,
    "bariatric_handling": boolean, "wound_or_ostomy_positioning": boolean,
    "behavioral_or_altered_mental": boolean
  },
  "monitoring": {
    "ventilator_or_airway": boolean, "airway_suctioning": boolean, "iv_medication": boolean,
    "cardiac_monitoring": boolean, "oxygen": boolean, "isolation_precautions": boolean
  },
  "filler_only": {
    "weakness_only": boolean, "fall_risk_only": boolean, "nonclinical_reason": boolean
  },
  "documentation": { "reason_documented": boolean, "note_is_operational_only": boolean },
  "evidence": [ {"field": string, "value": boolean, "quote": string} ]
}

FIELD NOTES
- mobility captures why the patient cannot travel by wheelchair van or car.
- monitoring captures clinical care needed during transport (drives level of service).
- filler_only captures vague terms that carry no necessity on their own.
- Set reason_documented = false when the field is empty or contains only filler.
- rationale: two or three sentences, plain English, aimed at a reviewer who has not read the text.
  State what the text establishes about mobility and about monitoring, and name what is absent or
  vague. Refer only to the text in front of you. Do not state or imply a necessity label, a billing
  outcome, or a recommendation. Example shape: "The text records an inability to bear weight, which
  speaks to mobility. Nothing is recorded about monitoring needs during transport. The reference to
  general weakness is not qualified by any functional consequence."

Now extract from the following order text."""

    def build_user_prompt(order_id, text):
        return f"order_id: {order_id}\n\nOrder text:\n{text}"

    # --- Validation: strict JSON, all sections present, every true flag carries evidence ---
    BOOL_SECTIONS = {
        "mobility": ["bed_confined", "cannot_ambulate", "cannot_sit", "bariatric_handling",
                     "wound_or_ostomy_positioning", "behavioral_or_altered_mental"],
        "monitoring": ["ventilator_or_airway", "airway_suctioning", "iv_medication",
                       "cardiac_monitoring", "oxygen", "isolation_precautions"],
        "filler_only": ["weakness_only", "fall_risk_only", "nonclinical_reason"],
    }
    REQUIRED_TOP_KEYS = {"order_id", "text_summary", "rationale", "mobility", "monitoring",
                         "filler_only", "documentation", "evidence"}
    # Named concepts (explicit in CMS text) - only these can carry an order to necessary.
    NAMED_FIELDS = {
        "mobility.bed_confined", "mobility.cannot_ambulate", "mobility.cannot_sit",
        "monitoring.ventilator_or_airway", "monitoring.airway_suctioning",
        "monitoring.iv_medication", "monitoring.cardiac_monitoring",
    }
    # Same weights as the CONCEPTS block, keyed to the LLM's field names.
    LLM_WEIGHTS = {
        "mobility.bed_confined": 3, "mobility.cannot_ambulate": 3, "mobility.cannot_sit": 3,
        "mobility.bariatric_handling": 2, "mobility.wound_or_ostomy_positioning": 1,
        "mobility.behavioral_or_altered_mental": 1,
        "monitoring.ventilator_or_airway": 3, "monitoring.airway_suctioning": 3,
        "monitoring.iv_medication": 3, "monitoring.cardiac_monitoring": 2,
        "monitoring.oxygen": 1, "monitoring.isolation_precautions": 1,
    }

    def validate_extraction(raw):
        try:
            obj = _json.loads(raw)
        except Exception as e:
            return False, {}, f"invalid JSON: {e}"
        missing = REQUIRED_TOP_KEYS - set(obj)
        if missing:
            return False, obj, f"missing keys: {sorted(missing)}"
        if not isinstance(obj.get("rationale"), str) or not obj["rationale"].strip():
            return False, obj, "rationale missing or empty"
        for section, keys in BOOL_SECTIONS.items():
            sec = obj.get(section, {})
            if not isinstance(sec, dict):
                return False, obj, f"{section} must be an object"
            for k in keys:
                if not isinstance(sec.get(k), bool):
                    return False, obj, f"{section}.{k} must be boolean"
        ev_fields = {e.get("field") for e in obj.get("evidence", []) if isinstance(e, dict)}
        for section, keys in BOOL_SECTIONS.items():
            for k in keys:
                if obj[section][k] and f"{section}.{k}" not in ev_fields:
                    return False, obj, f"true flag without evidence: {section}.{k}"
        return True, obj, "OK"

    # --- Rules engine: the SAME cutoffs as the score path, applied to the extracted facts ---
    def judge(obj):
        mob = sum(LLM_WEIGHTS[f"mobility.{k}"] for k in BOOL_SECTIONS["mobility"] if obj["mobility"][k])
        mon = sum(LLM_WEIGHTS[f"monitoring.{k}"] for k in BOOL_SECTIONS["monitoring"] if obj["monitoring"][k])
        total = mob + mon
        has_named = any(
            obj[sec][k] for sec in ("mobility", "monitoring") for k in BOOL_SECTIONS[sec]
            if f"{sec}.{k}" in NAMED_FIELDS
        )
        if total >= NECESSARY_CUTOFF and has_named:
            cls = "necessary"
        elif total == 0:
            cls = "not_necessary"
        else:
            cls = "indeterminate"
        return {"necessity_class_llm": cls, "mobility_score_llm": mob,
                "monitoring_score_llm": mon, "total_score_llm": total,
                "has_named_concept_llm": int(has_named)}

    # --- Select the target set ---
    id_col = ORDER_ID_COL or "OrderId"
    if LLM_TARGET == "unmatched":
        target = df.filter(F.col("unmatched_text") == 1)
    elif LLM_TARGET == "with_text":
        target = df.filter(F.col("_has_text"))
    elif LLM_TARGET == "all":
        target = df
    else:
        raise ValueError(f"LLM_TARGET must be unmatched / with_text / all, got {LLM_TARGET!r}")

    target_pdf = (target
        .select(F.col(id_col).alias("_order_id"), F.col(FREE_TEXT_COL).alias("_text"))
        .toPandas())
    target_total = len(target_pdf)
    if LLM_SAMPLE_N:
        target_pdf = target_pdf.head(LLM_SAMPLE_N)
    print(f"LLM determination: {len(target_pdf):,} of {target_total:,} orders in the "
          f"'{LLM_TARGET}' target set, via {LLM_MODEL}")
    if LLM_SAMPLE_N and target_total > LLM_SAMPLE_N:
        print(f"  SAMPLED - {target_total - LLM_SAMPLE_N:,} target orders carry llm_status='not_run'.")

    def process_one(rec):
        oid, text = rec
        out = {"_order_id": oid}
        try:
            raw = llm_call(MED_NEC_SYSTEM_PROMPT, build_user_prompt(str(oid), text or ""))
            ok, obj, msg = validate_extraction(raw)
        except Exception as e:
            ok, obj, msg = False, {}, f"call failed: {e}"
        if ok:
            out.update(judge(obj))
            out["llm_rationale"] = obj.get("rationale", "")
            out["llm_summary"] = obj.get("text_summary", "")
            out["llm_evidence"] = "; ".join(
                f'{e.get("field")}[{e.get("quote")}]' for e in obj.get("evidence", [])
                if isinstance(e, dict)
            )
            out["llm_status"] = "ok"
        else:
            out["llm_status"] = f"error: {msg}"
        return out

    records = list(zip(target_pdf["_order_id"], target_pdf["_text"]))
    with ThreadPoolExecutor(max_workers=LLM_WORKERS) as pool:
        rows = list(pool.map(process_one, records))

    llm_pdf = pd.DataFrame(rows)
    valid_rate = (llm_pdf["llm_status"] == "ok").mean() if len(llm_pdf) else 0.0
    print(f"Valid extractions: {valid_rate:.1%}  (target >= 95% before widening LLM_TARGET)")
    if valid_rate < 0.95 and len(llm_pdf):
        print("  Below target - review llm_status before scaling up:")
        print(llm_pdf.loc[llm_pdf.llm_status != "ok", "llm_status"].value_counts().head(10))

# --- Attach the LLM determination to the scored data ---
# One row per order carrying BOTH determinations. Orders the LLM did not read keep the score
# determination and carry llm_status = "not_run", so the two columns are never silently conflated.
_llm_schema_defaults = {
    "necessity_class_llm": (F.lit(None).cast("string")),
    "mobility_score_llm": (F.lit(None).cast("int")),
    "monitoring_score_llm": (F.lit(None).cast("int")),
    "total_score_llm": (F.lit(None).cast("int")),
    "has_named_concept_llm": (F.lit(None).cast("int")),
    "llm_rationale": (F.lit(None).cast("string")),
    "llm_evidence": (F.lit(None).cast("string")),
    "llm_summary": (F.lit(None).cast("string")),
}

if RUN_LLM and llm_pdf is not None and len(llm_pdf):
    _llm_sdf = spark.createDataFrame(sanitize_df(llm_pdf).astype(object).where(pd.notnull(llm_pdf), None))
    _join_col = ORDER_ID_COL or "OrderId"
    df = df.join(_llm_sdf, df[_join_col] == _llm_sdf["_order_id"], "left").drop("_order_id")
    df = df.withColumn("llm_status", F.coalesce(F.col("llm_status"), F.lit("not_run")))
else:
    for _c, _default in _llm_schema_defaults.items():
        df = df.withColumn(_c, _default)
    df = df.withColumn("llm_status", F.lit("not_run"))

# determination_agreement: how the two determinations relate for this order. This is the column the
# rule-vs-LLM comparison is built on - it is a comparison of two readings of the same text, not a
# statement that either one is correct.
df = df.withColumn(
    "determination_agreement",
    F.when(F.col("llm_status") != "ok", F.lit("llm_not_run"))
     .when(F.col("necessity_class") == F.col("necessity_class_llm"), F.lit("agree"))
     .otherwise(F.concat(F.lit("differ: score="), F.col("necessity_class"),
                         F.lit(", llm="), F.col("necessity_class_llm"))),
)

if RUN_LLM:
    print("Determination agreement:")
    display(df.groupBy("determination_agreement").agg(F.count("*").alias("orders"))
              .orderBy(F.desc("orders")))
# 10. Export summary workbook - med_nec_buckets_v2.xlsx

# Aggregates only. Row-level data lives in the review workbook and the detail CSV, so nothing
# here repeats it.

OUTPUT_XLSX = f"med_nec_buckets_v2_{RUN_TS}.xlsx"
# Destination directory. Workspace paths (/Workspace/...) are a real mounted filesystem, so the
# workbook is written there directly with pandas/openpyxl - no dbutils.fs.cp needed.
# For a Volume or DBFS destination instead, use e.g.:
#   "/Volumes/prod-sandbox/vivekkumar_patel/exports"   (Unity Catalog volume)
#   "dbfs:/FileStore/med_nec"                            (download via /files/ URL)
OUTPUT_DIR = "/Workspace/Users/josh.smitherman@gmr.net/med_nec/data"

total = df.count()

def with_pct(sdf, count_col="orders"):
    pdf = sdf.toPandas()
    pdf["pct_of_scope"] = (pdf[count_col] / total * 100).round(1)
    return pdf

# --- Summary tab: necessity, indeterminate reasons ---
nec_pdf = with_pct(
    df.groupBy("necessity_class").agg(F.count("*").alias("orders")).orderBy(F.desc("orders"))
)
indet_pdf = with_pct(
    df.groupBy("unmatched_text").agg(F.count("*").alias("orders")).orderBy(F.desc("orders"))
)

# --- Concepts tab: dictionary with counts and CMS citations (backs the concept slide) ---
concept_sums = df.agg(
    *[F.sum(F.col(f"c_{n}").cast("int")).alias(n) for n, *_ in CONCEPTS]
).collect()[0].asDict()
concept_pdf = pd.DataFrame([
    {"concept": n, "group": g, "axis": a, "weight": w, "basis": b,
     "cms_ref": ref, "source_url": url, "match_terms": pat,
     "tags": int(concept_sums[n] or 0),
     "pct_of_orders": round((concept_sums[n] or 0) / total * 100, 1)}
    for (n, a, w, g, b, ref, url, pat) in CONCEPTS
]).sort_values("tags", ascending=False).reset_index(drop=True)

# --- Mobility reasons tab: what reasons for mobility appear in the order text, and how often.
# Requested in the 10 Aug meeting (text analysis of reasons for mobility). Reuses the mobility-axis
# concept tags already computed - no new matching. Overall counts, plus counts among necessary
# orders, so the documented reasons that carry necessity are visible. ---
mobility_concepts = [(n, w, b) for n, a, w, g, b, *_ in CONCEPTS if a == "mobility"]
nec_df = df.filter(F.col("necessity_class") == "necessary")
nec_total = nec_df.count()
mob_sums_all = df.agg(
    *[F.sum(F.col(f"c_{n}").cast("int")).alias(n) for n, *_ in mobility_concepts]
).collect()[0].asDict()
mob_sums_nec = nec_df.agg(
    *[F.sum(F.col(f"c_{n}").cast("int")).alias(n) for n, *_ in mobility_concepts]
).collect()[0].asDict()
mobility_pdf = pd.DataFrame([
    {"mobility_reason": n, "weight": w, "basis": b,
     "orders": int(mob_sums_all[n] or 0),
     "pct_of_scope": round((mob_sums_all[n] or 0) / total * 100, 1),
     "orders_necessary": int(mob_sums_nec[n] or 0),
     "pct_of_necessary": round((mob_sums_nec[n] or 0) / max(nec_total, 1) * 100, 1)}
    for (n, w, b) in mobility_concepts
]).sort_values("orders", ascending=False).reset_index(drop=True)


# --- Where tab: by customer, by level of service, transport appropriateness ---
where_blocks = []
if CUSTOMER_COL is not None:
    where_blocks.append((
        "Necessity by customer",
        df.groupBy(CUSTOMER_COL, "necessity_class").agg(F.count("*").alias("orders"))
          .orderBy(CUSTOMER_COL, "necessity_class").toPandas()
    ))
if LOS_COL is not None:
    where_blocks.append((
        "Necessity by level of service",
        df.groupBy(LOS_COL, "necessity_class").agg(F.count("*").alias("orders"))
          .orderBy(LOS_COL, "necessity_class").toPandas()
    ))
    where_blocks.append((
        "Transport appropriateness (requested vs recommended)",
        df.groupBy(LOS_COL, "recommended_los").agg(F.count("*").alias("orders"))
          .orderBy(F.desc("orders")).toPandas()
    ))

# --- Column set for the row-level outputs (detail CSV and review workbook). This workbook holds
# aggregates only, so the same rows are not repeated across three files. ---
WRITE_FULL_DETAIL_CSV = True

concept_names = [n for n, *_ in CONCEPTS]
weight_col_names = [f"w_{n}" for n, a, w in SCORING_CONCEPTS]
context_cols = [c for c in [ORDER_ID_COL, LOS_COL, CUSTOMER_COL] if c is not None]
# concept hit columns, cast to 1/0 and named by the concept (the "labels")
label_cols = [F.col(f"c_{n}").cast("int").alias(n) for n in concept_names]
# weighted contribution per scoring concept, so the scores reconcile from the visible columns
weight_cols = [F.col(c) for c in weight_col_names]
derived_cols = ["mobility_score", "monitoring_score", "total_score",
                "named_score", "has_named_concept", "necessity_class",
                "unmatched_text", "recommended_los",
                "gy_disposition", "why_labeled", "classification_method"]
# Both determinations travel together everywhere: score columns, then LLM columns.
detail_select = ([F.col(c) for c in context_cols] + label_cols + weight_cols
                 + [F.col(c) for c in derived_cols]
                 + [F.col(c) for c in LLM_OUTPUT_COLS] + [F.col(FREE_TEXT_COL)])


# --- Determination comparison: score label vs LLM label. An aggregate, not row-level data.
# Only meaningful for orders the LLM actually read (llm_status = ok). ---
llm_read = df.filter(F.col("llm_status") == "ok")
llm_read_n = llm_read.count()
if llm_read_n:
    agreement_pdf = (llm_read.groupBy("determination_agreement")
                     .agg(F.count("*").alias("orders"))
                     .orderBy(F.desc("orders")).toPandas())
    agreement_pdf["pct_of_llm_read"] = (agreement_pdf["orders"] / llm_read_n * 100).round(1)
    crosstab_pdf = (llm_read.groupBy("necessity_class", "necessity_class_llm")
                    .agg(F.count("*").alias("orders"))
                    .orderBy(F.desc("orders")).toPandas())
    llm_status_pdf = (df.groupBy("llm_status").agg(F.count("*").alias("orders"))
                      .orderBy(F.desc("orders")).toPandas())
else:
    _none = pd.DataFrame({"note": ["LLM did not run - set RUN_LLM = True"]})
    agreement_pdf = crosstab_pdf = llm_status_pdf = _none

# --- Definitions tab: embedded so the workbook is self-explaining ---
definitions_pdf = pd.DataFrame([
    ["necessity_class = necessary", "A named CMS concept meets the clear threshold, or the full 10.2.3 test is met in the order text.", "BPM10 10.2.3 / 10.2.1; 414.605"],
    ["necessity_class = not_necessary", "Text present with a clinical reason field but only filler, no Group A concept.", "RSN AM600; MLN"],
    ["necessity_class = indeterminate", "Some signal but below the clear threshold, or text present that matched no term.", "BPM10 10.2.1"],
    ["unmatched_text", "1 = text present but nothing scored (words no rule recognized). Informational; the language-model target. Does not change the bucket.", "BPM10 10.2.1"],
    ["mobility_score", "Weighted sum of mobility-axis concepts - why other transport is contraindicated.", "BPM10 10.2.1 / 10.2.3"],
    ["monitoring_score", "Weighted sum of monitoring-axis concepts - why this level of service.", "414.605"],
    ["total_score", "mobility_score + monitoring_score. Drives the bucket via cutoffs.", "composite (provisional)"],
    ["named_score", "Points from named CMS concepts only. A named concept (named_score > 0) is required for necessary.", "BPM10 / 414.605"],
    ["has_named_concept", "1 if any named CMS concept matched. Gates the necessary band.", "BPM10 / 414.605"],
    ["recommended_los", "Level of service implied by the monitoring axis. Descriptive, not a billing call.", "414.605; 410.40(c)"],
    ["gy_disposition", "GY process relation: not_necessary = no documented reason (GY candidate); necessary = documented reason present; indeterminate = partial, review.", "review-process 4 Aug"],
    ["GY modifier", "Billing marks a non-medically-necessary Medicare trip GY: transport provided, payment not requested. Enables billing the facility/patient and building a record.", "review-process 4 Aug"],
    ["PCS", "Physician Certification Statement - the order-time certification. This analysis reads the order-time (PCS-side) free text.", "410.40(d); review-process"],
    ["PCR", "Patient Care Report - the crew's documentation in ImageTrend. A separate source, NOT in this table. Billing requires the PCR to support the PCS.", "review-process 4 Aug"],
    ["Routing", "Non-emergency Medicare trips go to Hemlata Khatri's team (under Jen); Medicaid, commercial and self-pay go to Integra. Coding, with medical-necessity confirmation, is the first step for both.", "review-process 4 Aug"],
    ["classification_method", "How labels were assigned: keyword_match today. text_unmatched orders are the set a language model would categorise next.", "n/a"],
    ["Group A vs Group B", "A = a clinical reason (12 concepts). B = vague filler that confers no necessity (3).", "BPM10 10.2.1; MLN"],
    ["named vs inferred", "named = explicit in CMS text. inferred = derived from the 10.2.1 general test.", "BPM10 10.2.1"],
    ["Limitation", "Necessity here reflects ORDER-TIME documentation only. The final billing determination also depends on the crew PCR supporting the PCS, which is not in this dataset.", "review-process 4 Aug"],
    ["Scope", f"Non-emergent ground ambulance only. {total:,} orders in scope, 2024 to present. Wheelchair, ambulatory, air and emergent codes excluded.", "BPM10 10.2.1"],
], columns=["term", "definition", "cms_ref"])

# --- Scoring tab: weight table, aggregation formula, and score distribution ---
weight_pdf = pd.DataFrame([
    {"concept": n, "axis": a, "group": g, "basis": b, "weight": w, "cms_ref": ref}
    for (n, a, w, g, b, ref, url, pat) in CONCEPTS
]).sort_values(["axis", "weight"], ascending=[True, False]).reset_index(drop=True)

formula_pdf = pd.DataFrame([
    ["mobility_score", "sum of weights of matched mobility-axis concepts", "BPM10 10.2.1 / 10.2.3"],
    ["monitoring_score", "sum of weights of matched monitoring-axis concepts", "414.605"],
    ["total_score", "mobility_score + monitoring_score", "composite (provisional)"],
    ["necessary", f"total_score >= {NECESSARY_CUTOFF} AND a named concept present", "BPM10 / 414.605"],
    ["not_necessary", "total_score == 0", "RSN AM600"],
    ["indeterminate", "everything else", "BPM10 10.2.1"],
    ["note", "weights and bonus are tunable modelling choices, not CMS figures - validate with SMEs", "n/a"],
], columns=["element", "definition", "cms_ref"])

total_dist_pdf = with_pct(
    df.groupBy("total_score").agg(F.count("*").alias("orders"))
      .orderBy("total_score")
)

score_by_class_pdf = (
    df.groupBy("necessity_class")
      .agg(F.round(F.avg("total_score"), 2).alias("avg_total_score"),
           F.min("total_score").alias("min_score"),
           F.max("total_score").alias("max_score"),
           F.count("*").alias("orders"))
      .orderBy(F.desc("avg_total_score")).toPandas()
)

# Aggregates only - every tab answers a different question. Row-level orders are in the review
# workbook and the detail CSV, and are deliberately not repeated here.
sheets = {
    "Definitions": [("Definitions and CMS basis", definitions_pdf)],
    "Summary":     [("Necessity - score determination", nec_pdf),
                    ("Indeterminate - why", indet_pdf)],
    "Scoring":     [("Concept weights", weight_pdf),
                    ("Aggregation formula", formula_pdf),
                    ("Total score distribution", total_dist_pdf),
                    ("Total score by necessity class", score_by_class_pdf)],
    "Concepts":    [("Concept dictionary - terms, weights, CMS basis", concept_pdf)],
    "Mobility reasons": [("Reasons for mobility - frequency overall and among necessary orders",
                          mobility_pdf)],
    "Determinations": [("Score vs LLM - agreement", agreement_pdf),
                       ("Score label vs LLM label", crosstab_pdf),
                       ("LLM coverage", llm_status_pdf)],
    "Where":       where_blocks or [("No customer/LOS column resolved", pd.DataFrame({"note": ["set FIELD_CANDIDATES"]}))],
}

def _coerce(v):
    if pd.isna(v):
        return None
    v = v.item() if hasattr(v, "item") else v
    return _clean(v)

def build_workbook(path, sheets):
    wb = Workbook(); wb.remove(wb.active)
    for name, blocks in sheets.items():
        ws = wb.create_sheet(name[:31]); r = 1
        for title, pdf in blocks:
            if title:
                ws.cell(r, 1, title).font = Font(bold=True); r += 1
            for j, col in enumerate(pdf.columns, 1):
                ws.cell(r, j, str(col)).font = Font(bold=True)
            r += 1
            for _, row in pdf.iterrows():
                for j, val in enumerate(row, 1):
                    ws.cell(r, j, _coerce(val))
                r += 1
            r += 1
        for col_cells in ws.columns:
            w = min(70, max((len(str(c.value)) for c in col_cells if c.value is not None), default=8) + 2)
            ws.column_dimensions[col_cells[0].column_letter].width = w
    wb.save(path)

dest = f"{OUTPUT_DIR.rstrip('/')}/{OUTPUT_XLSX}"
# Always build to a driver-local path first, then copy to the destination. Direct writes to
# /Workspace and /Volumes FUSE paths are unreliable on serverless compute.
local_path = f"/tmp/{OUTPUT_XLSX}"
build_workbook(local_path, sheets)
import os, shutil
if OUTPUT_DIR.startswith("/Workspace/") or OUTPUT_DIR.startswith("/dbfs/") or OUTPUT_DIR.startswith("/Volumes/"):
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        shutil.copyfile(local_path, dest)
    except Exception:
        dbutils.fs.cp(f"file:{local_path}", dest)
    print("Saved workbook to:", dest)
else:
    try:
        dbutils.fs.mkdirs(OUTPUT_DIR)
        dbutils.fs.cp(f"file:{local_path}", dest)
        print("Saved workbook to:", dest)
        if OUTPUT_DIR.startswith("dbfs:/FileStore"):
            print("Download via: <workspace-url>/files/" + dest.split("/FileStore/", 1)[1])
    except Exception as e:
        print("Wrote local copy:", local_path)
        print("Copy to", dest, "failed - set OUTPUT_DIR to a writable path.")
        print("Reason:", e)

# Full row-level detail as CSV beside the workbook (every scored order, not just the sample).
if WRITE_FULL_DETAIL_CSV:
    csv_name = OUTPUT_XLSX.replace(".xlsx", "_detail.csv")
    csv_dest = f"{OUTPUT_DIR.rstrip('/')}/{csv_name}"
    local_csv = f"/tmp/{csv_name}"
    sanitize_df(df.select(*detail_select).toPandas()).to_csv(local_csv, index=False)
    try:
        if OUTPUT_DIR.startswith("/Workspace/") or OUTPUT_DIR.startswith("/dbfs/") or OUTPUT_DIR.startswith("/Volumes/"):
            shutil.copyfile(local_csv, csv_dest)
        else:
            dbutils.fs.cp(f"file:{local_csv}", csv_dest)
        print("Saved full detail CSV to:", csv_dest)
    except Exception as e:
        print("Wrote local detail CSV:", local_csv, "| copy failed:", e)

# 11. Review workbook - one row per order, both determinations side by side

# The row-level workbook, aimed at the business review. ONE data tab, not several: the previous
# per-class tabs (Not necessary / Rules missed / Borderline) were all subsets of the same rows, so
# the same order appeared in three places. They are Excel filters on the single Orders tab instead.
# Free text is shown first, then the score determination, then the LLM determination and its
# rationale. Header row frozen, autofilter on, so the subsets are one click away.

from openpyxl.utils import get_column_letter

BUILD_REVIEW_WORKBOOK = True
REVIEW_XLSX = f"med_nec_review_{RUN_TS}.xlsx"
# Excel's own hard limit is 1,048,576 rows per sheet including the header, so the Orders tab
# cannot be truly uncapped. This is set to that ceiling rather than an arbitrary number, and any
# truncation is printed explicitly below so a capped tab is never mistaken for a full count.
# The uncapped, full-population outputs are the Genie table (section 12) and the detail CSV.
MAX_REVIEW_ROWS = 1_048_575

if BUILD_REVIEW_WORKBOOK:
    # Reading order: context, the text itself, then the SCORE determination, then the LLM
    # determination with its rationale, then how the two relate.
    review_cols = [c for c in [ORDER_ID_COL, LOS_COL, CUSTOMER_COL] if c is not None] + [
        FREE_TEXT_COL,
        "necessity_class", "total_score", "mobility_score", "monitoring_score",
        "named_score", "has_named_concept", "why_labeled", "gy_disposition",
        "unmatched_text", "recommended_los",
        "necessity_class_llm", "total_score_llm", "mobility_score_llm",
        "monitoring_score_llm", "has_named_concept_llm",
        "llm_rationale", "llm_evidence", "llm_summary", "llm_status",
        "determination_agreement",
    ]

    def review_frame(spark_df, limit=None):
        d = spark_df.select(*review_cols)
        if limit:
            d = d.limit(limit)
        return d.toPandas()

    with_text = df.filter(F.col("_has_text"))

    guide_pdf = pd.DataFrame([
        ["Purpose", "Walk through actual orders and see how each was read for medical-necessity documentation."],
        ["Tabs", "One data tab: Orders. Every order with free text, one row each. Use the column filters for subsets - there is no separate tab per subset."],
        ["", ""],
        ["TWO DETERMINATIONS", "Each order is read twice, independently, using the SAME cutoffs. Any difference comes from what was read out of the text, not from different thresholds."],
        ["free text (ClinicalData)", "The reason-for-transport text entered at order time. Read this first."],
        ["", ""],
        ["SCORE DETERMINATION", "Weighted keyword concepts."],
        ["necessity_class", "Score label: necessary / indeterminate / not_necessary."],
        ["total_score", "mobility_score + monitoring_score. 0 -> not_necessary; >= 3 with a named concept -> necessary; else indeterminate."],
        ["why_labeled", "The concepts that matched and the exact words that triggered each. Empty = the rules found nothing."],
        ["unmatched_text", "1 = text present that no keyword rule read. Filter on this for the rules' blind spot."],
        ["gy_disposition", "GY process relation: no documented reason (GY candidate) / documented reason present / partial, review."],
        ["", ""],
        ["LLM DETERMINATION", "A language model extracts the same facts from the same text; the same rules then assign the label."],
        ["necessity_class_llm", "LLM label: necessary / indeterminate / not_necessary. Blank where the LLM did not read the order."],
        ["llm_rationale", "Plain-English explanation of what the text does and does not establish. This is the WHY behind the LLM label."],
        ["llm_evidence", "Verbatim quote from the order text behind every fact the model asserted."],
        ["llm_status", "ok = the LLM read this order; not_run = it did not; error = the extraction failed validation."],
        ["", ""],
        ["determination_agreement", "agree / differ (with both labels named) / llm_not_run. Filter on 'differ' for the cases worth reviewing."],
        ["", ""],
        ["Limitation", "Reflects ORDER-TIME documentation only. Final billing also depends on the crew PCR supporting the PCS, which is not in this data."],
        ["Not a billing determination", "These labels describe what the order text documents. They are not a coverage, denial, or payment decision."],
    ], columns=["item", "meaning"])

    # Actual population for the Orders tab, so truncation is visible rather than implied.
    with_text_total = with_text.count()

    review_tabs = {
        "How to read": guide_pdf,
        "Orders": review_frame(with_text, MAX_REVIEW_ROWS),
    }

    if with_text_total > MAX_REVIEW_ROWS:
        print(f"NOTE: 'Orders' truncated to {MAX_REVIEW_ROWS:,} of {with_text_total:,} rows "
              "(Excel sheet limit). Use the Genie table or the detail CSV for the full population.")
    else:
        print(f"'Orders' holds the full population: {with_text_total:,} rows.")

    def build_review(path, tabs):
        # Write to a driver-local path first, then the caller copies it to the destination.
        # Direct pandas writes to /Workspace or /Volumes FUSE paths are unreliable on serverless.
        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            for name, pdf in tabs.items():
                sheet = (name[:31] or "Sheet")
                if pdf is None or len(pdf) == 0:
                    pdf = pd.DataFrame({"note": ["no rows for this view"]})
                pdf = sanitize_df(pdf)
                pdf.to_excel(xw, sheet_name=sheet, index=False)
                ws = xw.sheets[sheet]
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                ws.freeze_panes = "A2"
                # Autofilter on the data tab: the per-class tabs this replaces are now one click.
                if name != "How to read" and len(pdf.columns):
                    ws.auto_filter.ref = ws.dimensions
                for i, col in enumerate(pdf.columns, 1):
                    L = get_column_letter(i)
                    if col in (FREE_TEXT_COL, "why_labeled", "llm_rationale",
                               "llm_evidence", "llm_summary"):
                        ws.column_dimensions[L].width = 60
                    else:
                        ws.column_dimensions[L].width = min(26, max(len(str(col)) + 2, 10))

    review_dest = f"{OUTPUT_DIR.rstrip('/')}/{REVIEW_XLSX}"
    local_review = f"/tmp/{REVIEW_XLSX}"
    build_review(local_review, review_tabs)
    if OUTPUT_DIR.startswith("/Workspace/") or OUTPUT_DIR.startswith("/dbfs/") or OUTPUT_DIR.startswith("/Volumes/"):
        # Copy the finished file to the mounted destination (works on serverless).
        import shutil
        try:
            shutil.copyfile(local_review, review_dest)
        except Exception:
            dbutils.fs.cp(f"file:{local_review}", review_dest)
        print("Saved review workbook to:", review_dest)
        for nm, pdf in review_tabs.items():
            print(f"  {nm}: {len(pdf):,} rows")
    else:
        try:
            dbutils.fs.cp(f"file:{local_review}", review_dest)
            print("Saved review workbook to:", review_dest)
        except Exception as e:
            print("Wrote local review workbook:", local_review, "| copy failed:", e)

# 12. Genie table - persist the scored data as a curated table for AI/BI Genie

# Genie queries a real table or view, so this is the one place the pipeline writes. It writes only
# this derived output, to a schema you control - the source table is read only and is never touched.
# Set WRITE_GENIE_TABLE = False to restore the fully read-only posture.

WRITE_GENIE_TABLE = True

# Target for the scored table, split so only GENIE_SCHEMA needs editing.
# prod-sandbox uses one schema per person (von_aday, weilan_zeng, pavithra_dedigama, ...).
# vivekkumar_patel is the SOURCE owner's schema - writing there needs his agreement, so point
# this at your own schema. The source table is read only and is never modified either way.
GENIE_CATALOG    = "prod-sandbox"
GENIE_SCHEMA     = "josh_smitherman"
GENIE_TABLE_NAME = "med_nec_genie"
GENIE_TABLE = ".".join(f"`{p}`" for p in (GENIE_CATALOG, GENIE_SCHEMA, GENIE_TABLE_NAME))

if WRITE_GENIE_TABLE:
    # Pre-flight: an existing schema you cannot write to passes CREATE SCHEMA IF NOT EXISTS as a
    # no-op, so test the privilege that actually matters by creating and dropping a tiny table.
    # This surfaces PERMISSION_DENIED in seconds rather than after the full scoring run.
    _probe = f"`{GENIE_CATALOG}`.`{GENIE_SCHEMA}`.`_med_nec_write_probe`"
    try:
        spark.sql(f"CREATE OR REPLACE TABLE {_probe} AS SELECT 1 AS ok")
        spark.sql(f"DROP TABLE IF EXISTS {_probe}")
        print(f"Write access confirmed on {GENIE_CATALOG}.{GENIE_SCHEMA}")
    except Exception as e:
        print(f"Cannot write to {GENIE_CATALOG}.{GENIE_SCHEMA}: {e}")
        print("Set GENIE_SCHEMA to a schema you own, or request CREATE TABLE on this one.")
        raise

    concept_names_all = [n for n, *_ in CONCEPTS]
    weight_names_all = [f"w_{n}" for n, a, w in SCORING_CONCEPTS]
    # Column set is deliberate: order/level-of-service/customer context, the scoring outputs, the
    # concept flags, the per-concept weighted contributions (so the scores reconcile from the data
    # itself), and the order-time free text. Direct identifiers carried by the source table
    # (MRN, MRNSource, requester name and phone) are NOT selected - a Genie space invites open
    # questions and those fields have no analytical use here.
    genie_cols = (
        [c for c in [ORDER_ID_COL, LOS_COL, CUSTOMER_COL] if c is not None]
        + ["necessity_class", "gy_disposition", "total_score", "mobility_score",
           "monitoring_score", "named_score", "has_named_concept", "unmatched_text",
           "recommended_los", "why_labeled", "classification_method"]
        + [F.col(f"c_{n}").cast("int").alias(n) for n in concept_names_all]
        + [F.col(c) for c in weight_names_all]
        + [F.col(c) for c in LLM_OUTPUT_COLS]
        + [F.col(FREE_TEXT_COL).alias("clinical_text")]
    )
    genie_df = df.select(*[F.col(c) if isinstance(c, str) else c for c in genie_cols])

    # Run provenance, so anyone querying the space can see how current the data is and what it
    # came from. RUN_TS is the same constant used for every file output from this run.
    genie_df = (genie_df
                .withColumn("run_ts", F.lit(RUN_TS))
                .withColumn("source_table", F.lit(SOURCE_TABLE)))

    # Full in-scope population - no row cap. The Excel review workbook is capped by Excel's own
    # sheet limit; this table is not, so it is the authoritative count for any question of "how many".
    genie_df.write.mode("overwrite").option("overwriteSchema", "true").saveAsTable(GENIE_TABLE)
    genie_rows = spark.table(GENIE_TABLE).count()
    print(f"Wrote Genie table: {GENIE_TABLE} - {genie_rows:,} rows (full in-scope population)")
    if genie_rows != scope_count:
        print(f"  NOTE: table rows ({genie_rows:,}) != in-scope count ({scope_count:,}) - investigate.")

    # Column comments - Genie relies on these to understand the data. Keep them factual.
    col_comments = {
        "necessity_class": "Label for the order text: necessary / not_necessary / indeterminate.",
        "gy_disposition": "GY-process relation: not_necessary=no documented reason (GY candidate); necessary=documented reason present; indeterminate=partial.",
        "total_score": "mobility_score + monitoring_score. 0=not_necessary; >=3 with a named concept=necessary; else indeterminate.",
        "mobility_score": "Sum of matched mobility-axis concept weights (why other transport is contraindicated). BPM10 10.2.1/10.2.3.",
        "monitoring_score": "Sum of matched monitoring-axis concept weights (why this level of service). 42 CFR 414.605.",
        "named_score": "Points from named CMS concepts only. A SUBSET of total_score, not a third component - the named concepts are already counted in mobility_score and monitoring_score. Do not add it to them.",
        "has_named_concept": "1 if any named CMS concept matched, else 0.",
        "unmatched_text": "1 if text was entered but nothing scored (words no rule recognized). The language-model target set.",
        "recommended_los": "Level of service implied by monitoring concepts (descriptive, not a billing call).",
        "why_labeled": "The concepts that matched and the exact text that triggered each.",
        "clinical_text": "The free-text reason-for-transport entered at order time (order-time / PCS-side documentation).",
        "necessity_class_llm": "The LLM determination for the same order text: necessary / not_necessary / indeterminate. NULL where the LLM did not read this order - check llm_status. This is a SECOND, independent reading, not a correction of necessity_class.",
        "mobility_score_llm": "Mobility-axis score from the facts the LLM extracted. Same weights as mobility_score.",
        "monitoring_score_llm": "Monitoring-axis score from the facts the LLM extracted. Same weights as monitoring_score.",
        "total_score_llm": "mobility_score_llm + monitoring_score_llm. The SAME cutoffs as the score path are applied to it, so any label difference comes from what was read in the text.",
        "has_named_concept_llm": "1 if the LLM found any named CMS concept, else 0.",
        "llm_rationale": "Plain-English explanation of what the order text does and does not establish about mobility and monitoring. This is the stated reason behind necessity_class_llm.",
        "llm_evidence": "Verbatim quotes from the order text supporting each fact the LLM asserted, as field[quote] pairs.",
        "llm_summary": "One-line summary of the order text as read by the LLM.",
        "llm_status": "ok = the LLM read this order; not_run = it did not (LLM off, or outside the target set, or beyond the sample); error: ... = the extraction failed validation.",
        "determination_agreement": "How the two determinations relate: agree, differ (naming both labels), or llm_not_run. Only rows with llm_status = ok can agree or differ.",
        "run_ts": "Timestamp (YYYYMMDD_HHMMSS) of the scoring run that produced this table. All rows share one value; the table is overwritten each run.",
        "source_table": "The source table this run read from.",
    }
    for n, a, w, g, b, ref, url, pat in CONCEPTS:
        col_comments[n] = f"1 if the order text matched the {n} concept ({a} axis, weight {w}, {b}). Basis: {ref}."
    for n, a, w in SCORING_CONCEPTS:
        col_comments[f"w_{n}"] = (
            f"Weighted contribution of {n} to the score: {w} when matched, 0 otherwise. "
            f"Summing every w_ column on the {a} axis gives {a}_score."
        )
    for col, comment in col_comments.items():
        safe = comment.replace("'", "")
        try:
            spark.sql(f"ALTER TABLE {GENIE_TABLE} ALTER COLUMN {col} COMMENT '{safe}'")
        except Exception as e:
            print(f"  comment on {col} skipped:", e)
    spark.sql(
        f"COMMENT ON TABLE {GENIE_TABLE} IS "
        "'Non-emergent ground ambulance orders assessed for medical-necessity documentation against "
        "CMS criteria (order-time text only). One row per order, full in-scope population, no row cap. "
        "TWO determinations per order from the same text and the same cutoffs: necessity_class from "
        "weighted keyword concepts, and necessity_class_llm from language-model extraction with a "
        "stated rationale in llm_rationale. They are independent readings, not one correcting the other; "
        "determination_agreement compares them. Wheelchair, ambulatory, air and emergent codes are "
        "excluded from scope. Labels describe the documentation, not the trip, and are not a billing "
        "determination. Overwritten each run; see run_ts.'"
    )
    print("Applied column and table comments. Point a Genie space at:", GENIE_TABLE)
# Notes / open items

# - Weights and thresholds are provisional. They encode the CMS hierarchy (named [BPM10] 10.2.3
#   concepts strongest, inferred [BPM10] 10.2.1 concepts weakest) but should be validated by the
#   SMEs and, once available, against actual denial outcomes in [RSN].
# - text_unmatched is the LLM target - text present that no term captured. That count is the
#   concrete case for language-model classification over term matching.
# - recommended_los is descriptive until CMS BLS/ALS eligibility criteria in [414.605] and
#   [410.40](c) are confirmed; it surfaces requested-vs-recommended mismatches, not billing calls.
# - behavioral, oxygen, isolation, wound_ostomy, bariatric are inferred from [BPM10] 10.2.1 and
#   are the first candidates for SME challenge; behavioral is the weakest.
# - Second free-text field: the review raised expanding beyond ClinicalData. Add fields to the
#   match step only after confirming they are point-of-order per [BPM10] 10.2.4, not post-hoc.

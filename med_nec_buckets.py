# Databricks notebook source
# =============================================================================
# Medical Necessity EDA - bucketing the clinical data
#
# GOAL
# Sort the ClinicalData free text into buckets so we can see how much of it
# actually supports medical necessity.
#
# Output: one Excel file, 5 tabs, plain formatting.
#   1. Definitions   - what we did and why
#   2. Summary       - the headline numbers
#   3. Where         - buckets by customer and level of service
#   4. Examples      - real orders from each bucket
#   5. Concepts      - which clinical reasons show up in the text
#
# READ-ONLY. Temp views only. No tables created.
# Source: `prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster
# =============================================================================


# COMMAND ----------

# Run once, then comment out:
# %pip install openpyxl
# dbutils.library.restartPython()


# COMMAND ----------

TRIPMASTER  = "`prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster"
OUTPUT_DIR  = "/Workspace/Users/josh.smitherman@gmr.net/med_nec/data"
OUTPUT_XLSX = f"{OUTPUT_DIR}/med_nec_buckets.xlsx"
EXAMPLES_PER_BUCKET = 20

import os
from pyspark.sql import functions as F

spark.sql("USE CATALOG prod")
os.makedirs(OUTPUT_DIR, exist_ok=True)


# COMMAND ----------

# -----------------------------------------------------------------------------
# STEP 1 - SCOPE
# Only non-emergent ground transports need medical necessity documentation.
# Rideshare, air, and emergent do not, so they come out of the denominator.
# -----------------------------------------------------------------------------

# Check the codes first.
display(spark.sql(f"""
    SELECT LevelOfService, LevelOfServiceDescription, ServiceType, count(*) AS orders
    FROM {TRIPMASTER} GROUP BY 1,2,3 ORDER BY orders DESC
"""))


# COMMAND ----------

OUT_OF_SCOPE = """
    (
        upper(coalesce(ServiceType,''))               IN ('TAXI','ROTOR','EMG')
     OR upper(coalesce(LevelOfService,''))            IN ('TAXI','ROTOR','ALS-EMG','EMG','CCT-EMG')
     OR upper(coalesce(LevelOfServiceDescription,'')) LIKE '%EMERGEN%'
     OR upper(coalesce(LevelOfServiceDescription,'')) IN ('RIDESHARE','LYFT','ROTOR','FIXED WING QUOTE')
    )
"""

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW scoped AS
SELECT *, CASE WHEN {OUT_OF_SCOPE} THEN 'out' ELSE 'in' END AS scope
FROM {TRIPMASTER}
""")

display(spark.sql("""
    SELECT scope, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM scoped GROUP BY scope
"""))

print("Excluded:")
display(spark.sql("""
    SELECT LevelOfServiceDescription, count(*) AS orders
    FROM scoped WHERE scope='out' GROUP BY 1 ORDER BY orders DESC
"""))


# COMMAND ----------

# -----------------------------------------------------------------------------
# STEP 2 - CONCEPTS
# We search the nurse's free text for two kinds of words.
#
# Group A = a clinical reason. A specific medical fact that explains why the
#           patient cannot use a wheelchair van or car.
# Group B = vague filler. Words that describe a state but give no cause and no
#           functional limit, so they do not justify an ambulance on their own.
# -----------------------------------------------------------------------------

GROUP_A = {   # clinical reason - supports medical necessity
    "mobility_deficit": r"hemipar|hemipleg|paraly|non[\s-]*ambulat|unable to (bear weight|ambulate|walk|stand)|fracture|amputat|contracture|bear weight",
    "cannot_sit":       r"cannot sit|unable to sit|special positioning|supine|must lie|stretcher|cannot support trunk",
    "bed_confined":     r"bed[\s-]*(bound|confined)|unable to get up|cannot get out of bed",
    "oxygen":           r"oxygen|\bo2\b|\blpm\b|nasal cannula|\bbipap\b|\bcpap\b",
    "cardiac":          r"cardiac|telemetry|\bekg\b|\becg\b|\bnstemi\b|\bstemi\b|arrhythm|\bafib\b",
    "ventilator":       r"ventilat|\bvent\b|trach|intubat",
    "suctioning":       r"suction",
    "iv_medication":    r"\biv\b|infusion|\bdrip\b|heparin|antibiotic|\btpn\b",
    "wound_ostomy":     r"wound|ostomy|ulcer|decubitus|drain",
    "isolation":        r"isolation|\bmrsa\b|c\.? ?diff|precaution",
    "behavioral":       r"dementia|alzheimer|combative|agitat|altered mental|flight risk|elope",
    "bariatric":        r"bariatric|morbid",
}

GROUP_B = {   # vague filler - does not support medical necessity alone
    "weakness_only":  r"general(iz)?e?d? weakness|generally weak|^\s*weak",
    "fall_risk_only": r"fall risk|unsteady|deconditio",
    "nonclinical":    r"per protocol|convenience|no other transport|unable to arrange|family request",
}

# CMS source behind each concept. "inferred" = not named in CMS guidance, rests on the
# general contraindication test in Benefit Policy Manual 10.2.1. Review those with the SMEs.
CMS_REF = {
    "mobility_deficit": "Manual 10.2.3 prongs 1-2 (unable to get up / unable to ambulate)",
    "cannot_sit":       "Manual 10.2.3 prong 3 (unable to sit in a chair or wheelchair)",
    "bed_confined":     "Manual 10.2.3 (all three prongs required)",
    "cardiac":          "42 CFR 414.605 (ALS assessment / ALS2 procedures)",
    "iv_medication":    "42 CFR 414.605 (ALS2 - 3+ IV medications; central/intraosseous line)",
    "ventilator":       "42 CFR 414.605 (ALS2 intubation/surgical airway; SCT respiratory care)",
    "suctioning":       "42 CFR 414.605 (SCT - beyond EMT-Paramedic scope)",
    "oxygen":           "inferred - Manual 10.2.1 general test. Chronic O2 alone does not qualify.",
    "wound_ostomy":     "inferred - Manual 10.2.1. Works via positioning, not on its own.",
    "isolation":        "inferred - Manual 10.2.1. Infection control, not a patient contraindication.",
    "bariatric":        "inferred - Manual 10.2.1. Handling requirement, not a contraindication.",
    "behavioral":       "inferred - WEAKEST. CMS wants a physical limitation. Consider Group B.",
    "weakness_only":    "MAC guidance: 'Vague and general information is of little or no value.'",
    "fall_risk_only":   "MAC guidance (same). A risk is not a contraindication.",
    "nonclinical":      "Manual 10.2.1 - a physician order does not prove necessity; other transport "
                        "disqualifies 'whether or not actually available'.",
}

ALL_CONCEPTS = {**GROUP_A, **GROUP_B}

cols  = ",\n       ".join(f"CASE WHEN lower(ClinicalData) RLIKE '{r}' THEN 1 ELSE 0 END AS c_{n}"
                          for n, r in ALL_CONCEPTS.items())
a_sum = " + ".join(f"c_{n}" for n in GROUP_A)
b_sum = " + ".join(f"c_{n}" for n in GROUP_B)

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW tagged AS
SELECT TripRequestId, ContractName, RequesterFacility,
       LevelOfServiceDescription, ClinicalData,
       length(ClinicalData) AS text_len,
       {cols}
FROM scoped WHERE scope='in'
""")


# COMMAND ----------

# -----------------------------------------------------------------------------
# STEP 3 - BUCKETS
# Each order gets exactly one bucket.
# -----------------------------------------------------------------------------

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW bucketed AS
SELECT *,
    ({a_sum}) AS clinical_reasons,
    ({b_sum}) AS filler_terms,
    CASE
        WHEN ClinicalData IS NULL OR length(trim(ClinicalData))=0 THEN 'no_documentation'
        WHEN ({a_sum})=0 AND ({b_sum})>0                          THEN 'filler_only'
        WHEN ({a_sum})=0                                          THEN 'unrecognized'
        WHEN ({b_sum})>0                                          THEN 'reason_plus_filler'
        ELSE 'clear_reason'
    END AS bucket
FROM tagged
""")

display(spark.sql("""
    SELECT bucket, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM bucketed GROUP BY bucket ORDER BY orders DESC
"""))


# COMMAND ----------

# -----------------------------------------------------------------------------
# STEP 4 - BUILD THE 5 TABLES
# -----------------------------------------------------------------------------

import pandas as pd
def q(sql): return spark.sql(sql).toPandas()

# --- Tab 2: Summary ---
scope_counts = q("SELECT scope, count(*) AS orders FROM scoped GROUP BY scope").set_index("scope")["orders"].to_dict()
total_raw = scope_counts.get("in", 0) + scope_counts.get("out", 0)

summary = q("""
    SELECT bucket AS Bucket,
           CASE bucket
             WHEN 'no_documentation'   THEN 'Nothing typed'
             WHEN 'filler_only'        THEN 'Vague filler only'
             WHEN 'unrecognized'       THEN 'Text present, nothing recognized'
             WHEN 'reason_plus_filler' THEN 'Clinical reason plus filler'
             WHEN 'clear_reason'       THEN 'Clear clinical reason'
           END AS Meaning,
           CASE WHEN bucket IN ('no_documentation','filler_only') THEN 'AT RISK'
                WHEN bucket = 'unrecognized' THEN 'Unknown'
                ELSE 'Supported' END AS Status,
           count(*) AS Orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS Pct
    FROM bucketed GROUP BY bucket ORDER BY Orders DESC
""")

# --- Tab 3: Where (at-risk rate by customer and by service level) ---
by_contract = q("""
    SELECT 'By customer' AS View, ContractName AS Group_,
           count(*) AS Orders,
           sum(CASE WHEN bucket='no_documentation' THEN 1 ELSE 0 END) AS NothingTyped,
           sum(CASE WHEN bucket='filler_only'      THEN 1 ELSE 0 END) AS FillerOnly,
           sum(CASE WHEN bucket IN ('no_documentation','filler_only') THEN 1 ELSE 0 END) AS AtRisk,
           round(100.0*sum(CASE WHEN bucket IN ('no_documentation','filler_only') THEN 1 ELSE 0 END)
                 /count(*),1) AS PctAtRisk
    FROM bucketed GROUP BY ContractName
""")
by_los = q("""
    SELECT 'By level of service' AS View, LevelOfServiceDescription AS Group_,
           count(*) AS Orders,
           sum(CASE WHEN bucket='no_documentation' THEN 1 ELSE 0 END) AS NothingTyped,
           sum(CASE WHEN bucket='filler_only'      THEN 1 ELSE 0 END) AS FillerOnly,
           sum(CASE WHEN bucket IN ('no_documentation','filler_only') THEN 1 ELSE 0 END) AS AtRisk,
           round(100.0*sum(CASE WHEN bucket IN ('no_documentation','filler_only') THEN 1 ELSE 0 END)
                 /count(*),1) AS PctAtRisk
    FROM bucketed GROUP BY LevelOfServiceDescription HAVING count(*) >= 50
""")
where = pd.concat([by_contract.sort_values("AtRisk", ascending=False),
                   by_los.sort_values("AtRisk", ascending=False)], ignore_index=True)
where = where.rename(columns={"Group_": "Group"})

# --- Tab 4: Examples ---
examples = q(f"""
    SELECT bucket AS Bucket, ContractName AS Customer,
           LevelOfServiceDescription AS LevelOfService,
           clinical_reasons AS ClinicalReasons, filler_terms AS FillerTerms,
           ClinicalData AS WhatTheNurseTyped
    FROM (
      SELECT *, row_number() OVER (PARTITION BY bucket ORDER BY text_len DESC) AS rn
      FROM bucketed
    ) WHERE rn <= {EXAMPLES_PER_BUCKET}
    ORDER BY Bucket, rn
""")

# --- Tab 5: Concepts ---
concepts = pd.concat([
    spark.sql(f"""SELECT '{n}' AS Concept, sum(c_{n}) AS Orders,
                  round(100.0*sum(c_{n})/count(*),2) AS PctOfOrders FROM bucketed""").toPandas()
    for n in ALL_CONCEPTS], ignore_index=True)
concepts["Group"] = concepts["Concept"].apply(
    lambda c: "A - clinical reason" if c in GROUP_A else "B - vague filler")
concepts["CMS_Reference"] = concepts["Concept"].map(CMS_REF)
concepts["SearchTerms"] = concepts["Concept"].map(
    {k: v.replace("\\b", "").replace("[\\s-]*", " ") for k, v in ALL_CONCEPTS.items()})
concepts = concepts.sort_values("Orders", ascending=False)

print("tables built")
display(summary)


# COMMAND ----------

# -----------------------------------------------------------------------------
# STEP 4b - FULL ROW-LEVEL DETAIL FOR PIVOT TABLES
#
# One row per order with every dimension and every concept flag (0/1).
# Written as CSV - faster and far smaller than a 90k-row Excel tab, and Excel
# pivots off a CSV without any trouble.
#
# NOTE: WhatTheNurseTyped contains free text that may include patient
# information. Keep this file internal.
# -----------------------------------------------------------------------------

DETAIL_CSV  = f"{OUTPUT_DIR}/med_nec_detail.csv"
CONCEPT_CSV = f"{OUTPUT_DIR}/med_nec_concept_rows.csv"

# Include a date column if the table has one, so you can pivot by month.
have_date = "RequestDateTime" in spark.table("scoped").columns
date_cols = ("""
       ,date(RequestDateTime)                    AS RequestDate
       ,date_format(RequestDateTime,'yyyy-MM')   AS RequestMonth
""" if have_date else "")
print("date column available:", have_date)

concept_flag_cols = ",\n       ".join(f"c_{n} AS {n}" for n in ALL_CONCEPTS)

detail = spark.sql(f"""
    SELECT
        TripRequestId                                       AS OrderId,
        ContractName                                        AS Customer,
        RequesterFacility                                   AS Facility,
        LevelOfServiceDescription                           AS LevelOfService
        {date_cols},
        bucket                                              AS Bucket,
        CASE WHEN bucket IN ('no_documentation','filler_only') THEN 'AT RISK'
             WHEN bucket = 'unrecognized' THEN 'Unknown'
             ELSE 'Supported' END                           AS Status,
        clinical_reasons                                    AS ClinicalReasonCount,
        filler_terms                                        AS FillerCount,
        coalesce(text_len,0)                                AS TextLength,
        CASE WHEN ClinicalData IS NOT NULL
                  AND length(trim(ClinicalData))>0 THEN 'Y' ELSE 'N' END AS HasText,
        {concept_flag_cols},
        substr(regexp_replace(coalesce(ClinicalData,''), '[\\r\\n\\t]', ' '), 1, 300)
                                                            AS WhatTheNurseTyped
    FROM bucketed
""")

print(f"detail rows: {detail.count():,}   columns: {len(detail.columns)}")
display(detail.limit(20))


# COMMAND ----------

# Write the wide detail file (one row per order).
detail.toPandas().to_csv(DETAIL_CSV, index=False)
print("WROTE:", DETAIL_CSV)

# Long format: one row per order + concept that was found. Use this to pivot
# concept counts, or to see which concepts appear together.
concept_rows = spark.sql(f"""
    SELECT OrderId, Customer, Facility, LevelOfService, Bucket, Status, Concept,
           CASE WHEN Concept IN ({','.join(f"'{c}'" for c in GROUP_B)})
                THEN 'B - vague filler' ELSE 'A - clinical reason' END AS ConceptGroup
    FROM (
        SELECT TripRequestId AS OrderId, ContractName AS Customer,
               RequesterFacility AS Facility,
               LevelOfServiceDescription AS LevelOfService,
               bucket AS Bucket,
               CASE WHEN bucket IN ('no_documentation','filler_only') THEN 'AT RISK'
                    WHEN bucket = 'unrecognized' THEN 'Unknown'
                    ELSE 'Supported' END AS Status,
               stack({len(ALL_CONCEPTS)},
                     {','.join(f"'{n}', c_{n}" for n in ALL_CONCEPTS)}) AS (Concept, Hit)
        FROM bucketed
    ) WHERE Hit = 1
""")

print(f"concept rows: {concept_rows.count():,}")
concept_rows.toPandas().to_csv(CONCEPT_CSV, index=False)
print("WROTE:", CONCEPT_CSV)
display(concept_rows.limit(20))


# COMMAND ----------

# -----------------------------------------------------------------------------
# STEP 5 - WRITE EXCEL (5 tabs, plain)
# -----------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BOLD = Font(bold=True)

def write(wb, name, df, title):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=title).font = BOLD
    hdr = 3
    for j, c in enumerate(df.columns, 1):
        ws.cell(row=hdr, column=j, value=str(c)).font = BOLD
    for i, row in enumerate(df.itertuples(index=False), hdr + 1):
        for j, v in enumerate(row, 1):
            if isinstance(v, str) and len(v) > 900:
                v = v[:900] + "..."
            ws.cell(row=i, column=j, value=v)
    for j, c in enumerate(df.columns, 1):
        try:
            w = max(len(str(c)), int(df[c].astype(str).str.len().quantile(0.9)))
        except Exception:
            w = len(str(c))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 12), 80)
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)


wb = Workbook(); wb.remove(wb.active)

# ---------- Tab 1: Definitions ----------
ws = wb.create_sheet("Definitions")
ws.column_dimensions["A"].width = 110
lines = [
    ("Medical Necessity EDA - what we did and why", True),
    ("", False),
    ("WHY", True),
    ("Medicare pays for a non-emergency ambulance only when the patient's condition makes", False),
    ("any other transport unsafe. That reason has to be written down at the time of order.", False),
    ("If it is missing or vague, the claim gets denied 45 days later.", False),
    ("", False),
    ("WHAT WE DID", True),
    ("We took the free-text box the nurse fills in when ordering a transport (the ClinicalData", False),
    ("field) and sorted every order into one of five buckets based on what was written.", False),
    ("The point is to see how much of our documentation actually supports medical necessity.", False),
    ("", False),
    ("SOURCE", True),
    ("Table: prod-sandbox.vivekkumar_patel.temp_tnet_tripmaster (read-only)", False),
    ("Built from prod.silver_transbroker.triprequest + tripleg", False),
    ("Period: calendar year 2025 (confirm)", False),
    ("Customers: Texas Health Resources and MUSC, 52 facilities", False),
    ("", False),
    ("SCOPE", True),
    ("Only non-emergent GROUND transports need this documentation.", False),
    ("Rideshare, helicopter, fixed wing, and emergent trips are excluded from the counts.", False),
    ("", False),
    ("THE TWO WORD GROUPS", True),
    ("Group A - CLINICAL REASON", True),
    ("A specific medical fact that explains why the patient cannot use a wheelchair van or car.", False),
    ("Two tests, both must pass:", False),
    ("   1. Verifiable - could a payer check it against the medical record?", False),
    ("   2. Rules out the cheaper option - does it explain why a van will not work?", False),
    ("Passes: post-CVA hemiparesis, unable to bear weight, requires supine transport, on 3L oxygen", False),
    ("Fails:  patient is elderly, diabetic, going to rehab, doctor ordered ambulance", False),
    ("('Diabetic' is verifiable but does not stop someone riding in a car. Fails test 2.)", False),
    ("Source: Benefit Policy Manual 10.2.1 - medical necessity exists when any other method of", False),
    ("transportation is contraindicated.", False),
    ("", False),
    ("Group B - VAGUE FILLER", True),
    ("Words that describe a general state with no cause and no functional limit.", False),
    ("Test: can you picture what is physically wrong?", False),
    ("   'General weakness' - no. Can they sit? Stand? Walk?", False),
    ("   'Post-CVA hemiparesis, cannot bear weight' - yes.", False),
    ("Not false, just incomplete. Medicare needs the cause AND the functional deficit.", False),
    ("Three types: vague condition (general weakness), risk labels (fall risk, unsteady),", False),
    ("and non-clinical reasons (per protocol, no other transport, family request).", False),
    ("Source: MAC guidance - 'Vague and general information is of little or no value.'", False),
    ("", False),
    ("THE FIVE BUCKETS", True),
    ("clear_reason        Clinical reason present, no filler.            -> supported", False),
    ("reason_plus_filler  Clinical reason present, plus vague words.     -> supported", False),
    ("unrecognized        Text present but none of our words matched.    -> unknown", False),
    ("filler_only         Only vague words, no clinical reason.          -> AT RISK", False),
    ("no_documentation    The box is empty.                             -> AT RISK", False),
    ("", False),
    ("HOW THE BUCKET IS ASSIGNED", True),
    ("Count Group A hits and Group B hits in the text, then:", False),
    ("   box empty                  -> no_documentation", False),
    ("   0 group A, 1 or more B     -> filler_only", False),
    ("   0 group A, 0 group B       -> unrecognized", False),
    ("   1 or more A, 1 or more B   -> reason_plus_filler", False),
    ("   1 or more A, 0 group B     -> clear_reason", False),
    ("", False),
    ("LIMITS - READ THIS", True),
    ("1. This is a keyword search, not AI. It cannot tell that 'patient cannot support trunk", False),
    ("   in vehicle' means 'unable to sit upright'. That gap is what the LLM will fix, and the", False),
    ("   size of the unrecognized bucket is the measurement of it.", False),
    ("2. The word lists came from CMS rules and the Transport.net strategy deck. They have NOT", False),
    ("   been reviewed by Jen Jones or Michelle's team. Treat percentages as directional.", False),
    ("3. 'unrecognized' is a to-do list, not a verdict. Some of those orders are documented fine", False),
    ("   using words we did not look for.", False),
    ("4. Judgment call: 'behavioral' (dementia, flight risk) sits in Group A, but CMS strictly", False),
    ("   wants a physical limitation. This is the most likely reclassification.", False),
    ("", False),
    ("CMS REFERENCES", True),
    ("1. 42 CFR 410.40 - Coverage of ambulance services", False),
    ("   https://www.ecfr.gov/current/title-42/chapter-IV/subchapter-B/part-410/subpart-B/section-410.40", False),
    ("   Says a signed PCS alone does NOT prove medical necessity. The clinical text is what counts.", False),
    ("", False),
    ("2. Medicare Benefit Policy Manual, Chapter 10 - Ambulance Services (Pub 100-02)", False),
    ("   https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/bp102c10.pdf", False),
    ("   10.2.1 Necessity - other transport must be contraindicated", False),
    ("   10.2.3 Bed-confinement - three prongs, all required; not sufficient on its own", False),
    ("   10.2.4 Documentation requirements", False),
    ("", False),
    ("3. 42 CFR 414.605 - Definitions (BLS, ALS1, ALS2, Specialty Care Transport)", False),
    ("   https://www.ecfr.gov/current/title-42/chapter-IV/subchapter-B/part-414/subpart-H/section-414.605", False),
    ("   Note: the federal term is Specialty Care Transport (SCT). Our 'CCT' maps to SCT.", False),
    ("", False),
    ("4. CMS Prior Authorization Operational Guide (RSNAT)", False),
    ("   https://www.cms.gov/research-statistics-data-and-systems/monitoring-programs/", False),
    ("   medicare-ffs-compliance-programs/prior-authorization-initiatives/downloads/", False),
    ("   ambulancepriorauth_operationalguide_123115.pdf", False),
    ("", False),
    ("See the Concepts tab for the specific CMS source behind each individual concept.", False),
    ("", False),
    ("TAB GUIDE", True),
    ("Summary   - how many orders fell in each bucket", False),
    ("Where     - which customers and service levels have the problem", False),
    ("Examples  - real orders from each bucket, so the categories are concrete", False),
    ("Concepts  - which clinical reasons appear, with the CMS source and search terms", False),
    ("", False),
    ("FOR PIVOT TABLES - two CSV files are written alongside this workbook", True),
    ("med_nec_detail.csv - one row per order. Columns: OrderId, Customer, Facility,", False),
    ("   LevelOfService, RequestDate, RequestMonth, Bucket, Status, ClinicalReasonCount,", False),
    ("   FillerCount, TextLength, HasText, then a 0/1 column for each of the 15 concepts,", False),
    ("   then WhatTheNurseTyped (first 300 characters).", False),
    ("   Use for: orders by bucket and customer, at-risk rate by facility, trend by month.", False),
    ("", False),
    ("med_nec_concept_rows.csv - one row per order per concept found (long format).", False),
    ("   Use for: counting concepts, or seeing which concepts appear together.", False),
    ("   A wide file cannot pivot concepts as a single field - this one can.", False),
    ("", False),
    ("To build a pivot: open the CSV in Excel, Insert > PivotTable.", False),
    ("Example - at-risk rate by facility: Rows = Facility, Columns = Status, Values = Count of OrderId.", False),
    ("", False),
    ("Both CSVs contain free-text clinical notes. Keep internal.", False),
]
for i, (t, b) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=t)
    if b:
        c.font = BOLD

# ---------- Tabs 2-5 ----------
write(wb, "Summary", summary,
      f"Bucket counts - non-emergent ground only ({scope_counts.get('in',0):,} of {total_raw:,} orders in scope)")
write(wb, "Where", where,
      "Where the at-risk orders are - by customer and by level of service")
write(wb, "Examples", examples,
      "Real orders from each bucket - review for patient information before sharing")
write(wb, "Concepts", concepts[["Group", "Concept", "Orders", "PctOfOrders",
                               "CMS_Reference", "SearchTerms"]],
      "Which clinical reasons and filler terms appear, with the CMS source for each")

wb.save(OUTPUT_XLSX)
print("WROTE:", OUTPUT_XLSX)
print("tabs:", wb.sheetnames)


# COMMAND ----------

for f in sorted(os.listdir(OUTPUT_DIR)):
    print(f"{f}  ({os.path.getsize(os.path.join(OUTPUT_DIR,f)):,} bytes)")

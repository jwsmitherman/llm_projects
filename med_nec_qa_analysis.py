# Databricks notebook source
# =============================================================================
# Medical Necessity EDA v2 - QUESTION / ANSWER extraction
#
# WHY THIS VERSION EXISTS
# v1 searched keywords across ClinicalData + SpecialNeeds + LosQuestions mashed
# together. That was wrong: LosQuestions holds the ENTIRE questionnaire template
# (avg ~49,000 chars), so every order "matched" oxygen, ventilator, suction etc.
# simply because the form ASKS about them. We measured form templates, not
# patients.
#
# WHAT THIS VERSION DOES
#   PART A - discover the LosQuestions structure (what shape is the payload?)
#   PART B - extract QUESTION -> ANSWER pairs and keep only ANSWERED items
#   PART C - categorize the ANSWERS into clinical concepts
#   PART D - run concepts against ClinicalData ONLY (the nurse's free text)
#   PART E - export Excel with tables + charts
#
# This gives the real input for the AI: for each order, what did the nurse
# actually say, and does it meet medical necessity?
#
# READ-ONLY. Creates no tables. Temp views only.
# =============================================================================


# COMMAND ----------

# Run once, then comment out:
# %pip install openpyxl
# dbutils.library.restartPython()


# COMMAND ----------

# -----------------------------------------------------------------------------
# 0. CONFIG
# -----------------------------------------------------------------------------

TRIPMASTER  = "`prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster"
OUTPUT_DIR  = "/Workspace/Users/josh.smitherman@gmr.net/med_nec/data"
OUTPUT_XLSX = f"{OUTPUT_DIR}/med_nec_qa_summary.xlsx"

# How many payloads to pull to the driver for structure discovery (keep small -
# each payload averages ~49k characters).
STRUCTURE_SAMPLE = 5

import os, json, re
from pyspark.sql import functions as F, types as T

spark.sql("USE CATALOG prod")
os.makedirs(OUTPUT_DIR, exist_ok=True)
print("output ->", OUTPUT_XLSX)


# COMMAND ----------

# =============================================================================
# PART A - DISCOVER THE LosQuestions STRUCTURE
# Run this first. Everything downstream depends on what shape the payload is.
# =============================================================================

samples = (spark.sql(f"""
    SELECT ContractName, LosQuestions
    FROM {TRIPMASTER}
    WHERE LosQuestions IS NOT NULL AND length(trim(LosQuestions)) > 100
    LIMIT {STRUCTURE_SAMPLE}
""").collect())

print(f"pulled {len(samples)} sample payloads\n")
for i, r in enumerate(samples[:2]):
    p = r["LosQuestions"]
    print(f"--- sample {i} | contract={r['ContractName']} | length={len(p):,} ---")
    print(p[:1500])
    print("...\n")


# COMMAND ----------

# Is it JSON? What does the top level look like?
def probe(payload):
    out = {"length": len(payload), "starts_with": payload.lstrip()[:1]}
    try:
        obj = json.loads(payload)
        out["valid_json"] = True
        out["top_type"] = type(obj).__name__
        if isinstance(obj, dict):
            out["top_keys"] = list(obj.keys())[:30]
        elif isinstance(obj, list):
            out["list_len"] = len(obj)
            if obj and isinstance(obj[0], dict):
                out["item_keys"] = list(obj[0].keys())[:30]
    except Exception as e:
        out["valid_json"] = False
        out["error"] = str(e)[:120]
    return out

for i, r in enumerate(samples):
    print(f"sample {i}:", json.dumps(probe(r["LosQuestions"]), indent=2)[:900], "\n")


# COMMAND ----------

# Walk the JSON and report every key path plus an example value. This is the
# schema map - it tells us which keys hold the question text and the answer.
def walk(obj, path="$", depth=0, max_depth=6, acc=None):
    if acc is None:
        acc = {}
    if depth > max_depth:
        return acc
    if isinstance(obj, dict):
        for k, v in obj.items():
            p = f"{path}.{k}"
            if isinstance(v, (dict, list)):
                walk(v, p, depth + 1, max_depth, acc)
            else:
                if p not in acc:
                    acc[p] = str(v)[:80]
    elif isinstance(obj, list):
        for v in obj[:3]:
            walk(v, f"{path}[]", depth + 1, max_depth, acc)
    return acc

try:
    obj = json.loads(samples[0]["LosQuestions"])
    paths = walk(obj)
    print(f"{len(paths)} distinct key paths found:\n")
    for p, v in sorted(paths.items())[:80]:
        print(f"  {p:<60} e.g. {v}")
except Exception as e:
    print("not parseable as JSON:", e)
    print("\nInspect the raw sample above and tell me the format - I will adapt the parser.")


# COMMAND ----------

# Which key names look like QUESTION text vs ANSWER value? Heuristic scan.
QUESTION_HINTS = ("question", "prompt", "label", "text", "title", "name", "caption")
ANSWER_HINTS   = ("answer", "value", "response", "selected", "checked", "result", "chosen", "input")

try:
    q_keys = sorted({p.split(".")[-1] for p in paths
                     if any(h in p.split(".")[-1].lower() for h in QUESTION_HINTS)})
    a_keys = sorted({p.split(".")[-1] for p in paths
                     if any(h in p.split(".")[-1].lower() for h in ANSWER_HINTS)})
    print("candidate QUESTION keys:", q_keys)
    print("candidate ANSWER keys:  ", a_keys)
except Exception as e:
    print("run the previous cell first:", e)


# COMMAND ----------

# =============================================================================
# PART B - EXTRACT QUESTION -> ANSWER PAIRS
#
# The parser is deliberately adaptive: it walks the payload, finds every object
# that has BOTH something question-like and something answer-like, and emits the
# pair. Once Part A tells us the exact schema we can pin it down and drop the
# guessing - but this runs today without waiting.
#
# "Answered" means the answer is present and not empty/false/no/none.
# =============================================================================

NEGATIVE_ANSWERS = {"", "false", "0", "no", "none", "n/a", "na", "null", "unanswered",
                    "not applicable", "not selected", "unchecked"}

QA_SCHEMA = T.ArrayType(T.StructType([
    T.StructField("question", T.StringType()),
    T.StructField("answer",   T.StringType()),
    T.StructField("answered", T.BooleanType()),
]))


def extract_qa(payload):
    """Walk a payload and return [(question, answer, answered)]."""
    if not payload:
        return []
    try:
        obj = json.loads(payload)
    except Exception:
        return []

    out, seen = [], set()

    def qtext(d):
        for k in d:
            kl = k.lower()
            if any(h in kl for h in QUESTION_HINTS):
                v = d[k]
                if isinstance(v, str) and 3 < len(v) < 400:
                    return v.strip()
        return None

    def aval(d):
        for k in d:
            kl = k.lower()
            if any(h in kl for h in ANSWER_HINTS):
                v = d[k]
                if isinstance(v, (str, bool, int, float)) and v is not None:
                    return str(v).strip()
        return None

    def rec(o, depth=0):
        if depth > 12:
            return
        if isinstance(o, dict):
            q, a = qtext(o), aval(o)
            if q is not None and a is not None:
                key = (q[:200], a[:200])
                if key not in seen:
                    seen.add(key)
                    out.append((q[:400], a[:200],
                                a.strip().lower() not in NEGATIVE_ANSWERS))
            for v in o.values():
                rec(v, depth + 1)
        elif isinstance(o, list):
            for v in o:
                rec(v, depth + 1)

    rec(obj)
    return out


extract_qa_udf = F.udf(extract_qa, QA_SCHEMA)

qa = (spark.table(TRIPMASTER.replace("`", ""))
      if False else spark.sql(f"SELECT * FROM {TRIPMASTER}"))

qa_exploded = (qa
    .select("TripRequestId", "ContractName", "LevelOfServiceDescription",
            "ClinicalData", "LosQuestions")
    .withColumn("qa", extract_qa_udf(F.col("LosQuestions")))
    .withColumn("pair", F.explode_outer("qa"))
    .select("TripRequestId", "ContractName", "LevelOfServiceDescription",
            F.col("pair.question").alias("question"),
            F.col("pair.answer").alias("answer"),
            F.col("pair.answered").alias("answered"))
)

qa_exploded.createOrReplaceTempView("qa_pairs")

print("extraction complete. sanity check:")
display(spark.sql("""
    SELECT
        count(*)                                        AS total_pairs,
        count(DISTINCT TripRequestId)                   AS orders,
        sum(CASE WHEN answered THEN 1 ELSE 0 END)       AS answered_pairs,
        count(DISTINCT question)                        AS distinct_questions
    FROM qa_pairs
    WHERE question IS NOT NULL
"""))


# COMMAND ----------

# What questions exist, and how often are they answered YES? This is the
# question catalog we have been missing - the thing nobody had indexed.
display(spark.sql("""
    SELECT
        question,
        count(*)                                  AS times_asked,
        sum(CASE WHEN answered THEN 1 ELSE 0 END) AS times_answered,
        round(100.0*sum(CASE WHEN answered THEN 1 ELSE 0 END)/count(*),1) AS pct_answered
    FROM qa_pairs
    WHERE question IS NOT NULL
    GROUP BY question
    ORDER BY times_answered DESC
    LIMIT 200
"""))


# COMMAND ----------

# What do the ANSWERS actually look like? Free text? Yes/No? Coded values?
# This drives how the AI must read them.
display(spark.sql("""
    SELECT answer, count(*) AS times_used
    FROM qa_pairs
    WHERE answered AND answer IS NOT NULL
    GROUP BY answer
    ORDER BY times_used DESC
    LIMIT 200
"""))


# COMMAND ----------

# =============================================================================
# PART C - CATEGORIZE THE ANSWERED RESPONSES
# Tag each ANSWERED question by the clinical concept its QUESTION is about.
# Because we only count answered items, this reflects real patient need.
# =============================================================================

CONCEPTS = {
    "oxygen":              r"oxygen|\bo2\b|\blpm\b|nasal cannula|non[\s-]*rebreather|\bbipap\b|\bcpap\b",
    "ventilator":          r"ventilat|\bvent\b|intubat|\bett\b",
    "airway_trach":        r"trach|airway",
    "suctioning":          r"suction",
    "cardiac_monitoring":  r"cardiac|telemetry|\bekg\b|\becg\b|monitor.*heart|heart.*monitor",
    "iv_medication":       r"\biv\b|infusion|\bdrip\b|\bppn\b|\btpn\b|titrat|pca pump|medication",
    "bed_confined":        r"bed[\s-]*(bound|confined)|out of bed|get up",
    "cannot_sit":          r"sit|position|supine|stretcher|lie flat|recline",
    "mobility":            r"ambulat|walk|bear weight|transfer|mobil|wheelchair|stand",
    "wound_ostomy":        r"wound|ostomy|drain|ulcer|decubitus|dressing",
    "isolation":           r"isolation|precaution|\bmrsa\b|c\.? ?diff|contagio|infect",
    "bariatric":           r"bariatric|weight|\blbs\b|pound",
    "behavioral":          r"dementia|alzheimer|combative|agitat|confus|behavior|psych|"
                           r"flight risk|elope|sitter|restrain",
    "restraints":          r"restrain",
    "pain":                r"pain",
    "fall_risk":           r"fall",
    "weakness":            r"weak",
}

case_expr = "\n        ".join(
    f"WHEN lower(question) RLIKE '{rx}' THEN '{name}'" for name, rx in CONCEPTS.items()
)

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW qa_categorized AS
SELECT *,
    CASE
        {case_expr}
        ELSE 'other'
    END AS concept
FROM qa_pairs
WHERE question IS NOT NULL
""")

print("answered responses by clinical concept:")
display(spark.sql("""
    SELECT concept,
           count(*)                                  AS times_asked,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS times_answered_yes,
           round(100.0*sum(CASE WHEN answered THEN 1 ELSE 0 END)/count(*),1) AS pct_yes
    FROM qa_categorized
    GROUP BY concept
    ORDER BY times_answered_yes DESC
"""))


# COMMAND ----------

# Per-order rollup: how many clinical needs did this patient actually have?
spark.sql("""
CREATE OR REPLACE TEMP VIEW order_needs AS
SELECT
    TripRequestId,
    ContractName,
    LevelOfServiceDescription,
    count(*)                                        AS questions_asked,
    sum(CASE WHEN answered THEN 1 ELSE 0 END)       AS needs_confirmed,
    concat_ws(', ', collect_set(CASE WHEN answered THEN concept END)) AS confirmed_concepts
FROM qa_categorized
GROUP BY TripRequestId, ContractName, LevelOfServiceDescription
""")

display(spark.sql("""
    SELECT needs_confirmed, count(*) AS orders
    FROM order_needs
    GROUP BY needs_confirmed
    ORDER BY needs_confirmed
"""))


# COMMAND ----------

# =============================================================================
# PART D - CLINICAL FREE TEXT, ON ITS OWN
# The v1 mistake was mixing this with the questionnaire. Here we search ONLY
# ClinicalData - what the nurse actually typed.
# =============================================================================

TEXT_CONCEPTS = {
    "oxygen":            r"oxygen|\bo2\b|\blpm\b|nasal cannula|\bbipap\b|\bcpap\b",
    "ventilator":        r"ventilat|\bvent\b|trach|intubat",
    "suctioning":        r"suction",
    "cardiac":           r"cardiac|telemetry|\bekg\b|\bnstemi\b|\bstemi\b|arrhythm|\bafib\b",
    "iv_medication":     r"\biv\b|infusion|\bdrip\b|heparin|antibiotic|\btpn\b",
    "bed_confined":      r"bed[\s-]*(bound|confined)|unable to get up",
    "cannot_sit":        r"cannot sit|unable to sit|special positioning|supine|must lie|stretcher",
    "mobility_deficit":  r"hemipar|hemipleg|paraly|non[\s-]*ambulat|unable to (bear weight|ambulate|walk)|"
                         r"fracture|amputat|contracture",
    "wound_ostomy":      r"wound|ostomy|ulcer|decubitus|drain",
    "isolation":         r"isolation|\bmrsa\b|c\.? ?diff|precaution",
    "bariatric":         r"bariatric|morbid",
    "behavioral":        r"dementia|alzheimer|combative|agitat|altered mental|flight risk|elope",
    # documentation-quality problems
    "weakness_only":     r"general(iz)?e?d? weakness|generally weak|^\s*weak",
    "fall_risk_only":    r"fall risk|unsteady|deconditio",
    "nonclinical":       r"per protocol|convenience|no other transport|unable to arrange|family request",
}

WEAK = ["weakness_only", "fall_risk_only", "nonclinical"]
STRONG = [c for c in TEXT_CONCEPTS if c not in WEAK]

txt_cols = ",\n        ".join(
    f"CASE WHEN lower(ClinicalData) RLIKE '{rx}' THEN 1 ELSE 0 END AS t_{n}"
    for n, rx in TEXT_CONCEPTS.items())

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW text_tagged AS
SELECT TripRequestId, ContractName, LevelOfServiceDescription,
       ClinicalData, length(ClinicalData) AS clinical_len,
       {txt_cols}
FROM {TRIPMASTER}
""")

s_sum = " + ".join(f"t_{c}" for c in STRONG)
w_sum = " + ".join(f"t_{c}" for c in WEAK)

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW text_scored AS
SELECT *,
    ({s_sum}) AS strong_ct,
    ({w_sum}) AS weak_ct,
    CASE
        WHEN ClinicalData IS NULL OR length(trim(ClinicalData))=0 THEN 'no_documentation'
        WHEN ({s_sum})=0 AND ({w_sum})>0                          THEN 'weak_only'
        WHEN ({s_sum})=0                                          THEN 'unclassified'
        WHEN ({w_sum})>0                                          THEN 'mixed'
        ELSE 'specific'
    END AS documentation_quality
FROM text_tagged
""")

print("documentation quality from ClinicalData ONLY (the corrected number):")
display(spark.sql("""
    SELECT documentation_quality, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM text_scored
    GROUP BY documentation_quality
    ORDER BY orders DESC
"""))


# COMMAND ----------

# =============================================================================
# PART E - EXCEL EXPORT
# =============================================================================

import pandas as pd

def q(sql):
    return spark.sql(sql).toPandas()

overview = q(f"""
    SELECT count(*) AS total_orders,
           count(DISTINCT ContractName) AS contracts,
           sum(CASE WHEN ClinicalData IS NOT NULL AND length(trim(ClinicalData))>0 THEN 1 ELSE 0 END) AS has_clinical_text,
           sum(CASE WHEN LosQuestions IS NOT NULL AND length(trim(LosQuestions))>0 THEN 1 ELSE 0 END) AS has_questionnaire
    FROM {TRIPMASTER}
""").T.reset_index()
overview.columns = ["Metric", "Value"]

qa_health = q("""
    SELECT count(*) AS total_pairs,
           count(DISTINCT TripRequestId) AS orders_with_pairs,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS answered_pairs,
           count(DISTINCT question) AS distinct_questions
    FROM qa_pairs WHERE question IS NOT NULL
""").T.reset_index()
qa_health.columns = ["Metric", "Value"]

question_catalog = q("""
    SELECT question, count(*) AS times_asked,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS times_answered,
           round(100.0*sum(CASE WHEN answered THEN 1 ELSE 0 END)/count(*),1) AS pct_answered
    FROM qa_pairs WHERE question IS NOT NULL
    GROUP BY question ORDER BY times_answered DESC LIMIT 500
""")

answer_values = q("""
    SELECT answer, count(*) AS times_used
    FROM qa_pairs WHERE answered AND answer IS NOT NULL
    GROUP BY answer ORDER BY times_used DESC LIMIT 300
""")

concept_responses = q("""
    SELECT concept, count(*) AS times_asked,
           sum(CASE WHEN answered THEN 1 ELSE 0 END) AS answered_yes,
           round(100.0*sum(CASE WHEN answered THEN 1 ELSE 0 END)/count(*),1) AS pct_yes
    FROM qa_categorized GROUP BY concept ORDER BY answered_yes DESC
""")

needs_dist = q("""
    SELECT needs_confirmed, count(*) AS orders
    FROM order_needs GROUP BY needs_confirmed ORDER BY needs_confirmed
""")

needs_by_los = q("""
    SELECT LevelOfServiceDescription,
           count(*) AS orders,
           round(avg(needs_confirmed),2) AS avg_needs_confirmed
    FROM order_needs
    GROUP BY LevelOfServiceDescription
    ORDER BY orders DESC LIMIT 50
""")

doc_quality = q("""
    SELECT documentation_quality, count(*) AS orders,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM text_scored GROUP BY documentation_quality ORDER BY orders DESC
""")

doc_by_contract = q("""
    SELECT ContractName, documentation_quality, count(*) AS orders
    FROM text_scored GROUP BY ContractName, documentation_quality
    ORDER BY ContractName, orders DESC
""")

doc_by_los = q("""
    SELECT LevelOfServiceDescription, documentation_quality, count(*) AS orders
    FROM text_scored GROUP BY LevelOfServiceDescription, documentation_quality
    ORDER BY orders DESC LIMIT 100
""")

text_concepts = pd.concat([
    spark.sql(f"""SELECT '{n}' AS concept, sum(t_{n}) AS orders,
                  round(100.0*sum(t_{n})/count(*),2) AS pct_of_orders
                  FROM text_scored""").toPandas()
    for n in TEXT_CONCEPTS
], ignore_index=True).sort_values("orders", ascending=False)
text_concepts["type"] = text_concepts["concept"].apply(
    lambda c: "quality problem" if c in WEAK else "supports necessity")

sample_specific = q("""
    SELECT ContractName, LevelOfServiceDescription, documentation_quality, ClinicalData
    FROM text_scored WHERE documentation_quality='specific' LIMIT 200
""")
sample_weak = q("""
    SELECT ContractName, LevelOfServiceDescription, documentation_quality, ClinicalData
    FROM text_scored WHERE documentation_quality IN ('weak_only','unclassified') LIMIT 200
""")
sample_qa = q("""
    SELECT TripRequestId, ContractName, question, answer, answered
    FROM qa_pairs WHERE question IS NOT NULL AND answered LIMIT 500
""")

print("tables built")


# COMMAND ----------

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY     = Font(name="Arial", size=10)
TITLE    = Font(name="Arial", bold=True, size=14, color="1F4E79")


def write_sheet(wb, name, df, title=None, note=None):
    ws = wb.create_sheet(name[:31])
    r = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE
        r = 2
    if note:
        c = ws.cell(row=r, column=1, value=note)
        c.font = Font(name="Arial", size=9, italic=True, color="666666")
        r += 1
    r += 1
    hdr = r
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=hdr, column=j, value=str(col))
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(df.itertuples(index=False), start=hdr + 1):
        for j, val in enumerate(row, start=1):
            if isinstance(val, str) and len(val) > 800:
                val = val[:800] + "..."
            ws.cell(row=i, column=j, value=val).font = BODY
    for j, col in enumerate(df.columns, start=1):
        try:
            w = max(len(str(col)), int(df[col].astype(str).str.len().quantile(0.9)))
        except Exception:
            w = len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 3, 12), 70)
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    return ws, hdr, hdr + len(df)


def add_bar(ws, title, cat_col, val_col, hdr, last, anchor):
    ch = BarChart(); ch.type, ch.style, ch.title = "col", 10, title
    ch.add_data(Reference(ws, min_col=val_col, min_row=hdr, max_row=last), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=hdr + 1, max_row=last))
    ch.height, ch.width, ch.legend = 8, 18, None
    ws.add_chart(ch, anchor)


def add_pie(ws, title, cat_col, val_col, hdr, last, anchor):
    ch = PieChart(); ch.title = title
    ch.add_data(Reference(ws, min_col=val_col, min_row=hdr, max_row=last), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=hdr + 1, max_row=last))
    ch.height, ch.width = 9, 13
    ws.add_chart(ch, anchor)


wb = Workbook(); wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 118
for i, (t, f) in enumerate([
    ("Medical Necessity - Question / Answer Analysis", TITLE),
    ("", BODY),
    ("Source: `prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster (read-only, no tables created).", BODY),
    ("Scope: 2025 orders, Texas Health Resources + MUSC.", BODY),
    ("", BODY),
    ("WHAT CHANGED FROM THE FIRST VERSION", Font(name="Arial", bold=True, size=11)),
    ("v1 searched keywords across the clinical text AND the questionnaire mashed together.", BODY),
    ("The questionnaire is the full form template (~49,000 characters), listing every possible", BODY),
    ("question whether answered or not. So every order 'matched' oxygen, ventilator, suction etc.", BODY),
    ("simply because the form asks about them. v1 measured form templates, not patients.", BODY),
    ("", BODY),
    ("v2 does two separate things:", Font(name="Arial", bold=True, size=11)),
    ("  1. Parses the questionnaire into QUESTION -> ANSWER pairs and keeps only ANSWERED items.", BODY),
    ("  2. Searches the nurse's free text (ClinicalData) on its own.", BODY),
    ("", BODY),
    ("SHEET GUIDE", Font(name="Arial", bold=True, size=11)),
    ("  Question Catalog  - every question asked, and how often it is answered yes", BODY),
    ("  Answer Values     - what the answers look like (yes/no, coded, free text)", BODY),
    ("  Concept Responses - answered questions grouped by clinical concept", BODY),
    ("  Needs Per Order   - how many clinical needs each patient actually had confirmed", BODY),
    ("  Doc Quality       - quality of the nurse's free text, corrected", BODY),
    ("  Text Concepts     - clinical concepts found in the free text only", BODY),
    ("  Sample QA / Sample Specific / Sample Weak - raw examples for review", BODY),
    ("", BODY),
    ("CAVEAT: concept word lists come from CMS/GMR guidance and are NOT yet validated by", BODY),
    ("Jen Jones / Michelle's team. Treat percentages as directional until they are reviewed.", BODY),
], start=1):
    ws.cell(row=i, column=1, value=t).font = f

write_sheet(wb, "Overview", overview, "Population and field completeness")
write_sheet(wb, "QA Extraction Health", qa_health,
            "Question/answer extraction health",
            "If answered_pairs is 0 the parser did not match the payload schema - see Part A output.")

ws, h, l = write_sheet(wb, "Question Catalog", question_catalog,
                       "Question catalog - every question and how often answered yes",
                       "The index of clinical questions nobody had. Top 500 by answered volume.")
add_bar(ws, "Most-answered questions", 1, 3, h, min(l, h + 25), "F4")

write_sheet(wb, "Answer Values", answer_values,
            "What the answers look like",
            "Drives how the AI must read responses: yes/no flags, coded values, or free text.")

ws, h, l = write_sheet(wb, "Concept Responses", concept_responses,
                       "Answered responses by clinical concept",
                       "Counts only questions the nurse actually answered - real patient needs.")
add_bar(ws, "Confirmed needs by concept", 1, 3, h, l, "F4")

ws, h, l = write_sheet(wb, "Needs Per Order", needs_dist,
                       "How many clinical needs were confirmed per order",
                       "Orders with zero confirmed needs are the medical-necessity risk cases.")
add_bar(ws, "Orders by number of confirmed needs", 1, 2, h, l, "E4")

write_sheet(wb, "Needs By LOS", needs_by_los,
            "Average confirmed needs by level of service",
            "Higher levels of service should show more confirmed needs. If not, that is the gap.")

ws, h, l = write_sheet(wb, "Doc Quality", doc_quality,
                       "Documentation quality - free text only (CORRECTED)",
                       "Based solely on what the nurse typed in ClinicalData.")
add_pie(ws, "Documentation quality", 1, 2, h, l, "F4")

write_sheet(wb, "Doc Quality x Contract", doc_by_contract, "Documentation quality by contract")
write_sheet(wb, "Doc Quality x LOS", doc_by_los, "Documentation quality by level of service")

ws, h, l = write_sheet(wb, "Text Concepts", text_concepts,
                       "Clinical concepts in the nurse's free text only")
add_bar(ws, "Concepts in free text", 1, 2, h, l, "F4")

write_sheet(wb, "Sample QA", sample_qa,
            "Sample answered question/answer pairs",
            "The actual AI input. Review for PHI before sharing outside the team.")
write_sheet(wb, "Sample Specific", sample_specific, "Sample free text - specific documentation")
write_sheet(wb, "Sample Weak", sample_weak, "Sample free text - weak or unclassified")

wb.save(OUTPUT_XLSX)
print("WROTE:", OUTPUT_XLSX)
print("sheets:", wb.sheetnames)


# COMMAND ----------

import os
for f in sorted(os.listdir(OUTPUT_DIR)):
    p = os.path.join(OUTPUT_DIR, f)
    print(f"{f}  ({os.path.getsize(p):,} bytes)")

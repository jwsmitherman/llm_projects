# Databricks notebook source
# =============================================================================
# Medical Necessity EDA - clinical response categorization + Excel export
#
# READ-ONLY. This notebook creates NO tables. It reads the prebuilt TripMaster
# extract, categorizes the clinical responses, and writes a single formatted
# Excel workbook (tables + charts) to the med_nec/data workspace folder.
#
# CONFIRMED SOURCE (Vivekkumar's TripMaster_v2, 7/17/2026):
#   `prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster
#   built from prod.silver_transbroker.triprequest + .tripleg
#   filtered YEAR(RequestDateTime)=2025, ContractId IN (29,555,326,229)
#     229          = Texas Health Resources (THR)
#     29,555,326   = MUSC
#
# CLINICAL FIELDS:
#   ClinicalData       <- nurse free-text justification   *** the AI input ***
#   LosQuestions       <- structured question/answer payload per contract
#   LevelOfService     / LosCategory (ServiceType)
#   SpecialNeeds       <- equipment / needs
#   LOSOverride        <- BOOLEAN: user overrode system-derived LOS
#   LOSOverrideReason
#
# ORDER: run top to bottom. Last cell writes the workbook.
# =============================================================================


# COMMAND ----------

# Run once, then comment out:
# %pip install openpyxl
# dbutils.library.restartPython()


# COMMAND ----------

# -----------------------------------------------------------------------------
# 0. CONFIG - read-only
# -----------------------------------------------------------------------------

TRIPMASTER = "`prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster"

# Workspace folder you created: Users > josh.smitherman@gmr.net > med_nec > data
OUTPUT_DIR  = "/Workspace/Users/josh.smitherman@gmr.net/med_nec/data"
OUTPUT_XLSX = f"{OUTPUT_DIR}/med_nec_eda_summary.xlsx"

# Cap rows pulled to the driver for row-level work. Raise once you know volumes.
SAMPLE_LIMIT = 200000

import os
from pyspark.sql import functions as F

spark.sql("USE CATALOG prod")
os.makedirs(OUTPUT_DIR, exist_ok=True)
print("output ->", OUTPUT_XLSX)


# COMMAND ----------

# -----------------------------------------------------------------------------
# 1. CLINICAL CONCEPT TAXONOMY
# Categorizes the clinical responses (ClinicalData free text + SpecialNeeds +
# LosQuestions payload) into the concepts that actually drive medical necessity
# and level of service. Deterministic and auditable - no model needed, so it
# runs over the full population rather than a sample.
#
# Each entry: category -> regex. Validate the term lists with Jen Jones /
# Michelle's team before treating any number as final.
# -----------------------------------------------------------------------------

CONCEPTS = {
    # --- concepts that SUPPORT medical necessity -----------------------------
    "bed_confined":          r"bed[\s-]*(bound|confined)|cannot get out of bed|unable to get up",
    "cannot_sit_position":   r"special positioning|cannot sit|unable to sit|must lie|supine|stretcher only|flat",
    "mobility_deficit":      r"hemipar|hemipleg|paraly|quadripleg|parapleg|non[\s-]*ambulat|"
                             r"unable to (bear weight|ambulate|walk|stand)|fracture|amputat|contracture",
    "oxygen":                r"\bo2\b|oxygen|\blpm\b|nasal cannula|non[\s-]*rebreather|\bbipap\b|\bcpap\b",
    "cardiac_monitoring":    r"cardiac monitor|telemetry|\bekg\b|\becg\b|\bnstemi\b|\bstemi\b|arrhythmia|"
                             r"\bafib\b|chest pain",
    "iv_medication":         r"\biv\b|infusion|\bdrip\b|heparin|antibiotic|\bppn\b|\btpn\b|"
                             r"vasoactive|titrat|pca pump",
    "ventilator":            r"ventilat|\bvent\b|trach|intubat|\bett\b",
    "suctioning":            r"suction",
    "airway_other":          r"tracheostomy|airway",
    "wound_ostomy":          r"wound vac|\bostomy\b|colostomy|pressure ulcer|\bdecubitus\b|drain",
    "isolation":             r"isolation|\bmrsa\b|\bcdiff\b|c\.? diff|contact precaution|droplet",
    "bariatric":             r"bariatric|over ?weight limit|\bmorbid",
    "behavioral":            r"dementia|alzheimer|combative|agitat|confus|altered mental|"
                             r"flight risk|elope|psych|behavioral|sitter",
    "restraints":            r"restrain",
    "monitoring_general":    r"monitor(ing)?\b|vital signs|assessment en ?route",

    # --- documentation QUALITY problems (do NOT establish necessity alone) ----
    "weakness_only":         r"general(iz)?e?d? weakness|^\s*weak(ness)?\b|generally weak",
    "fall_risk_only":        r"fall risk|unsteady|deconditio",
    "nonclinical_reason":    r"per protocol|convenience|no other transport|unable to arrange|"
                             r"family (unable|request)|no ride|per (facility|physician) request",
}

# Concepts that, standing alone, are insufficient documentation.
WEAK_CONCEPTS = ["weakness_only", "fall_risk_only", "nonclinical_reason"]
# Concepts that support necessity.
STRONG_CONCEPTS = [c for c in CONCEPTS if c not in WEAK_CONCEPTS]

print(f"{len(CONCEPTS)} concepts defined ({len(STRONG_CONCEPTS)} supporting, "
      f"{len(WEAK_CONCEPTS)} quality-problem)")


# COMMAND ----------

# -----------------------------------------------------------------------------
# 2. BUILD THE CATEGORIZED VIEW (in-memory only - nothing written to the catalog)
# Concatenates the three clinical response fields into one searchable blob, then
# flags each concept. Uses a temp VIEW, which requires no CREATE TABLE rights.
# -----------------------------------------------------------------------------

concept_cols = ",\n        ".join(
    f"CASE WHEN lower(response_blob) RLIKE '{rx}' THEN 1 ELSE 0 END AS c_{name}"
    for name, rx in CONCEPTS.items()
)

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW mednec_base AS
SELECT
    TripRequestId, TripLegId, ContractId, ContractName,
    RequesterFacility, FacilityType,
    RequestType, LevelOfService, LevelOfServiceDescription, ServiceType,
    LOSOverride, LOSOverrideReason,
    SpecialNeeds, ClinicalData, LosQuestions,
    length(ClinicalData) AS clinical_len,
    lower(concat_ws(' ',
        coalesce(ClinicalData,''),
        coalesce(SpecialNeeds,''),
        coalesce(LosQuestions,''))) AS response_blob
FROM {TRIPMASTER}
""")

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW mednec_tagged AS
SELECT *,
        {concept_cols}
FROM mednec_base
""")

# Roll-ups: does the record have ANY supporting concept, and any weak-only signal?
strong_sum = " + ".join(f"c_{c}" for c in STRONG_CONCEPTS)
weak_sum   = " + ".join(f"c_{c}" for c in WEAK_CONCEPTS)

spark.sql(f"""
CREATE OR REPLACE TEMP VIEW mednec_scored AS
SELECT *,
    ({strong_sum}) AS strong_concept_count,
    ({weak_sum})   AS weak_concept_count,
    CASE
        WHEN ClinicalData IS NULL OR length(trim(ClinicalData)) = 0 THEN 'no_documentation'
        WHEN ({strong_sum}) = 0 AND ({weak_sum}) > 0               THEN 'weak_only'
        WHEN ({strong_sum}) = 0                                     THEN 'unclassified'
        WHEN ({weak_sum}) > 0                                       THEN 'mixed'
        ELSE 'specific'
    END AS documentation_quality
FROM mednec_tagged
""")

print("views ready: mednec_base, mednec_tagged, mednec_scored")
display(spark.sql("SELECT count(*) AS total_trips FROM mednec_scored"))


# COMMAND ----------

# -----------------------------------------------------------------------------
# 3. BUILD SUMMARY TABLES (collected to pandas for the workbook)
# -----------------------------------------------------------------------------

import pandas as pd

def q(sql):
    return spark.sql(sql).toPandas()


# 3a. Overview / field population
overview = q(f"""
    SELECT
        count(*)                                                              AS total_trips,
        count(DISTINCT ContractId)                                            AS contracts,
        count(DISTINCT RequesterFacility)                                     AS facilities,
        sum(CASE WHEN ClinicalData IS NOT NULL AND length(trim(ClinicalData))>0 THEN 1 ELSE 0 END) AS has_clinical_text,
        sum(CASE WHEN LosQuestions IS NOT NULL AND length(trim(LosQuestions))>0 THEN 1 ELSE 0 END) AS has_los_questions,
        sum(CASE WHEN SpecialNeeds IS NOT NULL AND length(trim(SpecialNeeds))>0 THEN 1 ELSE 0 END) AS has_special_needs,
        sum(CASE WHEN LOSOverride THEN 1 ELSE 0 END)                          AS los_overrides,
        round(avg(clinical_len),1)                                            AS avg_clinical_len,
        max(clinical_len)                                                     AS max_clinical_len
    FROM mednec_scored
""").T.reset_index()
overview.columns = ["Metric", "Value"]

# 3b. Documentation quality mix - the headline
doc_quality = q("""
    SELECT documentation_quality,
           count(*) AS trips,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM mednec_scored
    GROUP BY documentation_quality
    ORDER BY trips DESC
""")

# 3c. Documentation quality by contract
doc_by_contract = q("""
    SELECT ContractName, documentation_quality, count(*) AS trips
    FROM mednec_scored
    GROUP BY ContractName, documentation_quality
    ORDER BY ContractName, trips DESC
""")

# 3d. Concept frequency - the categorization of responses
concept_rows = []
for name in CONCEPTS:
    concept_rows.append(
        spark.sql(f"""
            SELECT '{name}' AS concept,
                   sum(c_{name}) AS trips,
                   round(100.0*sum(c_{name})/count(*),2) AS pct_of_trips
            FROM mednec_scored
        """).toPandas()
    )
concepts_df = pd.concat(concept_rows, ignore_index=True).sort_values("trips", ascending=False)
concepts_df["type"] = concepts_df["concept"].apply(
    lambda c: "quality problem" if c in WEAK_CONCEPTS else "supports necessity")

# 3e. Level of service mix
los_mix = q("""
    SELECT LevelOfService, LevelOfServiceDescription, ServiceType,
           count(*) AS trips,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM mednec_scored
    GROUP BY LevelOfService, LevelOfServiceDescription, ServiceType
    ORDER BY trips DESC
""")

# 3f. Documentation quality BY level of service (BLS vs wheelchair gray area)
doc_by_los = q("""
    SELECT LevelOfServiceDescription, documentation_quality, count(*) AS trips
    FROM mednec_scored
    GROUP BY LevelOfServiceDescription, documentation_quality
    ORDER BY LevelOfServiceDescription, trips DESC
""")

# 3g. Override analysis
override_summary = q("""
    SELECT CASE WHEN LOSOverride THEN 'overridden' ELSE 'system-derived' END AS los_source,
           count(*) AS trips,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct
    FROM mednec_scored
    GROUP BY 1 ORDER BY trips DESC
""")

override_reasons = q("""
    SELECT LOSOverrideReason, count(*) AS trips
    FROM mednec_scored
    WHERE LOSOverrideReason IS NOT NULL AND length(trim(LOSOverrideReason))>0
    GROUP BY LOSOverrideReason ORDER BY trips DESC LIMIT 100
""")

# 3h. RISK MATRIX - override x documentation quality. Top-left is the target.
risk_matrix = q("""
    SELECT CASE WHEN LOSOverride THEN 'overridden' ELSE 'system-derived' END AS los_source,
           documentation_quality,
           count(*) AS trips,
           round(100.0*count(*)/sum(count(*)) OVER (),1) AS pct_of_all
    FROM mednec_scored
    GROUP BY 1,2 ORDER BY trips DESC
""")

# 3i. Facility breakdown
by_facility = q("""
    SELECT RequesterFacility, FacilityType, ContractName,
           count(*) AS trips,
           sum(CASE WHEN documentation_quality IN ('weak_only','no_documentation') THEN 1 ELSE 0 END) AS weak_trips,
           round(100.0*sum(CASE WHEN documentation_quality IN ('weak_only','no_documentation') THEN 1 ELSE 0 END)
                 /count(*),1) AS pct_weak
    FROM mednec_scored
    GROUP BY RequesterFacility, FacilityType, ContractName
    ORDER BY trips DESC LIMIT 200
""")

# 3j. Sample narratives for review / few-shot seeding (no patient identifiers)
samples_weak = q("""
    SELECT ContractName, LevelOfServiceDescription, LOSOverride,
           documentation_quality, ClinicalData
    FROM mednec_scored
    WHERE documentation_quality IN ('weak_only','no_documentation')
    LIMIT 300
""")

samples_specific = q("""
    SELECT ContractName, LevelOfServiceDescription, LOSOverride,
           documentation_quality, ClinicalData
    FROM mednec_scored
    WHERE documentation_quality = 'specific'
    LIMIT 300
""")

# 3k. LosQuestions payload shape by contract
losq_shape = q("""
    SELECT ContractName,
           count(*) AS trips,
           sum(CASE WHEN LosQuestions IS NOT NULL AND length(trim(LosQuestions))>0 THEN 1 ELSE 0 END) AS has_payload,
           round(avg(length(LosQuestions)),0) AS avg_len,
           max(length(LosQuestions))          AS max_len
    FROM mednec_scored
    GROUP BY ContractName ORDER BY trips DESC
""")

losq_samples = q("""
    SELECT ContractName, LosQuestions
    FROM mednec_scored
    WHERE LosQuestions IS NOT NULL AND length(trim(LosQuestions))>0
    LIMIT 50
""")

print("summary tables built")
display(doc_quality)


# COMMAND ----------

display(concepts_df)


# COMMAND ----------

display(risk_matrix)


# COMMAND ----------

# -----------------------------------------------------------------------------
# 4. WRITE THE EXCEL WORKBOOK (tables + charts) to the data folder
# -----------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY     = Font(name="Arial", size=10)
TITLE    = Font(name="Arial", bold=True, size=14, color="1F4E79")


def write_sheet(wb, name, df, title=None, note=None, pct_cols=()):
    """Write a dataframe as a formatted table. Returns (ws, first_data_row, last_row)."""
    ws = wb.create_sheet(name[:31])
    r = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE
        r = 2
    if note:
        c = ws.cell(row=r, column=1, value=note)
        c.font = Font(name="Arial", size=9, italic=True, color="666666")
        r += 1
    r += 1
    header_row = r

    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=j, value=str(col))
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for j, val in enumerate(row, start=1):
            if isinstance(val, str) and len(val) > 800:
                val = val[:800] + "..."
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY
            if df.columns[j - 1] in pct_cols:
                c.number_format = "0.0"

    # column widths
    for j, col in enumerate(df.columns, start=1):
        try:
            width = max(len(str(col)), int(df[col].astype(str).str.len().quantile(0.9)))
        except Exception:
            width = len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(width + 3, 12), 70)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws, header_row, header_row + len(df)


def add_bar(ws, title, cat_col, val_col, hdr_row, last_row, anchor, y_title="Trips"):
    ch = BarChart()
    ch.type, ch.style, ch.title = "col", 10, title
    ch.y_axis.title, ch.x_axis.title = y_title, None
    data = Reference(ws, min_col=val_col, min_row=hdr_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=hdr_row + 1, max_row=last_row)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height, ch.width = 8, 18
    ch.legend = None
    ws.add_chart(ch, anchor)


def add_pie(ws, title, cat_col, val_col, hdr_row, last_row, anchor):
    ch = PieChart()
    ch.title = title
    data = Reference(ws, min_col=val_col, min_row=hdr_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=hdr_row + 1, max_row=last_row)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height, ch.width = 9, 13
    ws.add_chart(ch, anchor)


wb = Workbook()
wb.remove(wb.active)

# --- README ---
ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 120
lines = [
    ("Medical Necessity EDA - Summary", TITLE),
    ("", BODY),
    ("Source: `prod-sandbox`.vivekkumar_patel.temp_tnet_tripmaster (read-only).", BODY),
    ("Built from prod.silver_transbroker.triprequest + tripleg (Vivekkumar's TripMaster_v2).", BODY),
    ("Scope: RequestDateTime year 2025; ContractId 229 = Texas Health Resources, 29/555/326 = MUSC.", BODY),
    ("", BODY),
    ("Clinical response fields categorized: ClinicalData (nurse free text), SpecialNeeds, LosQuestions.", BODY),
    ("The three are concatenated into one blob, then tagged against a clinical concept taxonomy.", BODY),
    ("", BODY),
    ("documentation_quality definitions:", Font(name="Arial", bold=True, size=11)),
    ("  specific         - at least one concept supporting medical necessity, no weak-only language", BODY),
    ("  mixed            - supporting concept(s) present alongside vague language", BODY),
    ("  weak_only        - ONLY vague language (general weakness, fall risk, per protocol)", BODY),
    ("  unclassified     - text present but no concept matched (review these; may need new terms)", BODY),
    ("  no_documentation - ClinicalData empty", BODY),
    ("", BODY),
    ("CAVEAT: the concept term lists are drawn from CMS/GMR guidance, not yet validated against", BODY),
    ("these narratives. Review with Jen Jones / Michelle's team before treating any figure as final.", BODY),
    ("'weak_only' indicates documentation risk, NOT a confirmed denial.", BODY),
]
for i, (txt, font) in enumerate(lines, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = font

# --- sheets + charts ---
write_sheet(wb, "Overview", overview, "Overview - population and field completeness")

ws, h, l = write_sheet(wb, "Doc Quality", doc_quality,
                       "Documentation quality mix",
                       "Share of trips by quality of the clinical justification.",
                       pct_cols=("pct",))
add_pie(ws, "Documentation quality", 1, 2, h, l, "F4")

ws, h, l = write_sheet(wb, "Doc Quality x Contract", doc_by_contract,
                       "Documentation quality by contract")
add_bar(ws, "Trips by contract and quality", 2, 3, h, l, "F4")

ws, h, l = write_sheet(wb, "Concepts", concepts_df,
                       "Clinical concept frequency - categorized responses",
                       "How often each clinical concept appears across ClinicalData + SpecialNeeds + LosQuestions.",
                       pct_cols=("pct_of_trips",))
add_bar(ws, "Concept frequency", 1, 2, h, l, "F4")

ws, h, l = write_sheet(wb, "LOS Mix", los_mix, "Level of service mix", pct_cols=("pct",))
add_bar(ws, "Trips by level of service", 2, 4, h, l, "H4")

write_sheet(wb, "Doc Quality x LOS", doc_by_los,
            "Documentation quality by level of service",
            "BLS vs wheelchair is the denial-prone gray area.")

ws, h, l = write_sheet(wb, "Overrides", override_summary,
                       "Level of service overrides", pct_cols=("pct",))
add_pie(ws, "System-derived vs overridden", 1, 2, h, l, "F4")

write_sheet(wb, "Override Reasons", override_reasons, "Override reasons, ranked")

ws, h, l = write_sheet(wb, "Risk Matrix", risk_matrix,
                       "Risk matrix - override x documentation quality",
                       "Overridden + weak documentation = highest denial risk and the target for the front-end nudge.",
                       pct_cols=("pct_of_all",))
add_bar(ws, "Trips by risk cell", 2, 3, h, l, "G4")

write_sheet(wb, "By Facility", by_facility,
            "Facility breakdown (top 200 by volume)", pct_cols=("pct_weak",))
write_sheet(wb, "LosQuestions Shape", losq_shape, "LosQuestions payload size by contract")
write_sheet(wb, "Sample Weak", samples_weak,
            "Sample narratives - weak or missing documentation",
            "Few-shot / golden-set candidates. Review for PHI before sharing outside the team.")
write_sheet(wb, "Sample Specific", samples_specific,
            "Sample narratives - specific documentation",
            "Positive examples of compliant justification.")
write_sheet(wb, "Sample LosQuestions", losq_samples,
            "Sample LosQuestions payloads",
            "Inspect structure before parsing - format varies by contract.")

wb.save(OUTPUT_XLSX)
print("WROTE:", OUTPUT_XLSX)
print("sheets:", wb.sheetnames)


# COMMAND ----------

# Confirm the file landed in the data folder.
import os
for f in sorted(os.listdir(OUTPUT_DIR)):
    p = os.path.join(OUTPUT_DIR, f)
    print(f"{f}  ({os.path.getsize(p):,} bytes)")


# COMMAND ----------

# -----------------------------------------------------------------------------
# 5. OPTIONAL - LLM categorization of the narratives that the rules missed.
# Run only against 'unclassified' records to extend the taxonomy: whatever the
# model names repeatedly becomes a new regex in CONCEPTS above.
# Still read-only; results stay in the notebook.
# -----------------------------------------------------------------------------

# %pip install -q openai
# dbutils.library.restartPython()

# import json
# from openai import OpenAI
# client = OpenAI(api_key="sk-REPLACE_ME")   # move to a secret scope after testing
#
# rows = spark.sql("""
#     SELECT ClinicalData FROM mednec_scored
#     WHERE documentation_quality = 'unclassified'
#     LIMIT 50
# """).collect()
#
# SYS = ("Categorize this ambulance transport clinical justification. Respond ONLY as JSON: "
#        "{\"concept\": <short snake_case clinical concept>, "
#        "\"supports_medical_necessity\": <true|false>, \"why\": <=1 sentence}. "
#        "This is billing/reimbursement documentation review, not clinical advice.")
#
# out = []
# for r in rows:
#     resp = client.chat.completions.create(
#         model="gpt-4o-mini", temperature=0.0,
#         response_format={"type": "json_object"},
#         messages=[{"role": "system", "content": SYS},
#                   {"role": "user", "content": r["ClinicalData"][:1500]}])
#     out.append({"text": r["ClinicalData"][:200], **json.loads(resp.choices[0].message.content)})
#
# display(spark.createDataFrame(out))
